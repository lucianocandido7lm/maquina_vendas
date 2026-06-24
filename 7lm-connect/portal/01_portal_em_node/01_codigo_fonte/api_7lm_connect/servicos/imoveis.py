"""
Servicos do modulo de imoveis.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Iterable
import unicodedata
from urllib.parse import quote_plus
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile
from xml.sax.saxutils import escape as xml_escape

from fastapi import HTTPException, UploadFile
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image as PlatypusImage, KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    load_workbook = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None
    get_column_letter = None

from configuracoes import DIRETORIO_PUBLICO

from repositorios.imoveis import (
    atualizar_imovel as atualizar_imovel_repositorio,
    buscar_imovel_por_id,
    buscar_imovel_por_chave_importacao,
    criar_imovel as criar_imovel_repositorio,
    inserir_evolucao_obra,
    inserir_midia_imovel,
    listar_evolucao_obra,
    listar_imoveis_candidatos_mesmo_endereco,
    obter_ultima_evolucao_obra,
    sincronizar_percentual_conclusao_obra_atual,
)
from servicos.simulador import serializar_reserva
from validacoes.imoveis import (
    calcular_percentual_captacao_ate_entrega_por_valor,
    calcular_valor_garantido_padrao,
    calcular_valor_captacao_ate_entrega_planejado,
    EXTENSOES_PLANILHA_IMPORTACAO,
    calcular_data_entrega_dinamica,
    calcular_meses_ate_entrega,
    calcular_meses_ate_entrega_dinamica,
    classificar_arquivo_de_midia,
    linha_planilha_vazia,
    mapear_colunas_importacao,
    normalizar_nome_para_arquivo,
    normalizar_payload_imovel,
    normalizar_decimal,
    normalizar_decimal_intervalo,
    normalizar_inteiro,
    obter_valor_coluna_planilha,
    valor_planilha_como_texto,
)


def _decimal_para_float(valor: Decimal | None) -> float | None:
    if valor is None:
        return None
    return float(valor)


def _data_para_iso(valor: Any) -> str | None:
    if not valor:
        return None
    if hasattr(valor, "isoformat"):
        return valor.isoformat()
    texto = str(valor).strip()
    return texto or None


PADRAO_BLOCO = re.compile(r"\bbloco\s+([a-z0-9-]+)", re.IGNORECASE)
PADRAO_UNIDADE_TITULO = re.compile(r"\b([a-z0-9-]+)\s*-\s*bloco\b", re.IGNORECASE)
PADRAO_UNIDADE_ENDERECO = re.compile(r"\bunidade\s+([a-z0-9-]+)", re.IGNORECASE)
PADRAO_PAVIMENTO_ENDERECO = re.compile(r"\bpavimento\s+([^,\n\r.]+)", re.IGNORECASE)
PADRAO_CARACTERES_XML_INVALIDOS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _texto_limpo(valor: Any) -> str:
    return str(valor or "").strip()


def _sanitizar_texto_exportacao(valor: Any, *, preservar_quebras: bool = True) -> str:
    texto = _texto_limpo(valor)
    if not texto:
        return ""

    texto = (
        texto.replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\u2028", "\n")
        .replace("\u2029", "\n")
    )
    texto = PADRAO_CARACTERES_XML_INVALIDOS.sub("", texto)

    if preservar_quebras:
        linhas = [" ".join(linha.split()) for linha in texto.split("\n")]
        return "\n".join(linha for linha in linhas if linha).strip()

    return " ".join(texto.split()).strip()


def _resumir_texto_exportacao(valor: Any, limite: int, *, preservar_quebras: bool = False) -> str:
    texto = _sanitizar_texto_exportacao(valor, preservar_quebras=preservar_quebras)
    if not texto or len(texto) <= limite:
        return texto
    return f"{texto[: max(limite - 3, 0)].rstrip()}..."


def _extrair_campo_rotulado(texto: Any, rotulo: str) -> str:
    conteudo = _texto_limpo(texto)
    if not conteudo:
        return ""

    correspondencia = re.search(
        rf"{re.escape(rotulo)}\s*:\s*([^.\n\r]+)",
        conteudo,
        flags=re.IGNORECASE,
    )
    return correspondencia.group(1).strip() if correspondencia else ""


def _extrair_primeira_ocorrencia(fontes: Iterable[Any], padrao: re.Pattern[str]) -> str:
    for fonte in fontes:
        conteudo = _texto_limpo(fonte)
        if not conteudo:
            continue
        correspondencia = padrao.search(conteudo)
        if correspondencia:
            return correspondencia.group(1).strip()
    return ""


def _inferir_pavimento_por_unidade(unidade: str) -> str:
    digitos = re.sub(r"\D", "", _texto_limpo(unidade))
    if len(digitos) < 3:
        return ""

    andar = digitos[:-2].lstrip("0") or "0"
    return f"{andar}o PAVIMENTO"


def _montar_dados_agrupamento_imovel(registro: dict[str, Any]) -> dict[str, Any]:
    titulo = _texto_limpo(registro.get("titulo"))
    descricao = _texto_limpo(registro.get("descricao"))
    endereco = _texto_limpo(registro.get("endereco"))
    bairro = _texto_limpo(registro.get("bairro"))
    cidade = _texto_limpo(registro.get("cidade"))

    localidade = (
        _extrair_campo_rotulado(descricao, "Localidade")
        or bairro
        or cidade
        or "Sem localidade"
    )
    bloco = (
        _extrair_campo_rotulado(descricao, "Bloco")
        or _extrair_primeira_ocorrencia((titulo, endereco), PADRAO_BLOCO)
        or "Sem bloco"
    )
    unidade = (
        _extrair_campo_rotulado(descricao, "Unidade")
        or _extrair_primeira_ocorrencia((titulo,), PADRAO_UNIDADE_TITULO)
        or _extrair_primeira_ocorrencia((endereco,), PADRAO_UNIDADE_ENDERECO)
        or titulo
    )
    pavimento = (
        _extrair_campo_rotulado(descricao, "Pavimento")
        or _extrair_primeira_ocorrencia((endereco,), PADRAO_PAVIMENTO_ENDERECO)
        or _inferir_pavimento_por_unidade(unidade)
        or "Andar nao informado"
    )

    return {
        "localidade": localidade,
        "bloco": bloco,
        "andar": pavimento,
        "pavimento": pavimento,
        "unidade": unidade,
    }


PADRAO_COORDENADAS_PARES = re.compile(r"(-?\d{1,2}(?:[.,]\d+)?)\s*[,;]\s*(-?\d{1,3}(?:[.,]\d+)?)")
PADRAO_LATITUDE = re.compile(r"\b(?:lat|latitude)\s*[:=]\s*(-?\d{1,2}(?:[.,]\d+)?)", re.IGNORECASE)
PADRAO_LONGITUDE = re.compile(r"\b(?:lng|lon|long|longitude)\s*[:=]\s*(-?\d{1,3}(?:[.,]\d+)?)", re.IGNORECASE)
LATITUDE_BRASIL_MIN = -34.0
LATITUDE_BRASIL_MAX = 6.0
LONGITUDE_BRASIL_MIN = -74.0
LONGITUDE_BRASIL_MAX = -28.0


def _normalizar_caminho_publico(caminho: Any) -> str:
    valor = _texto_limpo(caminho)
    if not valor:
        return ""
    return f"/{valor.lstrip('/')}"


def _resolver_arquivo_publico_local(caminho_publico: Any) -> Path | None:
    caminho_normalizado = _normalizar_caminho_publico(caminho_publico)
    if not caminho_normalizado:
        return None

    try:
        caminho = (DIRETORIO_PUBLICO / caminho_normalizado.lstrip("/")).resolve()
        caminho.relative_to(DIRETORIO_PUBLICO)
    except Exception:
        return None

    return caminho if caminho.is_file() else None


def _float_seguro(valor: Any) -> float | None:
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        numero = float(valor)
        return numero if numero == numero else None

    texto = _texto_limpo(valor)
    if not texto:
        return None
    texto = texto.replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return None


def _coordenadas_no_brasil(latitude: float | None, longitude: float | None) -> bool:
    if latitude is None or longitude is None:
        return False
    return (
        LATITUDE_BRASIL_MIN <= latitude <= LATITUDE_BRASIL_MAX
        and LONGITUDE_BRASIL_MIN <= longitude <= LONGITUDE_BRASIL_MAX
    )


def _normalizar_coordenadas_brasil(latitude: float | None, longitude: float | None) -> tuple[float | None, float | None]:
    if _coordenadas_no_brasil(latitude, longitude):
        return latitude, longitude

    if _coordenadas_no_brasil(longitude, latitude):
        return longitude, latitude

    return None, None


def _montar_endereco_formatado(registro: dict[str, Any]) -> str:
    partes = [
        _texto_limpo(registro.get("endereco")),
        _texto_limpo(registro.get("bairro")),
        _texto_limpo(registro.get("cidade")),
        _texto_limpo(registro.get("estado")),
        _texto_limpo(registro.get("cep")),
    ]
    return ", ".join(parte for parte in partes if parte)


def _extrair_coordenadas(registro: dict[str, Any], descricao: str) -> tuple[float | None, float | None]:
    latitude, longitude = _normalizar_coordenadas_brasil(
        _float_seguro(registro.get("latitude")),
        _float_seguro(registro.get("longitude")),
    )

    if latitude is not None and longitude is not None:
        return latitude, longitude

    if descricao:
        lat_match = PADRAO_LATITUDE.search(descricao)
        lng_match = PADRAO_LONGITUDE.search(descricao)
        if lat_match and lng_match:
            latitude, longitude = _normalizar_coordenadas_brasil(
                _float_seguro(lat_match.group(1)),
                _float_seguro(lng_match.group(1)),
            )
            if latitude is not None and longitude is not None:
                return latitude, longitude

        par = PADRAO_COORDENADAS_PARES.search(descricao)
        if par:
            latitude, longitude = _normalizar_coordenadas_brasil(
                _float_seguro(par.group(1)),
                _float_seguro(par.group(2)),
            )
            if latitude is not None and longitude is not None:
                return latitude, longitude

    return latitude, longitude


def _montar_localizacao(registro: dict[str, Any]) -> dict[str, Any]:
    descricao = _texto_limpo(registro.get("descricao"))
    endereco_formatado = _montar_endereco_formatado(registro)
    latitude, longitude = _extrair_coordenadas(registro, descricao)

    consulta = f"{latitude},{longitude}" if latitude is not None and longitude is not None else endereco_formatado
    consulta = consulta.strip()
    consulta_encoded = quote_plus(consulta) if consulta else ""

    return {
        "endereco_formatado": endereco_formatado,
        "bairro": _texto_limpo(registro.get("bairro")),
        "cidade": _texto_limpo(registro.get("cidade")),
        "estado": _texto_limpo(registro.get("estado")),
        "cep": _texto_limpo(registro.get("cep")),
        "latitude": latitude,
        "longitude": longitude,
        "consulta": consulta,
        "google_maps_url": (
            f"https://www.google.com/maps/search/?api=1&query={consulta_encoded}" if consulta_encoded else ""
        ),
        "google_maps_embed_url": (
            f"https://www.google.com/maps?q={consulta_encoded}&output=embed" if consulta_encoded else ""
        ),
        "google_maps_rota_url": (
            f"https://www.google.com/maps/dir/?api=1&destination={consulta_encoded}" if consulta_encoded else ""
        ),
    }


def _montar_detalhes_comerciais(registro: dict[str, Any], agrupamento: dict[str, Any]) -> dict[str, Any]:
    descricao = _texto_limpo(registro.get("descricao"))
    area_privativa = _extrair_campo_rotulado(descricao, "Area privativa")
    area_total = (
        _extrair_campo_rotulado(descricao, "Area total construida")
        or _extrair_campo_rotulado(descricao, "Area total")
    )

    return {
        "empreendimento": (
            _extrair_campo_rotulado(descricao, "Empreendimento")
            or _extrair_campo_rotulado(descricao, "Localidade")
            or _texto_limpo(agrupamento.get("localidade"))
            or _texto_limpo(registro.get("bairro"))
        ),
        "bloco": _extrair_campo_rotulado(descricao, "Bloco") or _texto_limpo(agrupamento.get("bloco")),
        "unidade": _extrair_campo_rotulado(descricao, "Unidade") or _texto_limpo(agrupamento.get("unidade")),
        "pavimento": _extrair_campo_rotulado(descricao, "Pavimento") or _texto_limpo(agrupamento.get("pavimento")),
        "andar": _extrair_campo_rotulado(descricao, "Andar") or _texto_limpo(agrupamento.get("andar")),
        "orientacao": _extrair_campo_rotulado(descricao, "Orientacao"),
        "posicao": _extrair_campo_rotulado(descricao, "Posicao"),
        "vista": _extrair_campo_rotulado(descricao, "Vista"),
        "condicao_especial": (
            _extrair_campo_rotulado(descricao, "Condicao especial")
            or _extrair_campo_rotulado(descricao, "Condicao")
        ),
        "area_privativa": area_privativa,
        "area_total": area_total,
    }


def _montar_diferenciais_comerciais(registro: dict[str, Any], detalhes: dict[str, Any]) -> list[str]:
    titulo = _texto_limpo(registro.get("titulo")).lower()
    descricao = _texto_limpo(registro.get("descricao")).lower()
    tipo_imovel = _texto_limpo(registro.get("tipo_imovel")).lower()
    texto = " ".join([titulo, descricao, tipo_imovel]).strip()

    mapa_termos = [
        ("varanda", "Varanda"),
        ("garden", "Garden"),
        ("cobertura", "Cobertura"),
        ("sol da manha", "Sol da manha"),
        ("vista", "Vista valorizada"),
        ("oportunidade", "Oportunidade"),
        ("suite", "Suite"),
        ("mobiliado", "Mobiliado"),
        ("reformado", "Reformado"),
        ("familia", "Ideal para familia"),
        ("localizacao", "Boa localizacao"),
    ]

    diferenciais: list[str] = []
    for termo, rotulo in mapa_termos:
        if termo in texto and rotulo not in diferenciais:
            diferenciais.append(rotulo)

    if detalhes.get("orientacao"):
        diferenciais.append(f"Orientacao: {detalhes['orientacao']}")
    if detalhes.get("posicao"):
        diferenciais.append(f"Posicao: {detalhes['posicao']}")
    if detalhes.get("vista"):
        diferenciais.append(f"Vista: {detalhes['vista']}")
    if detalhes.get("condicao_especial"):
        diferenciais.append(f"Condicao especial: {detalhes['condicao_especial']}")

    # Preserva ordem e remove duplicados.
    vistos: set[str] = set()
    resultado: list[str] = []
    for item in diferenciais:
        chave = item.strip().lower()
        if not chave or chave in vistos:
            continue
        vistos.add(chave)
        resultado.append(item.strip())

    return resultado[:10]


def _montar_midias_comerciais(
    midias_serializadas: list[dict[str, Any]],
    foto_principal_registro: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], str]:
    lista_base = [dict(midia) for midia in (midias_serializadas or [])]
    foto_principal = _normalizar_caminho_publico(foto_principal_registro)

    # Primeiro fotos e depois videos, preservando a ordenacao da consulta.
    lista = [item for item in lista_base if _texto_limpo(item.get("tipo_arquivo")).lower() == "foto"]
    lista.extend(item for item in lista_base if _texto_limpo(item.get("tipo_arquivo")).lower() == "video")

    indice_principal = -1
    if foto_principal:
        for idx, item in enumerate(lista):
            if _normalizar_caminho_publico(item.get("caminho_arquivo")) == foto_principal:
                indice_principal = idx
                break

    if indice_principal < 0:
        for idx, item in enumerate(lista):
            if _texto_limpo(item.get("tipo_arquivo")).lower() == "foto":
                indice_principal = idx
                break

    if indice_principal < 0 and lista:
        indice_principal = 0

    for idx, item in enumerate(lista):
        item["caminho_arquivo"] = _normalizar_caminho_publico(item.get("caminho_arquivo"))
        item["ordem"] = idx + 1
        item["eh_principal"] = idx == indice_principal
        if item["eh_principal"] and not foto_principal:
            foto_principal = item["caminho_arquivo"]

    fotos = [item for item in lista if _texto_limpo(item.get("tipo_arquivo")).lower() == "foto"]
    videos = [item for item in lista if _texto_limpo(item.get("tipo_arquivo")).lower() == "video"]
    return lista, fotos, videos, foto_principal


def serializar_midia(linha: dict[str, Any] | Any) -> dict[str, Any]:
    registro = dict(linha)
    return {
        "id": str(registro["identificador_midia"]),
        "imovel_id": str(registro["identificador_imovel"]),
        "tipo_arquivo": registro["tipo_arquivo"],
        "nome_arquivo": registro["nome_arquivo"],
        "caminho_arquivo": _normalizar_caminho_publico(registro["caminho_arquivo"]),
        "mime_type": registro["mime_type"],
        "tamanho_bytes": int(registro["tamanho_bytes"] or 0),
        "data_hora_criacao": registro["data_hora_criacao"],
    }


def serializar_evolucao_obra(linha: dict[str, Any] | Any) -> dict[str, Any]:
    registro = dict(linha)
    return {
        "id": str(registro["identificador_evolucao_obra"]),
        "imovel_id": str(registro["identificador_imovel"]),
        "percentual_conclusao_obra": _decimal_para_float(registro.get("percentual_conclusao_obra") or Decimal("0")),
        "data_referencia": _data_para_iso(registro.get("data_referencia")),
        "observacoes": registro.get("observacoes"),
        "registrado_por": str(registro.get("registrado_por") or "") or None,
        "data_hora_criacao": registro.get("data_hora_criacao"),
    }


def serializar_imovel(
    linha: dict[str, Any] | Any,
    *,
    midias: Iterable[dict[str, Any]] | None = None,
    evolucao_obra: Iterable[dict[str, Any]] | None = None,
    usar_prazo_entrega_dinamico: bool = False,
) -> dict[str, Any]:
    registro = dict(linha)
    quantidade_midias = int(registro.get("quantidade_midias") or 0)
    midias_serializadas = [serializar_midia(midia) for midia in (midias or [])]
    evolucao_obra_serializada = [serializar_evolucao_obra(item) for item in (evolucao_obra or [])]
    agrupamento = _montar_dados_agrupamento_imovel(registro)
    detalhes_comerciais = _montar_detalhes_comerciais(registro, agrupamento)
    localizacao = _montar_localizacao(registro)
    data_entrega = registro.get("data_entrega")
    meses_pre_entrega = calcular_meses_ate_entrega(data_entrega)

    if usar_prazo_entrega_dinamico:
        data_entrega = calcular_data_entrega_dinamica()
        meses_pre_entrega = calcular_meses_ate_entrega_dinamica()
    elif meses_pre_entrega in (None, ""):
        meses_pre_entrega = registro.get("meses_pre_entrega")

    midias_comerciais, fotos, videos, foto_principal = _montar_midias_comerciais(
        midias_serializadas,
        registro.get("foto_principal"),
    )
    diferenciais_comerciais = _montar_diferenciais_comerciais(registro, detalhes_comerciais)

    if midias_comerciais:
        quantidade_midias = len(midias_comerciais)
    foto_principal_nome = _texto_limpo(registro.get("foto_principal_nome"))
    if not foto_principal_nome and fotos:
        foto_principal_nome = _texto_limpo(fotos[0].get("nome_arquivo"))

    return {
        "id": str(registro["identificador_imovel"]),
        "titulo": registro["titulo"],
        "descricao": registro.get("descricao"),
        "tipo_imovel": registro.get("tipo_imovel"),
        "endereco": registro.get("endereco"),
        "cidade": registro["cidade"],
        "bairro": registro["bairro"],
        "estado": registro.get("estado"),
        "cep": registro.get("cep"),
        "valor": _decimal_para_float(registro.get("valor")),
        "quartos": registro.get("quartos"),
        "banheiros": registro.get("banheiros"),
        "vagas_garagem": registro.get("vagas_garagem"),
        "tipo_garagem": registro.get("tipo_garagem") or "carro",
        "area_m2": _decimal_para_float(registro.get("area_m2")),
        "data_entrega": _data_para_iso(data_entrega),
        "meses_pre_entrega": int(meses_pre_entrega or 36),
        "meses_pos_entrega": int(24 if registro.get("meses_pos_entrega") in (None, "") else registro.get("meses_pos_entrega")),
        "percentual_conclusao_obra": _decimal_para_float(registro.get("percentual_conclusao_obra") or 0),
        "percentual_fechamento_minimo": _decimal_para_float(registro.get("percentual_fechamento_minimo") or Decimal("0.70")),
        "valor_garantido": _decimal_para_float(
            registro.get("valor_garantido")
            or calcular_valor_garantido_padrao(
                registro.get("valor"),
                registro.get("percentual_fechamento_minimo") or Decimal("0.70"),
            )
        ),
        "valor_garantido_planejado": _decimal_para_float(
            registro.get("valor_garantido")
            or calcular_valor_garantido_padrao(
                registro.get("valor"),
                registro.get("percentual_fechamento_minimo") or Decimal("0.70"),
            )
        ),
        "valor_garantido_pre_obra_planejado": _decimal_para_float(
            registro.get("valor_garantido_pre_obra_planejado")
            or calcular_valor_captacao_ate_entrega_planejado(
                registro.get("valor"),
                registro.get("percentual_captacao_ate_entrega"),
            )
        ),
        "percentual_captacao_ate_entrega": _decimal_para_float(
            registro.get("percentual_captacao_ate_entrega")
            or calcular_percentual_captacao_ate_entrega_por_valor(
                registro.get("valor"),
                registro.get("valor_garantido_pre_obra_planejado"),
            )
        ),
        "valor_parcela_minima_pre_obra": _decimal_para_float(registro.get("valor_parcela_minima_pre_obra") or Decimal("0.00")),
        "status": registro["status"],
        "agrupamento": agrupamento,
        "detalhes_comerciais": detalhes_comerciais,
        "diferenciais_comerciais": diferenciais_comerciais,
        "localizacao": localizacao,
        "endereco_formatado": localizacao.get("endereco_formatado"),
        "foto_principal": foto_principal,
        "foto_principal_nome": foto_principal_nome or None,
        "quantidade_midias": quantidade_midias,
        "quantidade_fotos": len(fotos) if midias_comerciais else 0,
        "quantidade_videos": len(videos) if midias_comerciais else 0,
        "data_hora_criacao": registro.get("data_hora_criacao"),
        "data_hora_atualizado_em": registro.get("data_hora_atualizado_em"),
        "midias": midias_comerciais,
        "midias_fotos": fotos,
        "midias_videos": videos,
        "evolucao_obra": evolucao_obra_serializada,
        "ultima_evolucao_obra": evolucao_obra_serializada[0] if evolucao_obra_serializada else None,
        "reserva_ativa": serializar_reserva(registro, prefixo="reserva_")
        if registro.get("reserva_identificador_reserva")
        else None,
    }


EXPORTACAO_IMOVEIS_COLUNAS = [
    {"chave": "titulo", "rotulo": "Imovel", "largura": 30, "tipo": "texto"},
    {"chave": "tipo_imovel", "rotulo": "Tipo", "largura": 16, "tipo": "texto"},
    {"chave": "status", "rotulo": "Status", "largura": 15, "tipo": "texto"},
    {"chave": "cidade", "rotulo": "Cidade", "largura": 18, "tipo": "texto"},
    {"chave": "bairro", "rotulo": "Bairro", "largura": 20, "tipo": "texto"},
    {"chave": "endereco_formatado", "rotulo": "Endereco", "largura": 42, "tipo": "texto"},
    {"chave": "valor", "rotulo": "Valor do imovel", "largura": 16, "tipo": "moeda"},
    {"chave": "valor_garantido", "rotulo": "Valor garantido", "largura": 16, "tipo": "moeda"},
    {"chave": "valor_parcela_minima_pre_obra", "rotulo": "Pre-obra minima", "largura": 16, "tipo": "moeda"},
    {"chave": "percentual_fechamento_minimo", "rotulo": "Captacao minima", "largura": 15, "tipo": "percentual"},
    {"chave": "data_entrega", "rotulo": "Data de entrega", "largura": 15, "tipo": "data"},
    {"chave": "meses_pre_entrega", "rotulo": "Meses ate entrega", "largura": 14, "tipo": "inteiro"},
    {"chave": "meses_pos_entrega", "rotulo": "Meses pos-entrega", "largura": 14, "tipo": "inteiro"},
    {"chave": "percentual_conclusao_obra", "rotulo": "Conclusao da obra", "largura": 15, "tipo": "percentual"},
    {"chave": "quartos", "rotulo": "Quartos", "largura": 10, "tipo": "inteiro"},
    {"chave": "banheiros", "rotulo": "Banheiros", "largura": 10, "tipo": "inteiro"},
    {"chave": "vagas_garagem", "rotulo": "Vagas", "largura": 10, "tipo": "inteiro"},
    {"chave": "tipo_garagem", "rotulo": "Tipo de garagem", "largura": 14, "tipo": "texto"},
    {"chave": "area_m2", "rotulo": "Area m2", "largura": 12, "tipo": "decimal"},
    {"chave": "quantidade_midias", "rotulo": "Midias", "largura": 10, "tipo": "inteiro"},
    {"chave": "reserva_cliente", "rotulo": "Cliente com reserva", "largura": 24, "tipo": "texto"},
    {"chave": "data_hora_atualizado_em", "rotulo": "Atualizado em", "largura": 19, "tipo": "datetime"},
]


def _numero_decimal(valor: Any) -> Decimal | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, str):
        texto = _texto_limpo(valor)
        if not texto:
            return None
        texto = re.sub(r"[^0-9,.\-]", "", texto)
        if not texto:
            return None
        if "," in texto and "." in texto:
            if texto.rfind(",") > texto.rfind("."):
                texto = texto.replace(".", "").replace(",", ".")
            else:
                texto = texto.replace(",", "")
        elif "," in texto:
            texto = texto.replace(",", ".")
        valor = texto
    try:
        return Decimal(str(valor))
    except Exception:
        return None


def _percentual_para_fracao(valor: Any) -> Decimal | None:
    numero = _numero_decimal(valor)
    if numero is None:
        return None
    if abs(numero) > Decimal("1.5"):
        return numero / Decimal("100")
    return numero


def _formatar_numero_br(valor: Any, casas: int = 2) -> str:
    numero = _numero_decimal(valor)
    if numero is None:
        return "-"
    texto = f"{float(numero):,.{casas}f}"
    return texto.replace(",", "_").replace(".", ",").replace("_", ".")


def _formatar_moeda_br(valor: Any) -> str:
    numero = _numero_decimal(valor)
    if numero is None:
        return "-"
    return f"R$ {_formatar_numero_br(numero, 2)}"


def _formatar_percentual_br(valor: Any) -> str:
    fracao = _percentual_para_fracao(valor)
    if fracao is None:
        return "-"
    return f"{_formatar_numero_br(fracao * Decimal('100'), 2)}%"


def _parse_data_ou_data_hora(valor: Any) -> datetime | None:
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor
    if hasattr(valor, "year") and hasattr(valor, "month") and hasattr(valor, "day"):
        try:
            return datetime(valor.year, valor.month, valor.day)
        except Exception:
            return None

    texto = str(valor).strip()
    if not texto:
        return None
    texto = texto.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(texto)
    except Exception:
        return None


def _normalizar_datetime_excel(valor: Any) -> datetime | None:
    data = _parse_data_ou_data_hora(valor)
    if not data:
        return None
    if getattr(data, "tzinfo", None) is None:
        return data
    try:
        return data.astimezone().replace(tzinfo=None)
    except Exception:
        return data.replace(tzinfo=None)


def _formatar_data_br(valor: Any) -> str:
    data = _parse_data_ou_data_hora(valor)
    if not data:
        return "-"
    return data.strftime("%d/%m/%Y")


def _formatar_data_hora_br(valor: Any) -> str:
    data = _parse_data_ou_data_hora(valor)
    if not data:
        return "-"
    return data.strftime("%d/%m/%Y %H:%M")


def _inteiro_seguro(valor: Any) -> int | None:
    numero = _numero_decimal(valor)
    if numero is None:
        return None
    try:
        return int(numero)
    except Exception:
        return None


def _formatar_inteiro(valor: Any) -> str:
    numero = _inteiro_seguro(valor)
    return str(numero) if numero is not None else "-"


def _primeiro_texto_exportacao(*valores: Any) -> str:
    for valor in valores:
        texto = _sanitizar_texto_exportacao(valor)
        if texto:
            return texto
    return "-"


def _mapear_imovel_para_exportacao(imovel: dict[str, Any]) -> dict[str, Any]:
    reserva = imovel.get("reserva_ativa") or {}
    return {
        "titulo": _primeiro_texto_exportacao(imovel.get("titulo")),
        "tipo_imovel": _primeiro_texto_exportacao(imovel.get("tipo_imovel")),
        "status": _primeiro_texto_exportacao(imovel.get("status")),
        "cidade": _primeiro_texto_exportacao(imovel.get("cidade")),
        "bairro": _primeiro_texto_exportacao(imovel.get("bairro")),
        "endereco_formatado": _primeiro_texto_exportacao(
            imovel.get("endereco_formatado"),
            imovel.get("endereco"),
        ),
        "valor": imovel.get("valor"),
        "valor_garantido": imovel.get("valor_garantido"),
        "valor_parcela_minima_pre_obra": imovel.get("valor_parcela_minima_pre_obra"),
        "percentual_fechamento_minimo": imovel.get("percentual_fechamento_minimo"),
        "data_entrega": imovel.get("data_entrega"),
        "meses_pre_entrega": imovel.get("meses_pre_entrega"),
        "meses_pos_entrega": imovel.get("meses_pos_entrega"),
        "percentual_conclusao_obra": imovel.get("percentual_conclusao_obra"),
        "quartos": imovel.get("quartos"),
        "banheiros": imovel.get("banheiros"),
        "vagas_garagem": imovel.get("vagas_garagem"),
        "tipo_garagem": _primeiro_texto_exportacao(imovel.get("tipo_garagem")),
        "area_m2": imovel.get("area_m2"),
        "quantidade_midias": imovel.get("quantidade_midias") or 0,
        "foto_principal": _normalizar_caminho_publico(imovel.get("foto_principal")),
        "foto_principal_nome": _primeiro_texto_exportacao(imovel.get("foto_principal_nome")),
        "reserva_cliente": _primeiro_texto_exportacao(
            reserva.get("cliente_nome"),
            reserva.get("cliente_cpf"),
            "-" if not reserva else "",
        ),
        "data_hora_atualizado_em": imovel.get("data_hora_atualizado_em"),
    }


def _valor_texto_exportacao(linha: dict[str, Any], coluna: dict[str, Any]) -> str:
    chave = coluna["chave"]
    tipo = coluna["tipo"]
    valor = linha.get(chave)

    if tipo == "moeda":
        return _formatar_moeda_br(valor)
    if tipo == "percentual":
        return _formatar_percentual_br(valor)
    if tipo == "data":
        return _formatar_data_br(valor)
    if tipo == "datetime":
        return _formatar_data_hora_br(valor)
    if tipo == "decimal":
        return _formatar_numero_br(valor, 2) if valor not in (None, "") else "-"
    if tipo == "inteiro":
        return _formatar_inteiro(valor)
    return _primeiro_texto_exportacao(valor)


def _valor_xlsx_exportacao(linha: dict[str, Any], coluna: dict[str, Any]) -> Any:
    chave = coluna["chave"]
    tipo = coluna["tipo"]
    valor = linha.get(chave)

    if tipo == "moeda":
        numero = _numero_decimal(valor)
        return float(numero) if numero is not None else None
    if tipo == "percentual":
        numero = _percentual_para_fracao(valor)
        return float(numero) if numero is not None else None
    if tipo == "decimal":
        numero = _numero_decimal(valor)
        return float(numero) if numero is not None else None
    if tipo == "inteiro":
        return _inteiro_seguro(valor)
    if tipo == "data":
        data = _normalizar_datetime_excel(valor)
        return data.date() if data else None
    if tipo == "datetime":
        return _normalizar_datetime_excel(valor)
    return None if valor in (None, "") else _sanitizar_texto_exportacao(valor)


def _media_exportacao(valores: list[float]) -> float:
    return (sum(valores) / len(valores)) if valores else 0.0


def _metricas_financeiras_exportacao(linhas: list[dict[str, Any]]) -> dict[str, Any]:
    valores = [float(numero) for numero in (_numero_decimal(linha.get("valor")) for linha in linhas) if numero is not None]
    garantidos = [
        float(numero)
        for numero in (_numero_decimal(linha.get("valor_garantido")) for linha in linhas)
        if numero is not None
    ]
    pre_obra = [
        float(numero)
        for numero in (_numero_decimal(linha.get("valor_parcela_minima_pre_obra")) for linha in linhas)
        if numero is not None
    ]
    conclusao_obra = [
        float(fracao * Decimal("100"))
        for fracao in (_percentual_para_fracao(linha.get("percentual_conclusao_obra")) for linha in linhas)
        if fracao is not None
    ]
    reservas = sum(1 for linha in linhas if _texto_limpo(linha.get("reserva_cliente")) not in {"", "-"})
    sem_midias = sum(1 for linha in linhas if (_inteiro_seguro(linha.get("quantidade_midias")) or 0) <= 0)
    total_midias = sum((_inteiro_seguro(linha.get("quantidade_midias")) or 0) for linha in linhas)
    datas_entrega = sorted(
        data for data in (_parse_data_ou_data_hora(linha.get("data_entrega")) for linha in linhas) if data is not None
    )

    return {
        "valor_total": sum(valores),
        "valor_garantido_total": sum(garantidos),
        "valor_medio": _media_exportacao(valores),
        "valor_garantido_medio": _media_exportacao(garantidos),
        "pre_obra_media": _media_exportacao(pre_obra),
        "conclusao_obra_media": _media_exportacao(conclusao_obra),
        "reservas": reservas,
        "sem_midias": sem_midias,
        "total_midias": total_midias,
        "proxima_entrega": datas_entrega[0] if datas_entrega else None,
    }


def _resumo_exportacao_imoveis(linhas: list[dict[str, Any]]) -> dict[str, Any]:
    total = len(linhas)
    disponiveis = sum(1 for linha in linhas if _texto_limpo(linha.get("status")).lower() == "disponivel")
    reservados = sum(1 for linha in linhas if _texto_limpo(linha.get("status")).lower() == "reservado")
    vendidos = sum(1 for linha in linhas if _texto_limpo(linha.get("status")).lower() == "vendido")
    inativos = sum(1 for linha in linhas if _texto_limpo(linha.get("status")).lower() == "inativo")
    valores = [float(_numero_decimal(linha.get("valor"))) for linha in linhas if _numero_decimal(linha.get("valor")) is not None]
    cidades = {str(linha.get("cidade")).strip() for linha in linhas if _texto_limpo(linha.get("cidade"))}
    return {
        "total": total,
        "disponiveis": disponiveis,
        "reservados": reservados,
        "vendidos": vendidos,
        "inativos": inativos,
        "ticket_medio": (sum(valores) / len(valores)) if valores else 0.0,
        "cidades": len(cidades),
    }


def _descricao_filtros_exportacao(filtros: dict[str, Any]) -> str:
    partes: list[str] = []
    busca = _sanitizar_texto_exportacao(filtros.get("q"), preservar_quebras=False)
    cidade = _sanitizar_texto_exportacao(filtros.get("cidade"), preservar_quebras=False)
    bairro = _sanitizar_texto_exportacao(filtros.get("bairro"), preservar_quebras=False)
    status = _sanitizar_texto_exportacao(filtros.get("status"), preservar_quebras=False)
    if busca:
        partes.append(f"Busca: {busca}")
    if cidade:
        partes.append(f"Cidade: {cidade}")
    if bairro:
        partes.append(f"Bairro: {bairro}")
    if status:
        partes.append(f"Status: {status}")
    return " | ".join(partes) if partes else "Sem filtros adicionais."


def _nome_arquivo_exportacao_imoveis(formato: str) -> str:
    carimbo = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"imoveis-cadastrados-{carimbo}.{formato}"


def gerar_csv_exportacao_imoveis(imoveis: list[dict[str, Any]]) -> bytes:
    import io

    stream = io.StringIO()
    writer = csv.writer(stream, delimiter=";", lineterminator="\r\n")
    writer.writerow([coluna["rotulo"] for coluna in EXPORTACAO_IMOVEIS_COLUNAS])
    for item in imoveis:
        linha = _mapear_imovel_para_exportacao(item)
        writer.writerow([_valor_texto_exportacao(linha, coluna) for coluna in EXPORTACAO_IMOVEIS_COLUNAS])
    conteudo = "\ufeff" + stream.getvalue()
    return conteudo.encode("utf-8")


def gerar_xlsx_exportacao_imoveis(imoveis: list[dict[str, Any]], filtros: dict[str, Any]) -> bytes:
    if Workbook is None or get_column_letter is None:
        raise HTTPException(status_code=503, detail="Exportacao Excel indisponivel neste ambiente.")

    resumo = _resumo_exportacao_imoveis([_mapear_imovel_para_exportacao(item) for item in imoveis])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Imoveis"

    ultima_coluna = get_column_letter(len(EXPORTACAO_IMOVEIS_COLUNAS))
    borda = Border(
        left=Side(style="thin", color="D6E4F0"),
        right=Side(style="thin", color="D6E4F0"),
        top=Side(style="thin", color="D6E4F0"),
        bottom=Side(style="thin", color="D6E4F0"),
    )
    fill_escuro = PatternFill("solid", fgColor="0B1524")
    fill_destaque = PatternFill("solid", fgColor="EAF7FF")
    fill_cabecalho = PatternFill("solid", fgColor="1677FF")
    fill_linha_par = PatternFill("solid", fgColor="F8FBFF")
    fonte_titulo = Font(color="FFFFFF", bold=True, size=16)
    fonte_meta = Font(color="1E293B", bold=True, size=10)
    fonte_cabecalho = Font(color="FFFFFF", bold=True, size=10)

    sheet.merge_cells(f"A1:{ultima_coluna}1")
    sheet["A1"] = "Relatorio de imoveis cadastrados"
    sheet["A1"].fill = fill_escuro
    sheet["A1"].font = fonte_titulo
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")

    sheet.merge_cells(f"A2:{ultima_coluna}2")
    sheet["A2"] = _sanitizar_texto_exportacao(
        f"Gerado em {_formatar_data_hora_br(datetime.now())} | "
        f"Filtros: {_descricao_filtros_exportacao(filtros)}"
    )
    sheet["A2"].fill = fill_destaque
    sheet["A2"].font = fonte_meta
    sheet["A2"].alignment = Alignment(wrap_text=True)

    metricas = [
        ("Total", resumo["total"]),
        ("Disponiveis", resumo["disponiveis"]),
        ("Reservados", resumo["reservados"]),
        ("Vendidos", resumo["vendidos"]),
        ("Inativos", resumo["inativos"]),
        ("Ticket medio", _formatar_moeda_br(resumo["ticket_medio"])),
        ("Cidades", resumo["cidades"]),
        ("Exportado", "Lista completa filtrada"),
    ]
    for indice, (rotulo, valor) in enumerate(metricas, start=1):
        linha = 3 + ((indice - 1) // 4)
        coluna_base = ((indice - 1) % 4) * 5 + 1
        sheet.cell(row=linha, column=coluna_base, value=rotulo)
        sheet.cell(row=linha, column=coluna_base + 1, value=valor)
        sheet.cell(row=linha, column=coluna_base).font = Font(color="5B6B82", bold=True, size=9)
        sheet.cell(row=linha, column=coluna_base + 1).font = Font(color="0F172A", bold=True, size=11)

    linha_cabecalho = 7
    for indice, coluna in enumerate(EXPORTACAO_IMOVEIS_COLUNAS, start=1):
        celula = sheet.cell(row=linha_cabecalho, column=indice, value=coluna["rotulo"])
        celula.fill = fill_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = borda
        sheet.column_dimensions[get_column_letter(indice)].width = coluna["largura"]

    for indice_linha, item in enumerate(imoveis, start=linha_cabecalho + 1):
        linha_exportacao = _mapear_imovel_para_exportacao(item)
        for indice_coluna, coluna in enumerate(EXPORTACAO_IMOVEIS_COLUNAS, start=1):
            celula = sheet.cell(
                row=indice_linha,
                column=indice_coluna,
                value=_valor_xlsx_exportacao(linha_exportacao, coluna),
            )
            celula.alignment = Alignment(
                horizontal="center" if coluna["tipo"] in {"inteiro", "percentual"} else "left",
                vertical="top",
                wrap_text=True,
            )
            celula.border = borda
            if indice_linha % 2 == 0:
                celula.fill = fill_linha_par
            if coluna["tipo"] == "moeda":
                celula.number_format = 'R$ #,##0.00'
            elif coluna["tipo"] == "percentual":
                celula.number_format = "0.00%"
            elif coluna["tipo"] == "data":
                celula.number_format = "DD/MM/YYYY"
            elif coluna["tipo"] == "datetime":
                celula.number_format = "DD/MM/YYYY HH:MM"
            elif coluna["tipo"] == "decimal":
                celula.number_format = "#,##0.00"

    sheet.freeze_panes = f"A{linha_cabecalho + 1}"
    sheet.auto_filter.ref = f"A{linha_cabecalho}:{ultima_coluna}{max(linha_cabecalho + 1, sheet.max_row)}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _xml_sheet_inline_text(valor: str) -> str:
    valor_limpo = _sanitizar_texto_exportacao(valor)
    return (
        '<c t="inlineStr"><is><t xml:space="preserve">'
        f"{xml_escape(valor_limpo)}"
        "</t></is></c>"
    )


def _xml_sheet_number(valor: Any) -> str:
    return f"<c><v>{xml_escape(str(valor))}</v></c>"


def _xml_sheet_data(rows: list[list[str]]) -> bytes:
    linhas = []
    for indice_linha, valores in enumerate(rows, start=1):
        celulas = []
        for valor in valores:
            numero = _numero_decimal(valor)
            if numero is not None and str(valor).strip().replace(".", "").replace(",", "").isdigit():
                celulas.append(_xml_sheet_number(numero))
            else:
                celulas.append(_xml_sheet_inline_text(str(valor)))
        linhas.append(f'<row r="{indice_linha}">{"".join(celulas)}</row>')

    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "<sheetData>"
        f"{''.join(linhas)}"
        "</sheetData>"
        "</worksheet>"
    )
    return xml.encode("utf-8")


def gerar_xlsx_exportacao_imoveis_fallback(imoveis: list[dict[str, Any]], filtros: dict[str, Any]) -> bytes:
    resumo = _resumo_exportacao_imoveis([_mapear_imovel_para_exportacao(item) for item in imoveis])
    linhas = [
        ["Relatorio de imoveis cadastrados"],
        [f"Gerado em {_formatar_data_hora_br(datetime.now())}"],
        [f"Filtros: {_descricao_filtros_exportacao(filtros)}"],
        [f"Total: {resumo['total']} | Disponiveis: {resumo['disponiveis']} | Reservados: {resumo['reservados']} | Ticket medio: {_formatar_moeda_br(resumo['ticket_medio'])}"],
        [""],
        [coluna["rotulo"] for coluna in EXPORTACAO_IMOVEIS_COLUNAS],
    ]
    for item in imoveis:
        linha = _mapear_imovel_para_exportacao(item)
        linhas.append([_valor_texto_exportacao(linha, coluna) for coluna in EXPORTACAO_IMOVEIS_COLUNAS])

    workbook = BytesIO()
    with ZipFile(workbook, "w", ZIP_DEFLATED) as zip_file:
        zip_file.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>""",
        )
        zip_file.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>""",
        )
        zip_file.writestr(
            "xl/workbook.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    <sheet name="Imoveis" sheetId="1" r:id="rId1"/>
  </sheets>
</workbook>""",
        )
        zip_file.writestr(
            "xl/_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>""",
        )
        zip_file.writestr("xl/worksheets/sheet1.xml", _xml_sheet_data(linhas))
        zip_file.writestr(
            "docProps/core.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>Relatorio de imoveis cadastrados</dc:title>
  <dc:creator>7LM Connect</dc:creator>
</cp:coreProperties>""",
        )
        zip_file.writestr(
            "docProps/app.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>7LM Connect</Application>
</Properties>""",
        )
    return workbook.getvalue()


def gerar_pdf_exportacao_imoveis(imoveis: list[dict[str, Any]], filtros: dict[str, Any]) -> bytes:
    linhas = [_mapear_imovel_para_exportacao(item) for item in imoveis]
    resumo = _resumo_exportacao_imoveis(linhas)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=10 * mm,
        title="Relatorio de imoveis cadastrados",
    )

    base = getSampleStyleSheet()
    azul_principal = colors.HexColor("#0F7BFF")
    azul_turquesa = colors.HexColor("#00B3DE")
    azul_claro = colors.HexColor("#EAF7FF")
    azul_fundo = colors.HexColor("#F6FBFE")
    azul_card = colors.HexColor("#F8FCFF")
    azul_card_secundario = colors.HexColor("#EEF8FD")
    borda = colors.HexColor("#D8E8F4")
    texto_escuro = colors.HexColor("#172033")
    texto_medio = colors.HexColor("#5B6B82")
    hero_fundo = colors.HexColor("#081222")

    estilo_hero_kicker = ParagraphStyle(
        "ExportHeroKicker",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#8EEEFF"),
        alignment=TA_LEFT,
    )
    estilo_hero_titulo = ParagraphStyle(
        "ExportHeroTitle",
        parent=base["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=23,
        textColor=colors.white,
        alignment=TA_LEFT,
        spaceAfter=0,
    )
    estilo_hero_subtitulo = ParagraphStyle(
        "ExportHeroSubtitle",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=9.2,
        leading=11.4,
        textColor=colors.HexColor("#DDEEFE"),
        alignment=TA_LEFT,
    )
    estilo_hero_meta = ParagraphStyle(
        "ExportHeroMeta",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.2,
        textColor=colors.HexColor("#DFF6FF"),
        alignment=TA_LEFT,
    )
    estilo_kpi_rotulo = ParagraphStyle(
        "ExportKpiLabel",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        textColor=texto_medio,
        alignment=TA_LEFT,
    )
    estilo_kpi_valor = ParagraphStyle(
        "ExportKpiValue",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11.5,
        leading=13,
        textColor=texto_escuro,
        alignment=TA_LEFT,
    )
    estilo_painel_rotulo = ParagraphStyle(
        "ExportPanelLabel",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        textColor=azul_principal,
        alignment=TA_LEFT,
    )
    estilo_painel_texto = ParagraphStyle(
        "ExportPanelText",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=texto_escuro,
        alignment=TA_LEFT,
    )
    estilo_card_kicker = ParagraphStyle(
        "ExportPropertyKicker",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        textColor=azul_principal,
        alignment=TA_LEFT,
    )
    estilo_card_titulo = ParagraphStyle(
        "ExportPropertyTitle",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=texto_escuro,
        alignment=TA_LEFT,
    )
    estilo_card_texto = ParagraphStyle(
        "ExportPropertyText",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.6,
        leading=10.6,
        textColor=texto_medio,
        alignment=TA_LEFT,
    )
    estilo_card_metric_rotulo = ParagraphStyle(
        "ExportMetricLabel",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=6.4,
        leading=8,
        textColor=texto_medio,
        alignment=TA_LEFT,
    )
    estilo_card_metric_valor = ParagraphStyle(
        "ExportMetricValue",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=10.6,
        textColor=texto_escuro,
        alignment=TA_LEFT,
    )
    estilo_reserva = ParagraphStyle(
        "ExportReserva",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=10,
        textColor=colors.HexColor("#9A6400"),
        alignment=TA_LEFT,
    )
    estilo_vazio = ParagraphStyle(
        "ExportEmpty",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        textColor=texto_medio,
        alignment=TA_LEFT,
    )
    estilo_logo_texto = ParagraphStyle(
        "ExportLogoText",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=18,
        textColor=texto_escuro,
        alignment=TA_LEFT,
    )
    estilo_hero_badge = ParagraphStyle(
        "ExportHeroBadge",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.3,
        leading=8.4,
        textColor=colors.white,
        alignment=TA_RIGHT,
    )
    estilo_hero_metric_rotulo = ParagraphStyle(
        "ExportHeroMetricLabel",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=7.7,
        textColor=colors.HexColor("#CBEAFA"),
        alignment=TA_LEFT,
    )
    estilo_hero_metric_valor = ParagraphStyle(
        "ExportHeroMetricValue",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11.2,
        leading=12.8,
        textColor=colors.white,
        alignment=TA_LEFT,
    )
    estilo_hero_media_titulo = ParagraphStyle(
        "ExportHeroMediaTitle",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=12.6,
        textColor=colors.white,
        alignment=TA_LEFT,
    )
    estilo_hero_media_texto = ParagraphStyle(
        "ExportHeroMediaText",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.1,
        leading=9.6,
        textColor=colors.HexColor("#DDEEFE"),
        alignment=TA_LEFT,
    )

    def _paragrafo_seguro(valor: Any, estilo: ParagraphStyle, fallback: str = "-") -> Paragraph:
        texto = _texto_limpo(valor) or fallback
        texto = xml_escape(texto).replace("\n", "<br/>")
        return Paragraph(texto, estilo)

    def _badge_status(status_texto: str) -> Table:
        status_normalizado = _texto_limpo(status_texto).lower()
        mapa = {
            "disponivel": ("#E8FFF7", "#0D7F5F"),
            "reservado": ("#FFF5E8", "#9A6400"),
            "vendido": ("#FFECEC", "#B42318"),
            "inativo": ("#EDF2F7", "#475467"),
        }
        fundo, cor = mapa.get(status_normalizado, ("#EDF2F7", "#475467"))
        pill = Table(
            [[[Paragraph(xml_escape((_texto_limpo(status_texto) or "Indefinido").upper()), ParagraphStyle(
                f"ExportStatus-{status_normalizado or 'default'}",
                parent=base["BodyText"],
                fontName="Helvetica-Bold",
                fontSize=7.2,
                leading=8.4,
                textColor=colors.HexColor(cor),
                alignment=TA_RIGHT,
            ))]]],
            colWidths=[30 * mm],
        )
        pill.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(fundo)),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor(fundo)),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return pill

    def _kpi_card(rotulo: str, valor: str, largura: float) -> Table:
        card = Table(
            [[[Paragraph(xml_escape(rotulo.upper()), estilo_kpi_rotulo), Spacer(1, 1.5 * mm), Paragraph(xml_escape(valor), estilo_kpi_valor)]]],
            colWidths=[largura],
        )
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.7, borda),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        return card

    def _hero_logo_badge() -> Table:
        logo_path = (DIRETORIO_PUBLICO / "assets" / "7lm_logo.png").resolve()
        conteudo: Any
        if logo_path.is_file():
            try:
                ImageReader(str(logo_path))
                conteudo = PlatypusImage(str(logo_path), width=24 * mm, height=10.5 * mm, kind="proportional")
            except Exception:
                conteudo = Paragraph("7LM", estilo_logo_texto)
        else:
            conteudo = Paragraph("7LM", estilo_logo_texto)

        badge = Table([[conteudo]], colWidths=[30 * mm], rowHeights=[13.5 * mm])
        badge.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#E5F0F7")),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return badge

    def _hero_status_badge(texto: str, largura: float = 28 * mm) -> Table:
        pill = Table(
            [[Paragraph(xml_escape(_texto_limpo(texto).upper() or "CARTEIRA"), estilo_hero_badge)]],
            colWidths=[largura],
            rowHeights=[11 * mm],
        )
        pill.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2A7394")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#70A9C5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return pill

    def _hero_metric_card(rotulo: str, valor: str, largura: float, *, destaque: bool = False) -> Table:
        cor_fundo = colors.HexColor("#2A7394" if destaque else "#245F80")
        cor_borda = colors.HexColor("#5B93AF" if destaque else "#4A86A5")
        card = Table(
            [[[Paragraph(xml_escape(rotulo.upper()), estilo_hero_metric_rotulo), Spacer(1, 1.1 * mm), Paragraph(xml_escape(valor), estilo_hero_metric_valor)]]],
            colWidths=[largura],
        )
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), cor_fundo),
            ("BOX", (0, 0), (-1, -1), 0.7, cor_borda),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        return card

    def _hero_media_card(linha: dict[str, Any], largura: float, altura: float) -> Table:
        caminho_midia = _resolver_arquivo_publico_local(linha.get("foto_principal"))
        destaque_titulo = _valor_texto_exportacao(linha, {"chave": "titulo", "tipo": "texto"})
        destaque_local = _resumir_texto_exportacao(descrever_localizacao(linha), 78) or "Local nao informado"
        destaque_detalhes = _resumir_texto_exportacao(descrever_detalhes(linha), 84) or "Sem detalhes adicionais"

        conteudo: list[Any] = []
        if caminho_midia is not None:
            try:
                ImageReader(str(caminho_midia))
                conteudo.append(
                    PlatypusImage(
                        str(caminho_midia),
                        width=max(largura - 16, 10),
                        height=max(altura - 16, 10),
                        kind="proportional",
                    )
                )
            except Exception:
                conteudo = []

        if not conteudo:
            conteudo = [
                Paragraph("UNIDADE EM DESTAQUE", estilo_hero_kicker),
                Spacer(1, 2 * mm),
                Paragraph(xml_escape(destaque_titulo), estilo_hero_media_titulo),
                Spacer(1, 1.4 * mm),
                Paragraph(xml_escape(destaque_local), estilo_hero_media_texto),
                Spacer(1, 1 * mm),
                Paragraph(xml_escape(destaque_detalhes), estilo_hero_media_texto),
            ]

        card = Table([[conteudo]], colWidths=[largura], rowHeights=[altura])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#294768")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#4E6F94")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return card

    def _card_metrica(label: str, value: str, largura: float) -> Table:
        bloco = Table(
            [[[Paragraph(xml_escape(label.upper()), estilo_card_metric_rotulo), Spacer(1, 1.2 * mm), Paragraph(xml_escape(value), estilo_card_metric_valor)]]],
            colWidths=[largura],
        )
        bloco.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), azul_card_secundario),
            ("BOX", (0, 0), (-1, -1), 0.55, borda),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        return bloco

    def _card_imovel(linha: dict[str, Any], largura: float) -> Table:
        titulo = _valor_texto_exportacao(linha, EXPORTACAO_IMOVEIS_COLUNAS[0])
        tipo_imovel = _valor_texto_exportacao(linha, EXPORTACAO_IMOVEIS_COLUNAS[1])
        status = _valor_texto_exportacao(linha, EXPORTACAO_IMOVEIS_COLUNAS[2])
        localizacao = " • ".join(
            parte for parte in [
                _texto_limpo(linha.get("bairro")),
                _texto_limpo(linha.get("cidade")),
            ] if parte
        ) or _valor_texto_exportacao(linha, {"chave": "endereco_formatado", "tipo": "texto"})
        detalhes = " • ".join(
            parte for parte in [
                tipo_imovel if tipo_imovel != "-" else "",
                _formatar_numero_br(linha.get("area_m2"), 2) + " m2" if linha.get("area_m2") not in (None, "") else "",
                f"{int(linha.get('quartos'))} quarto(s)" if linha.get("quartos") not in (None, "") else "",
                f"{int(linha.get('vagas_garagem'))} vaga(s)" if linha.get("vagas_garagem") not in (None, "") else "",
                _texto_limpo(linha.get("tipo_garagem")).capitalize() if _texto_limpo(linha.get("tipo_garagem")) else "",
            ] if parte
        )
        metricas = Table(
            [
                [
                    _card_metrica("Valor do imovel", _formatar_moeda_br(linha.get("valor")), (largura - 28) / 2),
                    _card_metrica("Valor garantido", _formatar_moeda_br(linha.get("valor_garantido")), (largura - 28) / 2),
                ],
                [
                    _card_metrica("Pre-obra minima", _formatar_moeda_br(linha.get("valor_parcela_minima_pre_obra")), (largura - 28) / 2),
                    _card_metrica(
                        "Entrega",
                        " • ".join(
                            parte for parte in [
                                _formatar_data_br(linha.get("data_entrega")) if linha.get("data_entrega") else "",
                                f"{linha.get('meses_pre_entrega') or '-'} meses",
                            ] if parte and parte != "-"
                        ) or "-",
                        (largura - 28) / 2,
                    ),
                ],
            ],
            colWidths=[(largura - 28) / 2, (largura - 28) / 2],
        )
        metricas.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))

        elementos: list[Any] = [
            Table(
                [[Paragraph("IMOVEL", estilo_card_kicker), _badge_status(status)]],
                colWidths=[largura - 34 * mm, 34 * mm],
            ),
            Spacer(1, 2.3 * mm),
            _paragrafo_seguro(titulo, estilo_card_titulo),
            Spacer(1, 1.2 * mm),
            _paragrafo_seguro(localizacao, estilo_card_texto),
        ]
        if detalhes:
            elementos.extend([Spacer(1, 1 * mm), _paragrafo_seguro(detalhes, estilo_card_texto)])
        elementos.extend([Spacer(1, 3 * mm), metricas])

        reserva_cliente = _texto_limpo(linha.get("reserva_cliente"))
        if reserva_cliente and reserva_cliente != "-":
            reserva_box = Table(
                [[[Paragraph(f"Reserva ativa para {xml_escape(reserva_cliente)}", estilo_reserva)]]],
                colWidths=[largura - 2],
            )
            reserva_box.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF8EC")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#F7D7A1")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            elementos.extend([Spacer(1, 3 * mm), reserva_box])

        atualizado = _formatar_data_hora_br(linha.get("data_hora_atualizado_em"))
        rodape = " • ".join(
            parte for parte in [
                f"Captacao minima {_formatar_percentual_br(linha.get('percentual_fechamento_minimo'))}" if linha.get("percentual_fechamento_minimo") not in (None, "") else "",
                f"Obra {_formatar_percentual_br(linha.get('percentual_conclusao_obra'))}" if linha.get("percentual_conclusao_obra") not in (None, "") else "",
                f"Atualizado em {atualizado}" if atualizado != "-" else "",
            ] if parte
        )
        if rodape:
            elementos.extend([Spacer(1, 3 * mm), _paragrafo_seguro(rodape, estilo_card_texto)])

        card = Table([[[elementos]]], colWidths=[largura])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), azul_card),
            ("BOX", (0, 0), (-1, -1), 0.8, borda),
            ("LINEABOVE", (0, 0), (-1, 0), 2.2, azul_principal),
            ("LEFTPADDING", (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 11),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 11),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        return card

    destaque_titulo = _valor_texto_exportacao(destaque, {"chave": "titulo", "tipo": "texto"})
    data_emissao = _formatar_data_hora_br(datetime.now())
    status_hero = "EM AJUSTE" if any([
        resumo["reservados"] > 0,
        resumo["vendidos"] > 0,
        resumo["inativos"] > 0,
        metricas["sem_midias"] > 0,
    ]) else "CARTEIRA ATIVA"

    top_row = Table(
        [[_hero_logo_badge(), _hero_status_badge(status_hero)]],
        colWidths=[40 * mm, doc.width - (40 * mm)],
    )
    top_row.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    hero_line = Table([[""]], colWidths=[32 * mm], rowHeights=[1.6 * mm])
    hero_line.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#35D5F7")),
        ("BOX", (0, 0), (-1, -1), 0, colors.white),
    ]))

    hero_media_width = doc.width * 0.54
    hero_metric_gap = 8
    hero_metric_width_total = doc.width - hero_media_width - hero_metric_gap
    hero_metric_half = (hero_metric_width_total - 6) / 2
    hero_metrics = Table(
        [
            [_hero_metric_card("Valor total da carteira", _formatar_moeda_br(metricas["valor_total"]), hero_metric_width_total, destaque=True), ""],
            [
                _hero_metric_card("Ticket medio", _formatar_moeda_br(resumo["ticket_medio"]), hero_metric_half),
                _hero_metric_card("Reservas ativas", _formatar_inteiro(metricas["reservas"]), hero_metric_half),
            ],
            [
                _hero_metric_card("Garantido total", _formatar_moeda_br(metricas["valor_garantido_total"]), hero_metric_half),
                _hero_metric_card("Proxima entrega", _formatar_data_br(metricas["proxima_entrega"]), hero_metric_half),
            ],
        ],
        colWidths=[hero_metric_half, hero_metric_half],
    )
    hero_metrics.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    hero_grid = Table(
        [[_hero_media_card(destaque, hero_media_width, 70 * mm), hero_metrics]],
        colWidths=[hero_media_width, hero_metric_width_total],
    )
    hero_grid.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    hero_conteudo = [
        top_row,
        Spacer(1, 5 * mm),
        Paragraph("Relatorio de imoveis cadastrados", estilo_hero_titulo),
        Spacer(1, 1.2 * mm),
        Paragraph(
            f"{xml_escape(destaque_titulo)} - Emitido em {xml_escape(data_emissao)}",
            estilo_hero_subtitulo,
        ),
        Spacer(1, 1.4 * mm),
        hero_line,
        Spacer(1, 4.5 * mm),
        hero_grid,
    ]

    hero = Table(
        [[hero_conteudo]],
        colWidths=[doc.width],
    )
    hero.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), hero_fundo),
        ("BOX", (0, 0), (-1, -1), 0, hero_fundo),
        ("LEFTPADDING", (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    painel_resumo = Table(
        [[
            [
                Paragraph("FILTROS APLICADOS", estilo_painel_rotulo),
                Spacer(1, 0.8 * mm),
                Paragraph(
                    _descricao_filtros_exportacao(filtros),
                    estilo_painel_texto,
                ),
            ],
            [
                Paragraph("INDICADORES", estilo_painel_rotulo),
                Spacer(1, 0.8 * mm),
                Paragraph(
                    f"Vendidos: {resumo['vendidos']} • Inativos: {resumo['inativos']} • Cidades: {resumo['cidades']} • Formatos: Excel, PDF e CSV",
                    estilo_painel_texto,
                ),
            ],
        ]],
        colWidths=[doc.width * 0.56, doc.width * 0.44],
    )
    painel_resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), azul_claro),
        ("BOX", (0, 0), (-1, -1), 0.7, borda),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    def desenhar_rodape(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(texto_medio)
        canvas.drawString(doc.leftMargin, 7 * mm, "7LM Connect | Exportacao comercial de imoveis")
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 7 * mm, f"Pagina {canvas.getPageNumber()}")
        canvas.restoreState()

    historia: list[Any] = [
        hero,
        Spacer(1, 3 * mm),
        painel_resumo,
        Spacer(1, 3 * mm),
    ]

    if not linhas:
        vazio = Table(
            [[[Paragraph("Nenhum imovel encontrado para os filtros informados.", estilo_vazio)]]],
            colWidths=[doc.width],
        )
        vazio.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), azul_fundo),
            ("BOX", (0, 0), (-1, -1), 0.7, borda),
            ("LEFTPADDING", (0, 0), (-1, -1), 16),
            ("RIGHTPADDING", (0, 0), (-1, -1), 16),
            ("TOPPADDING", (0, 0), (-1, -1), 18),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 18),
        ]))
        historia.append(vazio)
    else:
        largura_card = (doc.width - 8 * mm) / 2
        cards = [_card_imovel(linha, largura_card) for linha in linhas]
        for indice in range(0, len(cards), 2):
            linha_cards = cards[indice:indice + 2]
            if len(linha_cards) == 1:
                placeholder = Table([[""]], colWidths=[largura_card])
                placeholder.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0, colors.white)]))
                linha_cards.append(placeholder)
            grade = Table([linha_cards], colWidths=[largura_card, largura_card])
            grade.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            historia.append(KeepTogether([grade, Spacer(1, 4 * mm)]))

    doc.build(historia, onFirstPage=desenhar_rodape, onLaterPages=desenhar_rodape)
    return buffer.getvalue()


def gerar_pdf_exportacao_imoveis_simulado(imoveis: list[dict[str, Any]], filtros: dict[str, Any]) -> bytes:
    linhas = [_mapear_imovel_para_exportacao(item) for item in imoveis]
    resumo = _resumo_exportacao_imoveis(linhas)
    metricas = _metricas_financeiras_exportacao(linhas)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=9 * mm, rightMargin=9 * mm, topMargin=9 * mm, bottomMargin=9 * mm, title="Relatorio de imoveis cadastrados")
    base = getSampleStyleSheet()
    c_primary, c_dark, c_dark_alt = colors.HexColor("#0669F8"), colors.HexColor("#081222"), colors.HexColor("#102744")
    c_soft, c_border, c_text, c_muted = colors.HexColor("#EAF7FF"), colors.HexColor("#D8E8F4"), colors.HexColor("#172033"), colors.HexColor("#5B6B82")
    s_kicker = ParagraphStyle("PdfSimKicker", parent=base["BodyText"], fontName="Helvetica-Bold", fontSize=7.5, leading=9, textColor=colors.HexColor("#8EEEFF"))
    s_title = ParagraphStyle("PdfSimTitle", parent=base["Heading1"], fontName="Helvetica-Bold", fontSize=20, leading=23, textColor=colors.white)
    s_light = ParagraphStyle("PdfSimLight", parent=base["BodyText"], fontName="Helvetica", fontSize=8.8, leading=11, textColor=colors.HexColor("#DDEEFE"))
    s_label = ParagraphStyle("PdfSimLabel", parent=base["BodyText"], fontName="Helvetica-Bold", fontSize=7, leading=9, textColor=c_primary)
    s_text = ParagraphStyle("PdfSimText", parent=base["BodyText"], fontName="Helvetica", fontSize=8.8, leading=11, textColor=c_text)
    s_section = ParagraphStyle("PdfSimSection", parent=base["Heading2"], fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=c_text)
    s_kpi_l = ParagraphStyle("PdfSimKpiLabel", parent=base["BodyText"], fontName="Helvetica-Bold", fontSize=6.8, leading=8.2, textColor=c_muted)
    s_kpi_v = ParagraphStyle("PdfSimKpiValue", parent=base["BodyText"], fontName="Helvetica-Bold", fontSize=11.2, leading=13, textColor=c_text)
    s_head = ParagraphStyle("PdfSimHead", parent=base["BodyText"], fontName="Helvetica-Bold", fontSize=7, leading=8.2, textColor=colors.HexColor("#627086"))
    s_cell = ParagraphStyle("PdfSimCell", parent=base["BodyText"], fontName="Helvetica", fontSize=8, leading=9.7, textColor=c_text)
    s_muted = ParagraphStyle("PdfSimMuted", parent=base["BodyText"], fontName="Helvetica", fontSize=7.4, leading=8.9, textColor=c_muted)
    s_reserva = ParagraphStyle("PdfSimReserva", parent=base["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=9.7, textColor=colors.HexColor("#9A6400"))
    def p(valor: Any, estilo: ParagraphStyle, fallback: str = "-") -> Paragraph: return Paragraph(xml_escape(_texto_limpo(valor) or fallback).replace("\n", "<br/>"), estilo)
    def secao(kicker: str, titulo: str) -> list[Any]: return [Paragraph(xml_escape(kicker.upper()), s_label), Spacer(1, 1 * mm), Paragraph(xml_escape(titulo), s_section), Spacer(1, 3 * mm)]
    def local(linha: dict[str, Any]) -> str: return " - ".join(parte for parte in [_texto_limpo(linha.get("bairro")), _texto_limpo(linha.get("cidade"))] if parte) or _texto_limpo(linha.get("endereco_formatado")) or "Local nao informado"
    def detalhes(linha: dict[str, Any]) -> str: return " • ".join(parte for parte in [_texto_limpo(linha.get("tipo_imovel")), f"{_formatar_numero_br(linha.get('area_m2'), 2)} m2" if linha.get("area_m2") not in (None, "") else "", f"{_formatar_inteiro(linha.get('quartos'))} quarto(s)" if _inteiro_seguro(linha.get("quartos")) is not None else "", f"{_formatar_inteiro(linha.get('vagas_garagem'))} vaga(s)" if _inteiro_seguro(linha.get("vagas_garagem")) is not None else ""] if parte) or "Sem detalhes adicionais."
    def card(rotulo: str, valor: str, largura: float) -> Table:
        t = Table([[[Paragraph(xml_escape(rotulo.upper()), s_kpi_l), Spacer(1, 1.2 * mm), Paragraph(xml_escape(valor), s_kpi_v)]]], colWidths=[largura])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.white), ("BOX", (0, 0), (-1, -1), 0.7, c_border), ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 10), ("BOTTOMPADDING", (0, 0), (-1, -1), 10)]))
        return t
    def grade(campos: list[tuple[str, str]], colunas: int) -> Table:
        largura = doc.width / colunas; itens = [card(rotulo, valor, largura - 2) for rotulo, valor in campos]
        while len(itens) % colunas != 0:
            vazio = Table([[""]], colWidths=[largura - 2]); vazio.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0, c_soft)])); itens.append(vazio)
        g = Table([itens[i:i + colunas] for i in range(0, len(itens), colunas)], colWidths=[largura] * colunas)
        g.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
        return g
    destaque = linhas[0] if linhas else {}
    hero_bloco = Table([[[Paragraph("UNIDADE EM DESTAQUE", s_kicker), Spacer(1, 1.2 * mm), Paragraph(xml_escape(_valor_texto_exportacao(destaque, {"chave": "titulo", "tipo": "texto"})), s_title), Spacer(1, 1.2 * mm), Paragraph(xml_escape(local(destaque)), s_light), Spacer(1, 1 * mm), Paragraph(xml_escape(detalhes(destaque)), s_light)]]], colWidths=[doc.width])
    hero_bloco.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), c_dark_alt), ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#1F476E")), ("LEFTPADDING", (0, 0), (-1, -1), 12), ("RIGHTPADDING", (0, 0), (-1, -1), 12), ("TOPPADDING", (0, 0), (-1, -1), 12), ("BOTTOMPADDING", (0, 0), (-1, -1), 12)]))
    hero = Table([[[Paragraph("SIMULADOR VISUAL", s_kicker), Spacer(1, 1.2 * mm), Paragraph("Base de imoveis cadastrados", s_title), Spacer(1, 1.1 * mm), Paragraph("Relatorio no padrao visual do simulador, com hierarquia comercial mais clara para apresentar a carteira.", s_light), Spacer(1, 2.5 * mm), Paragraph(f"Gerado em {_formatar_data_hora_br(datetime.now())}<br/>Filtros ativos: {xml_escape(_descricao_filtros_exportacao(filtros))}", s_light), Spacer(1, 3 * mm), hero_bloco, Spacer(1, 3 * mm), grade([("Valor total da carteira", _formatar_moeda_br(metricas["valor_total"])), ("Ticket medio", _formatar_moeda_br(resumo["ticket_medio"])), ("Reservas ativas", str(metricas["reservas"])), ("Proxima entrega", _formatar_data_br(metricas["proxima_entrega"]))], 2)]]], colWidths=[doc.width])
    hero.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), c_dark), ("BOX", (0, 0), (-1, -1), 0.7, c_dark_alt), ("LEFTPADDING", (0, 0), (-1, -1), 14), ("RIGHTPADDING", (0, 0), (-1, -1), 14), ("TOPPADDING", (0, 0), (-1, -1), 14), ("BOTTOMPADDING", (0, 0), (-1, -1), 14)]))
    alertas = []
    if resumo["reservados"] > 0: alertas.append(f"{resumo['reservados']} unidade(s) exigem acompanhamento comercial imediato.")
    if metricas["sem_midias"] > 0: alertas.append(f"{metricas['sem_midias']} unidade(s) ainda nao possuem midias publicadas.")
    if resumo["inativos"] > 0: alertas.append(f"{resumo['inativos']} unidade(s) estao marcadas como inativas e precisam revisao.")
    if resumo["disponiveis"] > 0: alertas.append(f"{resumo['disponiveis']} unidade(s) estao livres para simulacao e negociacao.")
    if not alertas: alertas.append("Nenhum alerta critico foi identificado para os filtros desta exportacao.")
    def rodape(canvas, doc_obj):
        canvas.saveState(); canvas.setFont("Helvetica", 8); canvas.setFillColor(c_muted); canvas.drawString(doc.leftMargin, 6 * mm, "7LM Connect | Exportacao comercial de imoveis"); canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 6 * mm, f"Pagina {canvas.getPageNumber()}"); canvas.restoreState()
    historia: list[Any] = [hero, Spacer(1, 5 * mm)]
    historia.extend(secao("Resumo comercial", "Carteira e filtros"))
    historia.append(Table([[card("Filtros", _descricao_filtros_exportacao(filtros), doc.width * 0.5 - 2), card("Distribuicao", f"Disponiveis {resumo['disponiveis']} • Reservados {resumo['reservados']} • Vendidos {resumo['vendidos']} • Inativos {resumo['inativos']} • Cidades {resumo['cidades']}", doc.width * 0.5 - 2)]], colWidths=[doc.width * 0.5, doc.width * 0.5]))
    historia.append(Spacer(1, 4 * mm))
    historia.extend(secao("Indicadores da operacao", "Resumo financeiro"))
    historia.append(grade([("Valor total", _formatar_moeda_br(metricas["valor_total"])), ("Garantido total", _formatar_moeda_br(metricas["valor_garantido_total"])), ("Ticket medio", _formatar_moeda_br(metricas["valor_medio"])), ("Garantido medio", _formatar_moeda_br(metricas["valor_garantido_medio"])), ("Pre-obra media", _formatar_moeda_br(metricas["pre_obra_media"])), ("Conclusao media", f"{metricas['conclusao_obra_media']:.2f}%".replace(".", ",")), ("Midias totais", str(metricas["total_midias"])), ("Sem midias", str(metricas["sem_midias"]))], 4))
    historia.append(Spacer(1, 4 * mm))
    historia.extend(secao("Orientacao", "Alertas e recomendacoes"))
    historia.append(Table([[[Paragraph(xml_escape("\n".join(f"- {item}" for item in alertas)).replace("\n", "<br/>"), s_text)]]], colWidths=[doc.width]))
    if not linhas:
        historia.append(Spacer(1, 4 * mm)); historia.append(Table([[[Paragraph("Nenhum imovel encontrado para os filtros informados.", s_text)]]], colWidths=[doc.width]))
    else:
        historia.append(PageBreak()); historia.extend(secao("Detalhamento da carteira", "Demonstrativo de imoveis"))
        tabela_linhas: list[list[Any]] = [[Paragraph("Imovel", s_head), Paragraph("Localizacao", s_head), Paragraph("Status", s_head), Paragraph("Valor", s_head), Paragraph("Garantido", s_head), Paragraph("Entrega", s_head), Paragraph("% obra", s_head), Paragraph("Reserva", s_head)]]
        for linha in linhas:
            entrega = " • ".join(parte for parte in [_formatar_data_br(linha.get("data_entrega")) if linha.get("data_entrega") else "", f"{_formatar_inteiro(linha.get('meses_pre_entrega'))} meses" if _inteiro_seguro(linha.get("meses_pre_entrega")) is not None else ""] if parte and parte != "-") or "-"
            reserva_texto = _valor_texto_exportacao(linha, {"chave": "reserva_cliente", "tipo": "texto"})
            tabela_linhas.append([[p(_valor_texto_exportacao(linha, {"chave": "titulo", "tipo": "texto"}), s_cell), Spacer(1, 0.8 * mm), p(detalhes(linha), s_muted)], [p(local(linha), s_cell), Spacer(1, 0.8 * mm), p(_texto_limpo(linha.get("endereco_formatado")) or "Endereco nao informado", s_muted)], p(_valor_texto_exportacao(linha, {"chave": "status", "tipo": "texto"}), s_cell), p(_formatar_moeda_br(linha.get("valor")), s_cell), p(_formatar_moeda_br(linha.get("valor_garantido")), s_cell), p(entrega, s_cell), p(_formatar_percentual_br(linha.get("percentual_conclusao_obra")), s_cell), p(reserva_texto, s_reserva if reserva_texto != "-" else s_cell)])
        tabela = Table(tabela_linhas, colWidths=[38 * mm, 34 * mm, 18 * mm, 23 * mm, 23 * mm, 18 * mm, 16 * mm, 20 * mm], repeatRows=1)
        tabela.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), c_soft), ("BOX", (0, 0), (-1, -1), 0.65, c_border), ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E7EEF5")), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFDFE")])]))
        historia.append(tabela)
    doc.build(historia, onFirstPage=rodape, onLaterPages=rodape)
    return buffer.getvalue()


def gerar_pdf_exportacao_imoveis_simulado(imoveis: list[dict[str, Any]], filtros: dict[str, Any]) -> bytes:
    linhas = [_mapear_imovel_para_exportacao(item) for item in imoveis]
    resumo = _resumo_exportacao_imoveis(linhas)
    metricas = _metricas_financeiras_exportacao(linhas)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=9 * mm,
        bottomMargin=9 * mm,
        title="Relatorio de imoveis cadastrados",
    )

    base = getSampleStyleSheet()
    cor_primaria = colors.HexColor("#0669F8")
    cor_hero = colors.HexColor("#081222")
    cor_hero_secundaria = colors.HexColor("#102744")
    cor_superficie = colors.HexColor("#EAF7FF")
    cor_borda = colors.HexColor("#D8E8F4")
    cor_texto = colors.HexColor("#172033")
    cor_texto_suave = colors.HexColor("#5B6B82")
    cor_texto_hero = colors.HexColor("#DDEEFE")
    hero_padding_horizontal = 12
    hero_padding_vertical = 10
    hero_largura_util = doc.width - (hero_padding_horizontal * 2)

    estilo_hero_kicker = ParagraphStyle(
        "PdfImoveisHeroKickerV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        textColor=colors.HexColor("#8EEEFF"),
    )
    estilo_hero_titulo = ParagraphStyle(
        "PdfImoveisHeroTituloV2",
        parent=base["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=20,
        textColor=colors.white,
    )
    estilo_hero_texto = ParagraphStyle(
        "PdfImoveisHeroTextoV2",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.1,
        leading=9.6,
        textColor=cor_texto_hero,
    )
    estilo_secao_rotulo = ParagraphStyle(
        "PdfImoveisSecaoRotuloV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=9,
        textColor=cor_primaria,
    )
    estilo_secao_titulo = ParagraphStyle(
        "PdfImoveisSecaoTituloV2",
        parent=base["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12.5,
        leading=14,
        textColor=cor_texto,
    )
    estilo_corpo = ParagraphStyle(
        "PdfImoveisCorpoV2",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=9.4,
        textColor=cor_texto,
    )
    estilo_corpo_suave = ParagraphStyle(
        "PdfImoveisCorpoSuaveV2",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=6.8,
        leading=8,
        textColor=cor_texto_suave,
    )
    estilo_kpi_rotulo = ParagraphStyle(
        "PdfImoveisKpiRotuloV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=7.5,
        textColor=cor_texto_suave,
    )
    estilo_kpi_valor = ParagraphStyle(
        "PdfImoveisKpiValorV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=10.2,
        leading=11.5,
        textColor=cor_texto,
    )
    estilo_tabela_head = ParagraphStyle(
        "PdfImoveisTabelaHeadV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=6.4,
        leading=7.2,
        textColor=colors.HexColor("#627086"),
    )
    estilo_tabela_titulo = ParagraphStyle(
        "PdfImoveisTabelaTituloV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=8.1,
        textColor=cor_texto,
    )
    estilo_tabela_celula = ParagraphStyle(
        "PdfImoveisTabelaCelulaV2",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=7.1,
        leading=8.1,
        textColor=cor_texto,
    )
    estilo_tabela_numero = ParagraphStyle(
        "PdfImoveisTabelaNumeroV2",
        parent=estilo_tabela_celula,
        alignment=TA_RIGHT,
    )
    estilo_tabela_centro = ParagraphStyle(
        "PdfImoveisTabelaCentroV2",
        parent=estilo_tabela_celula,
        alignment=TA_LEFT,
    )
    estilo_reserva = ParagraphStyle(
        "PdfImoveisReservaV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.1,
        leading=8.1,
        textColor=colors.HexColor("#9A6400"),
    )
    estilo_logo_texto = ParagraphStyle(
        "PdfImoveisLogoTextoV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=18,
        textColor=cor_texto,
    )
    estilo_hero_subtitulo = ParagraphStyle(
        "PdfImoveisHeroSubtituloV2",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.9,
        leading=10.4,
        textColor=cor_texto_hero,
    )
    estilo_hero_badge = ParagraphStyle(
        "PdfImoveisHeroBadgeV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.3,
        leading=8.4,
        textColor=colors.white,
        alignment=TA_RIGHT,
    )
    estilo_hero_metric_rotulo = ParagraphStyle(
        "PdfImoveisHeroMetricRotuloV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=6.5,
        leading=7.7,
        textColor=colors.HexColor("#CBEAFA"),
    )
    estilo_hero_metric_valor = ParagraphStyle(
        "PdfImoveisHeroMetricValorV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11.2,
        leading=12.8,
        textColor=colors.white,
    )
    estilo_hero_media_titulo = ParagraphStyle(
        "PdfImoveisHeroMidiaTituloV2",
        parent=base["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=12.6,
        textColor=colors.white,
    )
    estilo_hero_media_texto = ParagraphStyle(
        "PdfImoveisHeroMidiaTextoV2",
        parent=base["BodyText"],
        fontName="Helvetica",
        fontSize=8.1,
        leading=9.6,
        textColor=colors.HexColor("#DDEEFE"),
    )

    def paragrafo_seguro(
        valor: Any,
        estilo: ParagraphStyle,
        fallback: str = "-",
        *,
        preservar_quebras: bool = True,
    ) -> Paragraph:
        texto = _sanitizar_texto_exportacao(valor, preservar_quebras=preservar_quebras) or fallback
        texto = xml_escape(texto).replace("\n", "<br/>")
        return Paragraph(texto, estilo)

    def montar_secao(rotulo: str, titulo: str) -> list[Any]:
        return [
            paragrafo_seguro(rotulo.upper(), estilo_secao_rotulo, preservar_quebras=False),
            Spacer(1, 0.6 * mm),
            paragrafo_seguro(titulo, estilo_secao_titulo, preservar_quebras=False),
            Spacer(1, 2 * mm),
        ]

    def descrever_localizacao(linha: dict[str, Any]) -> str:
        partes = [
            _sanitizar_texto_exportacao(linha.get("bairro"), preservar_quebras=False),
            _sanitizar_texto_exportacao(linha.get("cidade"), preservar_quebras=False),
        ]
        return (
            " - ".join(parte for parte in partes if parte)
            or _sanitizar_texto_exportacao(linha.get("endereco_formatado"), preservar_quebras=False)
            or "Local nao informado"
        )

    def descrever_detalhes(linha: dict[str, Any]) -> str:
        partes = [
            _sanitizar_texto_exportacao(linha.get("tipo_imovel"), preservar_quebras=False),
            f"{_formatar_numero_br(linha.get('area_m2'), 2)} m2" if linha.get("area_m2") not in (None, "") else "",
            (
                f"{_formatar_inteiro(linha.get('quartos'))} qtos"
                if _inteiro_seguro(linha.get("quartos")) is not None
                else ""
            ),
            (
                f"{_formatar_inteiro(linha.get('vagas_garagem'))} vaga"
                if _inteiro_seguro(linha.get("vagas_garagem")) is not None
                else ""
            ),
        ]
        return " | ".join(parte for parte in partes if parte) or "Sem detalhes adicionais."

    def descrever_endereco_compacto(linha: dict[str, Any]) -> str:
        return (
            _resumir_texto_exportacao(linha.get("endereco_formatado"), 74)
            or "Endereco nao informado"
        )

    def descrever_entrega_compacta(linha: dict[str, Any]) -> str:
        meses = _inteiro_seguro(linha.get("meses_pre_entrega"))
        data = _formatar_data_br(linha.get("data_entrega"))
        if meses is not None and data != "-":
            return f"{meses} m | {data}"
        if meses is not None:
            return f"{meses} meses"
        return data

    def criar_card_info(
        rotulo: str,
        valor: str,
        largura: float,
        *,
        padding_horizontal: float = 8,
        padding_vertical: float = 7,
    ) -> Table:
        card = Table(
            [[[
                paragrafo_seguro(rotulo.upper(), estilo_kpi_rotulo, preservar_quebras=False),
                Spacer(1, 0.7 * mm),
                paragrafo_seguro(valor, estilo_kpi_valor),
            ]]],
            colWidths=[largura],
        )
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.7, cor_borda),
            ("LEFTPADDING", (0, 0), (-1, -1), padding_horizontal),
            ("RIGHTPADDING", (0, 0), (-1, -1), padding_horizontal),
            ("TOPPADDING", (0, 0), (-1, -1), padding_vertical),
            ("BOTTOMPADDING", (0, 0), (-1, -1), padding_vertical),
        ]))
        return card

    def criar_grade_cards(
        campos: list[tuple[str, str]],
        colunas: int,
        *,
        largura_total: float | None = None,
        padding_horizontal: float = 8,
        padding_vertical: float = 7,
    ) -> Table:
        largura_base = largura_total or doc.width
        largura_coluna = largura_base / colunas
        cards = [
            criar_card_info(
                rotulo,
                valor,
                largura_coluna - 2,
                padding_horizontal=padding_horizontal,
                padding_vertical=padding_vertical,
            )
            for rotulo, valor in campos
        ]
        while len(cards) % colunas != 0:
            vazio = Table([[""]], colWidths=[largura_coluna - 2])
            vazio.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0, colors.white)]))
            cards.append(vazio)
        grade = Table(
            [cards[indice:indice + colunas] for indice in range(0, len(cards), colunas)],
            colWidths=[largura_coluna] * colunas,
        )
        grade.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        return grade

    def criar_badge_logo_hero() -> Table:
        logo_path = (DIRETORIO_PUBLICO / "assets" / "7lm_logo.png").resolve()
        conteudo: Any
        if logo_path.is_file():
            try:
                ImageReader(str(logo_path))
                conteudo = PlatypusImage(
                    str(logo_path),
                    width=24 * mm,
                    height=10.5 * mm,
                    kind="proportional",
                )
            except Exception:
                conteudo = Paragraph("7LM", estilo_logo_texto)
        else:
            conteudo = Paragraph("7LM", estilo_logo_texto)

        badge = Table([[conteudo]], colWidths=[30 * mm], rowHeights=[13.5 * mm])
        badge.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#E5F0F7")),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return badge

    def criar_badge_status_hero(texto: str, largura: float = 28 * mm) -> Table:
        pill = Table(
            [[Paragraph(xml_escape(_texto_limpo(texto).upper() or "CARTEIRA"), estilo_hero_badge)]],
            colWidths=[largura],
            rowHeights=[11 * mm],
        )
        pill.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#2A7394")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#70A9C5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return pill

    def criar_card_metric_hero(rotulo: str, valor: str, largura: float, *, destaque: bool = False) -> Table:
        cor_fundo_metric = colors.HexColor("#2A7394" if destaque else "#245F80")
        cor_borda_metric = colors.HexColor("#5B93AF" if destaque else "#4A86A5")
        card = Table(
            [[[Paragraph(xml_escape(rotulo.upper()), estilo_hero_metric_rotulo), Spacer(1, 1.1 * mm), Paragraph(xml_escape(valor), estilo_hero_metric_valor)]]],
            colWidths=[largura],
        )
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), cor_fundo_metric),
            ("BOX", (0, 0), (-1, -1), 0.7, cor_borda_metric),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        return card

    def criar_card_midia_hero(linha: dict[str, Any], largura: float, altura: float) -> Table:
        caminho_midia = _resolver_arquivo_publico_local(linha.get("foto_principal"))
        destaque_titulo = _valor_texto_exportacao(linha, {"chave": "titulo", "tipo": "texto"})
        destaque_local = _resumir_texto_exportacao(descrever_localizacao(linha), 78) or "Local nao informado"
        destaque_detalhes = _resumir_texto_exportacao(descrever_detalhes(linha), 84) or "Sem detalhes adicionais"
        conteudo: Any = None

        if caminho_midia is not None:
            try:
                ImageReader(str(caminho_midia))
                conteudo = PlatypusImage(
                    str(caminho_midia),
                    width=max(largura - 16, 10),
                    height=max(altura - 16, 10),
                    kind="proportional",
                )
            except Exception:
                conteudo = None

        if conteudo is None:
            conteudo = [
                Paragraph("UNIDADE EM DESTAQUE", estilo_hero_kicker),
                Spacer(1, 2 * mm),
                Paragraph(xml_escape(destaque_titulo or "Carteira comercial"), estilo_hero_media_titulo),
                Spacer(1, 1.4 * mm),
                Paragraph(xml_escape(destaque_local), estilo_hero_media_texto),
                Spacer(1, 1 * mm),
                Paragraph(xml_escape(destaque_detalhes), estilo_hero_media_texto),
            ]

        card = Table([[conteudo]], colWidths=[largura], rowHeights=[altura])
        card.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#294768")),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#4E6F94")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return card

    destaque = linhas[0] if linhas else {}
    destaque_titulo = _valor_texto_exportacao(destaque, {"chave": "titulo", "tipo": "texto"})
    data_emissao = _formatar_data_hora_br(datetime.now())
    status_hero = (
        "EM AJUSTE"
        if any([
            resumo["reservados"] > 0,
            resumo["vendidos"] > 0,
            resumo["inativos"] > 0,
            metricas["sem_midias"] > 0,
        ])
        else "CARTEIRA ATIVA"
    )

    top_row = Table(
        [[criar_badge_logo_hero(), criar_badge_status_hero(status_hero)]],
        colWidths=[40 * mm, hero_largura_util - (40 * mm)],
    )
    top_row.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    hero_line = Table([[""]], colWidths=[32 * mm], rowHeights=[1.6 * mm])
    hero_line.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#35D5F7")),
        ("BOX", (0, 0), (-1, -1), 0, colors.white),
    ]))

    hero_media_width = hero_largura_util * 0.56
    hero_metric_width_total = hero_largura_util - hero_media_width
    hero_metric_half = (hero_metric_width_total - 6) / 2
    hero_metrics = Table(
        [
            [criar_card_metric_hero("Valor total da carteira", _formatar_moeda_br(metricas["valor_total"]), hero_metric_width_total, destaque=True), ""],
            [
                criar_card_metric_hero("Garantido total", _formatar_moeda_br(metricas["valor_garantido_total"]), hero_metric_half),
                criar_card_metric_hero("Ticket medio", _formatar_moeda_br(resumo["ticket_medio"]), hero_metric_half),
            ],
            [
                criar_card_metric_hero("Reservas ativas", _formatar_inteiro(metricas["reservas"]), hero_metric_half),
                criar_card_metric_hero("Proxima entrega", _formatar_data_br(metricas["proxima_entrega"]), hero_metric_half),
            ],
        ],
        colWidths=[hero_metric_half, hero_metric_half],
    )
    hero_metrics.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    hero_grid = Table(
        [[criar_card_midia_hero(destaque, hero_media_width, 68 * mm), hero_metrics]],
        colWidths=[hero_media_width, hero_metric_width_total],
    )
    hero_grid.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    hero_conteudo = [
        top_row,
        Spacer(1, 4.8 * mm),
        paragrafo_seguro("Base de imoveis cadastrados", estilo_hero_titulo, preservar_quebras=False),
        Spacer(1, 1.1 * mm),
        paragrafo_seguro(
            f"{_resumir_texto_exportacao(destaque_titulo, 58) or 'Carteira comercial'} - Emitido em {data_emissao}",
            estilo_hero_subtitulo,
            preservar_quebras=False,
        ),
        Spacer(1, 1.6 * mm),
        hero_line,
        Spacer(1, 4.3 * mm),
        hero_grid,
    ]

    hero = Table([[hero_conteudo]], colWidths=[doc.width])
    hero.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), cor_hero),
        ("BOX", (0, 0), (-1, -1), 0.7, cor_hero_secundaria),
        ("LEFTPADDING", (0, 0), (-1, -1), hero_padding_horizontal),
        ("RIGHTPADDING", (0, 0), (-1, -1), hero_padding_horizontal),
        ("TOPPADDING", (0, 0), (-1, -1), hero_padding_vertical + 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), hero_padding_vertical + 2),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    distribuicao = (
        f"Disponiveis {resumo['disponiveis']} | "
        f"Reservados {resumo['reservados']} | "
        f"Vendidos {resumo['vendidos']} | "
        f"Inativos {resumo['inativos']} | "
        f"Cidades {resumo['cidades']}"
    )

    alertas: list[str] = []
    if resumo["reservados"] > 0:
        alertas.append(f"{resumo['reservados']} unidade(s) exigem acompanhamento comercial imediato.")
    if metricas["sem_midias"] > 0:
        alertas.append(f"{metricas['sem_midias']} unidade(s) ainda nao possuem midias publicadas.")
    if resumo["inativos"] > 0:
        alertas.append(f"{resumo['inativos']} unidade(s) estao marcadas como inativas e precisam de revisao.")
    if resumo["disponiveis"] > 0:
        alertas.append(f"{resumo['disponiveis']} unidade(s) estao livres para simulacao e negociacao.")
    if not alertas:
        alertas.append("Nenhum alerta critico foi identificado para os filtros desta exportacao.")

    def rodape(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(cor_texto_suave)
        canvas.drawString(doc.leftMargin, 6 * mm, "7LM Connect | Exportacao comercial de imoveis")
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 6 * mm, f"Pagina {canvas.getPageNumber()}")
        canvas.restoreState()

    historia: list[Any] = [hero, Spacer(1, 3.5 * mm)]
    historia.extend(montar_secao("Resumo comercial", "Carteira e filtros"))
    historia.append(
        Table(
            [[
                criar_card_info("Filtros", _descricao_filtros_exportacao(filtros), doc.width * 0.5 - 2),
                criar_card_info("Distribuicao", distribuicao, doc.width * 0.5 - 2),
            ]],
            colWidths=[doc.width * 0.5, doc.width * 0.5],
        )
    )
    historia.append(Spacer(1, 3 * mm))
    historia.extend(montar_secao("Indicadores da operacao", "Resumo financeiro"))
    historia.append(
        criar_grade_cards(
            [
                ("Valor total", _formatar_moeda_br(metricas["valor_total"])),
                ("Garantido total", _formatar_moeda_br(metricas["valor_garantido_total"])),
                ("Ticket medio", _formatar_moeda_br(metricas["valor_medio"])),
                ("Garantido medio", _formatar_moeda_br(metricas["valor_garantido_medio"])),
                ("Pre-obra media", _formatar_moeda_br(metricas["pre_obra_media"])),
                ("Conclusao media", _formatar_percentual_br(metricas["conclusao_obra_media"])),
                ("Midias totais", _formatar_inteiro(metricas["total_midias"])),
                ("Sem midias", _formatar_inteiro(metricas["sem_midias"])),
            ],
            4,
            padding_horizontal=7,
            padding_vertical=6,
        )
    )
    historia.append(Spacer(1, 3 * mm))
    historia.extend(montar_secao("Orientacao", "Alertas e recomendacoes"))
    historia.append(
        Table(
            [[[
                Paragraph(
                    xml_escape("\n".join(f"- {item}" for item in alertas)).replace("\n", "<br/>"),
                    estilo_corpo,
                )
            ]]],
            colWidths=[doc.width],
        )
    )

    if not linhas:
        historia.append(Spacer(1, 3 * mm))
        vazio = Table(
            [[[
                Paragraph("Nenhum imovel encontrado para os filtros informados.", estilo_corpo)
            ]]],
            colWidths=[doc.width],
        )
        vazio.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.65, cor_borda),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]))
        historia.append(vazio)
    else:
        historia.append(Spacer(1, 3 * mm))
        historia.extend(montar_secao("Detalhamento da carteira", "Demonstrativo de imoveis"))
        tabela_linhas: list[list[Any]] = [[
            Paragraph("Imovel", estilo_tabela_head),
            Paragraph("Localizacao", estilo_tabela_head),
            Paragraph("Status", estilo_tabela_head),
            Paragraph("Valor", estilo_tabela_head),
            Paragraph("Garantido", estilo_tabela_head),
            Paragraph("Entrega", estilo_tabela_head),
            Paragraph("% obra", estilo_tabela_head),
            Paragraph("Reserva", estilo_tabela_head),
        ]]

        for linha in linhas:
            titulo = _resumir_texto_exportacao(
                _valor_texto_exportacao(linha, {"chave": "titulo", "tipo": "texto"}),
                54,
            )
            detalhes = _resumir_texto_exportacao(descrever_detalhes(linha), 58)
            localizacao = _resumir_texto_exportacao(descrever_localizacao(linha), 58)
            endereco = descrever_endereco_compacto(linha)
            entrega = descrever_entrega_compacta(linha)
            reserva_texto = _resumir_texto_exportacao(
                _valor_texto_exportacao(linha, {"chave": "reserva_cliente", "tipo": "texto"}),
                28,
            ) or "-"
            tabela_linhas.append([
                [
                    paragrafo_seguro(
                        titulo or "-",
                        estilo_tabela_titulo,
                        preservar_quebras=False,
                    ),
                    Spacer(1, 0.45 * mm),
                    paragrafo_seguro(detalhes or "Sem detalhes adicionais.", estilo_corpo_suave, preservar_quebras=False),
                ],
                [
                    paragrafo_seguro(localizacao or "Local nao informado", estilo_tabela_titulo, preservar_quebras=False),
                    Spacer(1, 0.45 * mm),
                    paragrafo_seguro(
                        endereco,
                        estilo_corpo_suave,
                        preservar_quebras=False,
                    ),
                ],
                paragrafo_seguro(
                    _valor_texto_exportacao(linha, {"chave": "status", "tipo": "texto"}),
                    estilo_tabela_centro,
                    preservar_quebras=False,
                ),
                paragrafo_seguro(_formatar_moeda_br(linha.get("valor")), estilo_tabela_numero, preservar_quebras=False),
                paragrafo_seguro(_formatar_moeda_br(linha.get("valor_garantido")), estilo_tabela_numero, preservar_quebras=False),
                paragrafo_seguro(entrega, estilo_tabela_centro, preservar_quebras=False),
                paragrafo_seguro(_formatar_percentual_br(linha.get("percentual_conclusao_obra")), estilo_tabela_numero, preservar_quebras=False),
                paragrafo_seguro(
                    reserva_texto,
                    estilo_reserva if reserva_texto != "-" else estilo_tabela_centro,
                    preservar_quebras=False,
                ),
            ])

        tabela = Table(
            tabela_linhas,
            colWidths=[45 * mm, 47 * mm, 16 * mm, 19 * mm, 20 * mm, 15 * mm, 12 * mm, 16 * mm],
            repeatRows=1,
        )
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), cor_superficie),
            ("BOX", (0, 0), (-1, -1), 0.65, cor_borda),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E7EEF5")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3.2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3.2),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFDFE")]),
        ]))
        historia.append(tabela)

    doc.build(historia, onFirstPage=rodape, onLaterPages=rodape)
    return buffer.getvalue()


def gerar_arquivo_exportacao_imoveis(formato: str, imoveis: list[dict[str, Any]], filtros: dict[str, Any]) -> tuple[bytes, str, str]:
    formato_normalizado = _texto_limpo(formato).lower()
    if formato_normalizado not in {"csv", "xlsx", "pdf"}:
        raise HTTPException(status_code=400, detail="Formato de exportacao invalido. Use csv, xlsx ou pdf.")

    if formato_normalizado == "csv":
        return (
            gerar_csv_exportacao_imoveis(imoveis),
            "text/csv; charset=utf-8",
            _nome_arquivo_exportacao_imoveis("csv"),
        )

    if formato_normalizado == "xlsx":
        if Workbook is None:
            return (
                gerar_xlsx_exportacao_imoveis_fallback(imoveis, filtros),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                _nome_arquivo_exportacao_imoveis("xlsx"),
            )
        try:
            return (
                gerar_xlsx_exportacao_imoveis(imoveis, filtros),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                _nome_arquivo_exportacao_imoveis("xlsx"),
            )
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail="Falha ao gerar o arquivo Excel.") from exc

    return (
        gerar_pdf_exportacao_imoveis_simulado(imoveis, filtros),
        "application/pdf",
        _nome_arquivo_exportacao_imoveis("pdf"),
    )


def _normalizar_payload_evolucao_obra(payload: dict[str, Any]) -> dict[str, Any]:
    percentual = normalizar_decimal_intervalo(
        payload.get("percentual_conclusao_obra"),
        "percentual de conclusao da obra",
        Decimal("0"),
        Decimal("100"),
        Decimal("0"),
    )
    data_referencia = payload.get("data_referencia") or date.today()
    observacoes = _texto_limpo(payload.get("observacoes"))[:500] or None
    return {
        "percentual_conclusao_obra": percentual,
        "data_referencia": data_referencia,
        "observacoes": observacoes,
    }


def _normalizar_texto_endereco_lote(valor: Any) -> str:
    texto = str(valor or "").strip()
    if not texto:
        return ""
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(caractere for caractere in texto if unicodedata.category(caractere) != "Mn")
    texto = re.sub(r"[^a-zA-Z0-9]+", " ", texto).lower().strip()
    return re.sub(r"\s+", " ", texto)


def _cep_normalizado(valor: Any) -> str:
    return re.sub(r"\D+", "", str(valor or ""))


def _base_endereco_para_lote(valor: Any) -> str:
    texto = str(valor or "").strip()
    if not texto:
        return ""

    partes = [parte.strip() for parte in re.split(r"\s*,\s*", texto) if parte.strip()]
    partes_base: list[str] = []
    for parte in partes:
        parte_normalizada = _normalizar_texto_endereco_lote(parte)
        if re.match(r"^(bloco|pavimento|unidade|apartamento|apto)\b", parte_normalizada):
            continue
        partes_base.append(parte)

    if not partes_base and partes:
        partes_base = [partes[0]]

    return _normalizar_texto_endereco_lote(", ".join(partes_base))


def _dados_endereco_lote(registro: dict[str, Any]) -> dict[str, str]:
    return {
        "base": _base_endereco_para_lote(registro.get("endereco")),
        "cidade": _normalizar_texto_endereco_lote(registro.get("cidade")),
        "bairro": _normalizar_texto_endereco_lote(registro.get("bairro")),
        "estado": _normalizar_texto_endereco_lote(registro.get("estado")),
        "cep": _cep_normalizado(registro.get("cep")),
    }


def _mesmo_endereco_lote(referencia: dict[str, str], candidato: dict[str, Any]) -> bool:
    dados = _dados_endereco_lote(candidato)
    if not referencia["base"] or dados["base"] != referencia["base"]:
        return False
    for chave in ("cidade", "bairro", "estado"):
        if referencia[chave] and dados[chave] and referencia[chave] != dados[chave]:
            return False
    if referencia["cep"] and dados["cep"] and referencia["cep"] != dados["cep"]:
        return False
    return True


async def registrar_evolucao_obra(
    conexao,
    esquema: str,
    identificador_imovel: str,
    payload: dict[str, Any],
    *,
    registrado_por: str | None = None,
):
    dados = _normalizar_payload_evolucao_obra(payload)
    registro = await inserir_evolucao_obra(
        conexao,
        esquema,
        identificador_imovel=identificador_imovel,
        percentual_conclusao_obra=dados["percentual_conclusao_obra"],
        data_referencia=dados["data_referencia"],
        observacoes=dados["observacoes"],
        registrado_por=registrado_por,
    )
    imovel = await sincronizar_percentual_conclusao_obra_atual(conexao, esquema, identificador_imovel)
    return registro, imovel


async def registrar_evolucao_obra_lote_mesmo_endereco(
    conexao,
    esquema: str,
    identificador_imovel: str,
    payload: dict[str, Any],
    *,
    registrado_por: str | None = None,
):
    referencia = await buscar_imovel_por_id(conexao, esquema, identificador_imovel)
    if not referencia:
        return None

    dados_referencia = _dados_endereco_lote(referencia)
    if not dados_referencia["base"]:
        raise HTTPException(
            status_code=400,
            detail="Este imovel nao possui endereco suficiente para atualizacao em lote.",
        )

    dados = _normalizar_payload_evolucao_obra(payload)
    candidatos = await listar_imoveis_candidatos_mesmo_endereco(
        conexao,
        esquema,
        cep_normalizado=dados_referencia["cep"],
        cidade=referencia.get("cidade"),
        bairro=referencia.get("bairro"),
        estado=referencia.get("estado"),
    )

    alvos = [item for item in candidatos if _mesmo_endereco_lote(dados_referencia, item)]
    if not any(str(item.get("identificador_imovel")) == str(identificador_imovel) for item in alvos):
        alvos.append(referencia)

    registros = []
    imoveis_atualizados = []
    imovel_referencia_atualizado = None
    for alvo in alvos:
        alvo_id = str(alvo.get("identificador_imovel") or "")
        if not alvo_id:
            continue
        registro = await inserir_evolucao_obra(
            conexao,
            esquema,
            identificador_imovel=alvo_id,
            percentual_conclusao_obra=dados["percentual_conclusao_obra"],
            data_referencia=dados["data_referencia"],
            observacoes=dados["observacoes"],
            registrado_por=registrado_por,
        )
        imovel = await sincronizar_percentual_conclusao_obra_atual(conexao, esquema, alvo_id)
        registros.append(registro)
        if imovel:
            imoveis_atualizados.append(imovel)
            if alvo_id == str(identificador_imovel):
                imovel_referencia_atualizado = imovel

    return {
        "registros": registros,
        "imoveis": imoveis_atualizados,
        "imovel_referencia": imovel_referencia_atualizado or referencia,
        "total_atualizados": len(registros),
    }


async def _registrar_evolucao_obra_se_necessario(
    conexao,
    esquema: str,
    identificador_imovel: str,
    percentual_conclusao_obra: Decimal,
    *,
    registrado_por: str | None = None,
    observacoes: str | None = None,
) -> bool:
    ultima = await obter_ultima_evolucao_obra(conexao, esquema, identificador_imovel)
    if ultima and Decimal(str(ultima.get("percentual_conclusao_obra") or 0)) == percentual_conclusao_obra:
        return False

    await inserir_evolucao_obra(
        conexao,
        esquema,
        identificador_imovel=identificador_imovel,
        percentual_conclusao_obra=percentual_conclusao_obra,
        data_referencia=date.today(),
        observacoes=observacoes,
        registrado_por=registrado_por,
    )
    await sincronizar_percentual_conclusao_obra_atual(conexao, esquema, identificador_imovel)
    return True


async def criar_imovel(conexao, esquema: str, payload: dict[str, Any], registrado_por: str | None = None):
    dados_normalizados = normalizar_payload_imovel(payload)
    imovel = await criar_imovel_repositorio(conexao, esquema, dados_normalizados)
    await _registrar_evolucao_obra_se_necessario(
        conexao,
        esquema,
        imovel["identificador_imovel"],
        dados_normalizados["percentual_conclusao_obra"],
        registrado_por=registrado_por,
        observacoes="Registro inicial da obra.",
    )
    return await sincronizar_percentual_conclusao_obra_atual(conexao, esquema, imovel["identificador_imovel"]) or imovel


async def atualizar_imovel(
    conexao,
    esquema: str,
    identificador_imovel: str,
    payload: dict[str, Any],
    registrado_por: str | None = None,
):
    dados_normalizados = normalizar_payload_imovel(payload)
    imovel = await atualizar_imovel_repositorio(conexao, esquema, identificador_imovel, dados_normalizados)
    if not imovel:
        return None

    historico_criado = await _registrar_evolucao_obra_se_necessario(
        conexao,
        esquema,
        identificador_imovel,
        dados_normalizados["percentual_conclusao_obra"],
        registrado_por=registrado_por,
        observacoes="Atualizacao pelo cadastro do imovel.",
    )
    if historico_criado:
        return await sincronizar_percentual_conclusao_obra_atual(conexao, esquema, identificador_imovel) or imovel
    return imovel


def _primeiro_texto_preenchido(*valores: Any) -> str:
    for valor in valores:
        texto = valor_planilha_como_texto(valor)
        if texto:
            return texto
    return ""


def _primeiro_valor_preenchido(*valores: Any) -> Any:
    for valor in valores:
        if valor is None:
            continue
        if isinstance(valor, str) and not valor.strip():
            continue
        return valor
    return None


def _formatar_medida(valor: Any) -> str:
    texto = valor_planilha_como_texto(valor)
    if not texto:
        return ""
    return texto.replace(".", ",")


def _gerar_titulo_importacao(
    *,
    tipo_imovel: str,
    unidade: str,
    bloco: str,
    posicao: str,
    linha_planilha: int,
) -> str:
    base = valor_planilha_como_texto(tipo_imovel) or "Imovel"

    if unidade and bloco:
        return f"{base} {unidade} - Bloco {bloco}"
    if unidade:
        return f"{base} {unidade}"
    if posicao:
        return f"{base} {posicao}"
    return f"{base} importado {linha_planilha}"


def _gerar_endereco_importacao(*, endereco_base: str, bloco: str, pavimento: str, unidade: str) -> str | None:
    partes = []
    if endereco_base:
        partes.append(endereco_base)
    if bloco:
        partes.append(f"Bloco {bloco}")
    if pavimento:
        partes.append(f"Pavimento {pavimento}")
    if unidade:
        partes.append(f"Unidade {unidade}")

    texto = ", ".join(parte for parte in partes if parte)
    return texto or None


def _gerar_descricao_importacao(
    *,
    descricao_manual: str,
    localidade: str,
    unidade: str,
    pavimento: str,
    bloco: str,
    orientacao: str,
    posicao: str,
    suites: Any,
    vaga_vinculada: str,
    area_real: Any,
    area_equivalente: Any,
    area_total: Any,
) -> str | None:
    detalhes = ["Importado em lote via planilha Excel."]
    pares = [
        ("Localidade", localidade),
        ("Unidade", unidade),
        ("Pavimento", pavimento),
        ("Bloco", bloco),
        ("Orientacao", orientacao),
        ("Posicao", posicao),
        ("Suites", valor_planilha_como_texto(suites)),
        ("Vaga vinculada", vaga_vinculada),
        ("Area real", f"{_formatar_medida(area_real)} m2" if valor_planilha_como_texto(area_real) else ""),
        (
            "Area equivalente",
            f"{_formatar_medida(area_equivalente)} m2" if valor_planilha_como_texto(area_equivalente) else "",
        ),
        ("Area total construida", f"{_formatar_medida(area_total)} m2" if valor_planilha_como_texto(area_total) else ""),
    ]

    for rotulo, valor in pares:
        texto = valor_planilha_como_texto(valor)
        if texto:
            detalhes.append(f"{rotulo}: {texto}.")

    bloco_descritivo = " ".join(detalhes).strip()
    if descricao_manual and bloco_descritivo:
        return f"{descricao_manual}\n\n{bloco_descritivo}"
    return descricao_manual or bloco_descritivo or None


def _normalizar_chave_importacao(*valores: Any) -> str:
    partes = [normalizar_nome_para_arquivo(valor_planilha_como_texto(valor)) for valor in valores if valor_planilha_como_texto(valor)]
    return "::".join(partes)


def _montar_payload_importacao_linha(
    linha: list[Any] | tuple[Any, ...],
    colunas: dict[str, int],
    contexto: dict[str, Any],
    linha_planilha: int,
) -> dict[str, Any]:
    titulo_manual = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "titulo"))
    descricao_manual = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "descricao"))
    tipo_imovel = _primeiro_texto_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "tipo_imovel"),
        contexto.get("tipo_imovel_padrao"),
        "Apartamento",
    )
    unidade = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "unidade_autonoma"))
    bloco = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "bloco"))
    posicao = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "posicao"))
    pavimento = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "pavimento"))
    localidade = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "localidade"))
    orientacao = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "orientacao"))
    vaga_vinculada = _primeiro_texto_preenchido(obter_valor_coluna_planilha(linha, colunas, "vaga_vinculada"))

    titulo = titulo_manual or _gerar_titulo_importacao(
        tipo_imovel=tipo_imovel,
        unidade=unidade,
        bloco=bloco,
        posicao=posicao,
        linha_planilha=linha_planilha,
    )
    endereco = _primeiro_texto_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "endereco"),
        _gerar_endereco_importacao(
            endereco_base=valor_planilha_como_texto(contexto.get("endereco_base")),
            bloco=bloco,
            pavimento=pavimento,
            unidade=unidade,
        ),
    )

    cidade = _primeiro_texto_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "cidade"),
        contexto.get("cidade_padrao"),
    )
    bairro = _primeiro_texto_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "bairro"),
        contexto.get("bairro_padrao"),
    )
    estado = _primeiro_texto_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "estado"),
        contexto.get("estado_padrao"),
    )
    cep = _primeiro_texto_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "cep"),
        contexto.get("cep_padrao"),
    )
    status = _primeiro_texto_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "status"),
        contexto.get("status_padrao"),
        "Disponivel",
    )

    valor = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "valor"),
        obter_valor_coluna_planilha(linha, colunas, "vgv"),
    )
    quartos = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "quartos"),
        obter_valor_coluna_planilha(linha, colunas, "dormitorios"),
    )
    banheiros = _primeiro_valor_preenchido(obter_valor_coluna_planilha(linha, colunas, "banheiros"))
    vagas = _primeiro_valor_preenchido(obter_valor_coluna_planilha(linha, colunas, "vagas_garagem"))
    if vagas is None and vaga_vinculada:
        vagas = 1
    tipo_garagem = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "tipo_garagem"),
        contexto.get("tipo_garagem_padrao"),
        "carro",
    )

    area_real = _primeiro_valor_preenchido(obter_valor_coluna_planilha(linha, colunas, "ap_real"))
    area_total = _primeiro_valor_preenchido(obter_valor_coluna_planilha(linha, colunas, "area_total_construida"))
    area_equivalente = _primeiro_valor_preenchido(obter_valor_coluna_planilha(linha, colunas, "ap_equivalente"))
    area_m2 = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "area_m2"),
        area_real,
        area_total,
        area_equivalente,
    )
    meses_pre_entrega = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "meses_pre_entrega"),
        contexto.get("meses_pre_entrega_padrao"),
    )
    data_entrega = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "data_entrega"),
        contexto.get("data_entrega_padrao"),
    )
    meses_pos_entrega = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "meses_pos_entrega"),
        contexto.get("meses_pos_entrega_padrao"),
    )
    percentual_conclusao_obra = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "percentual_conclusao_obra"),
        contexto.get("percentual_conclusao_obra_padrao"),
    )
    percentual_fechamento_minimo = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "percentual_fechamento_minimo"),
        contexto.get("percentual_fechamento_minimo_padrao"),
    )
    valor_garantido = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "valor_garantido"),
        contexto.get("valor_garantido_padrao"),
    )
    valor_garantido_pre_obra_planejado = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "valor_garantido_pre_obra_planejado"),
        contexto.get("valor_garantido_pre_obra_planejado_padrao"),
    )
    percentual_captacao_ate_entrega = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "percentual_captacao_ate_entrega"),
        contexto.get("percentual_captacao_ate_entrega_padrao"),
    )
    valor_parcela_minima_pre_obra = _primeiro_valor_preenchido(
        obter_valor_coluna_planilha(linha, colunas, "valor_parcela_minima_pre_obra"),
        contexto.get("valor_parcela_minima_pre_obra_padrao"),
    )

    payload = normalizar_payload_imovel(
        {
            "titulo": titulo,
            "descricao": _gerar_descricao_importacao(
                descricao_manual=descricao_manual,
                localidade=localidade,
                unidade=unidade,
                pavimento=pavimento,
                bloco=bloco,
                orientacao=orientacao,
                posicao=posicao,
                suites=obter_valor_coluna_planilha(linha, colunas, "suite"),
                vaga_vinculada=vaga_vinculada,
                area_real=area_real,
                area_equivalente=area_equivalente,
                area_total=area_total,
            ),
            "tipo_imovel": tipo_imovel,
            "endereco": endereco,
            "cidade": cidade,
            "bairro": bairro,
            "estado": estado,
            "cep": cep,
            "valor": normalizar_decimal(valor, "valor"),
            "quartos": normalizar_inteiro(quartos, "quartos"),
            "banheiros": normalizar_inteiro(banheiros, "banheiros"),
            "vagas_garagem": normalizar_inteiro(vagas, "vagas de garagem"),
            "tipo_garagem": tipo_garagem,
            "area_m2": normalizar_decimal(area_m2, "area em m2"),
            "data_entrega": data_entrega,
            "meses_pre_entrega": meses_pre_entrega,
            "meses_pos_entrega": meses_pos_entrega,
            "percentual_conclusao_obra": percentual_conclusao_obra,
            "percentual_fechamento_minimo": percentual_fechamento_minimo,
            "valor_garantido": valor_garantido,
            "valor_garantido_pre_obra_planejado": valor_garantido_pre_obra_planejado,
            "percentual_captacao_ate_entrega": percentual_captacao_ate_entrega,
            "valor_parcela_minima_pre_obra": valor_parcela_minima_pre_obra,
            "status": status,
        }
    )

    return {
        "linha": linha_planilha,
        "referencia": titulo,
        "payload": payload,
        "chave": _normalizar_chave_importacao(
            payload["titulo"],
            payload.get("cidade"),
            payload.get("bairro"),
            payload.get("endereco"),
        ),
    }


def carregar_planilha_imoveis(
    *,
    conteudo_arquivo: bytes,
    nome_arquivo: str,
    contexto: dict[str, Any],
) -> dict[str, Any]:
    if load_workbook is None:
        raise HTTPException(
            status_code=500,
            detail="A dependencia openpyxl nao esta disponivel para processar planilhas Excel.",
        )

    nome_normalizado = Path(nome_arquivo or "").name
    extensao = Path(nome_normalizado).suffix.lower()
    if extensao not in EXTENSOES_PLANILHA_IMPORTACAO:
        raise HTTPException(status_code=400, detail="Envie uma planilha Excel valida nos formatos .xlsx ou .xlsm.")

    if not conteudo_arquivo:
        raise HTTPException(status_code=400, detail="A planilha enviada esta vazia.")

    try:
        workbook = load_workbook(BytesIO(conteudo_arquivo), data_only=True, read_only=True)
    except Exception as erro:
        raise HTTPException(status_code=400, detail="Nao foi possivel ler a planilha Excel enviada.") from erro

    try:
        if not workbook.sheetnames:
            raise HTTPException(status_code=400, detail="A planilha enviada nao possui abas disponiveis.")

        worksheet = workbook[workbook.sheetnames[0]]
        iterador = worksheet.iter_rows(values_only=True)
        cabecalhos = next(iterador, None)
        if not cabecalhos or linha_planilha_vazia(cabecalhos):
            raise HTTPException(status_code=400, detail="A planilha enviada nao possui cabecalho valido.")

        colunas = mapear_colunas_importacao(cabecalhos)
        itens = []
        total_linhas = 0

        for linha_planilha, linha in enumerate(iterador, start=2):
            if not linha or linha_planilha_vazia(linha):
                continue
            total_linhas += 1
            itens.append(_montar_payload_importacao_linha(linha, colunas, contexto, linha_planilha))

        if total_linhas <= 0:
            raise HTTPException(status_code=400, detail="A planilha enviada nao possui linhas de dados para importacao.")

        return {
            "arquivo": nome_normalizado or "planilha.xlsx",
            "aba": worksheet.title,
            "total_linhas": total_linhas,
            "itens": itens,
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass


async def importar_imoveis_em_lote(
    conexao,
    esquema: str,
    *,
    conteudo_arquivo: bytes,
    nome_arquivo: str,
    contexto: dict[str, Any],
    registrado_por: str | None = None,
) -> dict[str, Any]:
    planilha = carregar_planilha_imoveis(
        conteudo_arquivo=conteudo_arquivo,
        nome_arquivo=nome_arquivo,
        contexto=contexto,
    )

    importados: list[dict[str, Any]] = []
    ignorados: list[dict[str, Any]] = []
    erros: list[dict[str, Any]] = []
    chaves_processadas: set[str] = set()

    for item in planilha["itens"]:
        linha_planilha = item["linha"]
        referencia = item["referencia"]
        payload = item["payload"]
        chave = item["chave"]

        if chave and chave in chaves_processadas:
            ignorados.append(
                {
                    "linha": linha_planilha,
                    "referencia": referencia,
                    "motivo": "Duplicado dentro da propria planilha.",
                }
            )
            continue

        try:
            async with conexao.transaction():
                existente = await buscar_imovel_por_chave_importacao(
                    conexao,
                    esquema,
                    titulo=payload["titulo"],
                    endereco=payload.get("endereco"),
                    cidade=payload["cidade"],
                    bairro=payload["bairro"],
                )
                if existente:
                    ignorados.append(
                        {
                            "linha": linha_planilha,
                            "referencia": referencia,
                            "motivo": "Ja existe um imovel com o mesmo titulo e localizacao cadastrados.",
                        }
                    )
                    if chave:
                        chaves_processadas.add(chave)
                    continue

                registro = await criar_imovel_repositorio(conexao, esquema, payload)
                await _registrar_evolucao_obra_se_necessario(
                    conexao,
                    esquema,
                    registro["identificador_imovel"],
                    payload["percentual_conclusao_obra"],
                    registrado_por=registrado_por,
                    observacoes="Registro inicial via importacao.",
                )
            importados.append(serializar_imovel(registro))
            if chave:
                chaves_processadas.add(chave)
        except HTTPException as erro:
            erros.append(
                {
                    "linha": linha_planilha,
                    "referencia": referencia,
                    "motivo": str(erro.detail),
                }
            )
        except Exception as erro:
            erros.append(
                {
                    "linha": linha_planilha,
                    "referencia": referencia,
                    "motivo": f"Falha inesperada ao importar a linha: {erro}",
                }
            )

    return {
        "arquivo": planilha["arquivo"],
        "aba": planilha["aba"],
        "total_linhas": planilha["total_linhas"],
        "total_importados": len(importados),
        "total_ignorados": len(ignorados),
        "total_erros": len(erros),
        "amostra_importados": [
            {"id": item["id"], "titulo": item["titulo"], "cidade": item["cidade"], "bairro": item["bairro"]}
            for item in importados[:12]
        ],
        "linhas_ignoradas": ignorados,
        "linhas_com_erro": erros,
    }


async def _gravar_upload_em_disco(
    arquivo: UploadFile,
    destino: Path,
    limite_bytes: int,
) -> int:
    total = 0
    destino.parent.mkdir(parents=True, exist_ok=True)

    with destino.open("wb") as ponteiro:
        while True:
            bloco = await arquivo.read(1024 * 1024)
            if not bloco:
                break
            total += len(bloco)
            if total > limite_bytes:
                ponteiro.close()
                destino.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=400,
                    detail=f"O arquivo {arquivo.filename or 'informado'} excede o limite permitido.",
                )
            ponteiro.write(bloco)

    if total <= 0:
        destino.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="Nao e possivel enviar arquivos vazios.")

    return total


def _resolver_caminho_publico(base_publica: Path, caminho_publico: str) -> Path | None:
    relativo = str(caminho_publico or "").replace("\\", "/").lstrip("/")
    if not relativo:
        return None

    base_resolvida = base_publica.resolve()
    destino = (base_resolvida / relativo).resolve()

    if destino != base_resolvida and base_resolvida not in destino.parents:
        return None

    return destino


def remover_arquivo_publico(base_publica: Path, caminho_publico: str) -> None:
    destino = _resolver_caminho_publico(base_publica, caminho_publico)
    if not destino:
        return

    try:
        destino.unlink(missing_ok=True)
    except Exception:
        return

    diretorio = destino.parent
    while diretorio != base_publica.resolve():
        try:
            diretorio.rmdir()
        except OSError:
            break
        diretorio = diretorio.parent


def remover_arquivos_publicos(base_publica: Path, caminhos_publicos: Iterable[str]) -> None:
    for caminho_publico in caminhos_publicos:
        remover_arquivo_publico(base_publica, caminho_publico)


async def salvar_midias_do_imovel(
    conexao,
    esquema: str,
    *,
    identificador_imovel: str,
    arquivos: list[UploadFile],
    diretorio_uploads: Path,
    diretorio_publico: Path,
    url_base_uploads: str,
    limite_imagem_bytes: int,
    limite_video_bytes: int,
) -> list[dict[str, Any]]:
    if not arquivos:
        raise HTTPException(status_code=400, detail="Nenhuma midia foi enviada.")

    diretorio_imovel = Path(diretorio_uploads) / identificador_imovel
    url_base = str(url_base_uploads or "/uploads/imoveis").rstrip("/")

    caminhos_salvos: list[Path] = []
    registros: list[dict[str, Any]] = []

    try:
        for arquivo in arquivos:
            metadados = classificar_arquivo_de_midia(arquivo.filename, arquivo.content_type)
            limite_bytes = limite_imagem_bytes if metadados["tipo_arquivo"] == "foto" else limite_video_bytes
            nome_base = normalizar_nome_para_arquivo(Path(arquivo.filename or metadados["tipo_arquivo"]).stem)
            nome_final = f"{nome_base}-{uuid4().hex[:12]}{metadados['extensao']}"
            destino = diretorio_imovel / nome_final

            tamanho_bytes = await _gravar_upload_em_disco(arquivo, destino, limite_bytes)
            caminhos_salvos.append(destino)

            caminho_publico = f"{url_base}/{identificador_imovel}/{nome_final}"
            registro = await inserir_midia_imovel(
                conexao,
                esquema,
                identificador_imovel=identificador_imovel,
                tipo_arquivo=metadados["tipo_arquivo"],
                nome_arquivo=Path(arquivo.filename or nome_final).name,
                caminho_arquivo=caminho_publico,
                mime_type=metadados["mime_type"],
                tamanho_bytes=tamanho_bytes,
            )
            registros.append(serializar_midia(registro))
    except Exception:
        for caminho in caminhos_salvos:
            try:
                caminho.unlink(missing_ok=True)
            except Exception:
                continue
        raise
    finally:
        for arquivo in arquivos:
            try:
                await arquivo.close()
            except Exception:
                continue

    return registros
