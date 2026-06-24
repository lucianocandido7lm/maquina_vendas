"""
Servicos de dominio para Metas e Resultados.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from calendar import monthrange
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

from fastapi import HTTPException, status


INDICADORES_PADRAO = (
    "LEADS",
    "VISITAS",
    "PROPOSTAS",
    "CANCELAMENTOS",
    "VENDAS_FINALIZADAS",
    "DISTRATOS",
    "REPASSES",
    "SLA_FINALIZACAO",
    "SLA_REPASSE",
    "IPC",
    "ALUGUEL",
    "VENDA",
    "SOBREPRECO",
)

TIPOS_USUARIO = {"CORRETOR", "GESTOR", "COORDENADOR"}
ORIGENS_META = {"MANUAL", "AUTOMATICA"}
TIPOS_META_GERENCIAL = {"REGIONAL", "EMPREENDIMENTO", "GLOBAL"}
ORIGENS_RESULTADO = {"MANUAL", "SISTEMA", "IMPORTACAO", "CALCULADO"}
REGRAS_META_GERENCIAL = {">", "<", "=", "BETWEEN", "FATO_1_BETWEEN_FATO_2"}


def decimalizar(valor: Any, padrao: Decimal | None = Decimal("0")) -> Decimal | None:
    if valor is None or valor == "":
        return padrao
    if isinstance(valor, Decimal):
        return valor
    try:
        texto = str(valor).strip()
        if not texto:
            return padrao
        if "," in texto and "." in texto:
            if texto.rfind(",") > texto.rfind("."):
                texto = texto.replace(".", "").replace(",", ".")
            else:
                texto = texto.replace(",", "")
        elif "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        return Decimal(texto)
    except Exception:
        return padrao


def arredondar_decimal(valor: Any, casas: str = "0.01") -> Decimal:
    return (decimalizar(valor) or Decimal("0")).quantize(Decimal(casas), rounding=ROUND_HALF_UP)


def normalizar_texto(valor: Any) -> str:
    return str(valor or "").strip()


def remover_acentos(valor: Any) -> str:
    texto = unicodedata.normalize("NFD", normalizar_texto(valor))
    return "".join(ch for ch in texto if unicodedata.category(ch) != "Mn")


def normalizar_codigo(valor: Any) -> str:
    texto = remover_acentos(valor).upper()
    texto = re.sub(r"[^A-Z0-9]+", "_", texto)
    return texto.strip("_")


def normalizar_tipo_usuario(valor: Any) -> str:
    tipo = normalizar_codigo(valor)
    if tipo not in TIPOS_USUARIO:
        raise HTTPException(status_code=422, detail="Tipo de usuário inválido.")
    return tipo


def normalizar_origem_meta(valor: Any) -> str:
    origem = normalizar_codigo(valor or "MANUAL")
    if origem not in ORIGENS_META:
        raise HTTPException(status_code=422, detail="Origem da meta inválida.")
    return origem


def normalizar_tipo_meta_gerencial(valor: Any) -> str:
    tipo = normalizar_codigo(valor)
    if tipo not in TIPOS_META_GERENCIAL:
        raise HTTPException(status_code=422, detail="Tipo de meta gerencial inválido.")
    return tipo


def normalizar_origem_resultado(valor: Any) -> str:
    origem = normalizar_codigo(valor or "MANUAL")
    if origem not in ORIGENS_RESULTADO:
        raise HTTPException(status_code=422, detail="Origem do resultado inválida.")
    return origem


def normalizar_regra_meta(valor: Any) -> str | None:
    texto = remover_acentos(valor).lower().strip()
    if not texto:
        return None
    if "between" in texto:
        return "BETWEEN"
    if texto in {">", "<", "="}:
        return texto
    codigo = normalizar_codigo(texto)
    if codigo in REGRAS_META_GERENCIAL:
        return "BETWEEN" if codigo == "FATO_1_BETWEEN_FATO_2" else codigo
    raise HTTPException(status_code=422, detail="Regra da meta gerencial inválida.")


def validar_mes_ano(mes: int, ano: int) -> None:
    if int(mes) < 1 or int(mes) > 12:
        raise HTTPException(status_code=422, detail="Mês de referência inválido.")
    if int(ano) < 2000 or int(ano) > 2100:
        raise HTTPException(status_code=422, detail="Ano de referência inválido.")


def validar_vigencia(data_inicio: date, data_fim: date) -> None:
    if data_inicio > data_fim:
        raise HTTPException(status_code=422, detail="Data inicial não pode ser maior que a data final.")


def ultimo_dia_mes(mes: int, ano: int) -> date:
    validar_mes_ano(mes, ano)
    return date(int(ano), int(mes), monthrange(int(ano), int(mes))[1])


def calcular_limite_alteracao(mes: int, ano: int) -> datetime:
    fechamento = ultimo_dia_mes(mes, ano)
    limite = fechamento - timedelta(days=5)
    return datetime.combine(limite, time(23, 59, 59))


def validar_prazo_alteracao(mes: int, ano: int, referencia: datetime | None = None) -> dict[str, Any]:
    referencia = referencia or datetime.now()
    limite = calcular_limite_alteracao(mes, ano)
    mes_atual = referencia.month
    ano_atual = referencia.year
    periodo = (int(ano), int(mes))
    periodo_atual = (ano_atual, mes_atual)

    if periodo < periodo_atual:
        return {
            "permitido": False,
            "limite": limite,
            "mensagem": "O prazo para alteração desta meta foi encerrado.",
        }

    if periodo == periodo_atual and referencia > limite:
        return {
            "permitido": False,
            "limite": limite,
            "mensagem": "O prazo para alteração desta meta foi encerrado.",
        }

    return {
        "permitido": True,
        "limite": limite,
        "mensagem": f"Esta meta pode ser alterada até {limite.strftime('%d/%m/%Y')}.",
    }


def exigir_prazo_alteracao(mes: int, ano: int, referencia: datetime | None = None) -> None:
    prazo = validar_prazo_alteracao(mes, ano, referencia)
    if not prazo["permitido"]:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=prazo["mensagem"])


def calcular_atingimento(valor_realizado: Any, meta_oficial: Any) -> Decimal | None:
    meta = decimalizar(meta_oficial, None)
    realizado = decimalizar(valor_realizado, Decimal("0")) or Decimal("0")
    if meta is None or meta <= 0:
        return None
    return ((realizado / meta) * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calcular_faltante(valor_realizado: Any, meta_oficial: Any) -> Decimal | None:
    meta = decimalizar(meta_oficial, None)
    realizado = decimalizar(valor_realizado, Decimal("0")) or Decimal("0")
    if meta is None or meta <= 0:
        return None
    faltante = meta - realizado
    return faltante if faltante > 0 else Decimal("0")


def classificar_status_meta(valor_realizado: Any, meta_oficial: Any, resultado_existe: bool = True) -> str:
    meta = decimalizar(meta_oficial, None)
    if meta is None or meta <= 0:
        return "Sem meta cadastrada"
    if not resultado_existe:
        return "Sem resultado lançado"

    atingimento = calcular_atingimento(valor_realizado, meta)
    if atingimento is None:
        return "Sem meta cadastrada"
    if atingimento >= Decimal("100"):
        return "Acima da meta"
    if atingimento >= Decimal("75"):
        return "Dentro do esperado"
    if atingimento >= Decimal("50"):
        return "Abaixo do esperado"
    return "Crítico"


def resolver_meta_oficial_gestor(meta_manual: Any, meta_automatica: Any) -> dict[str, Any]:
    manual = decimalizar(meta_manual, None)
    automatica = decimalizar(meta_automatica, Decimal("0")) or Decimal("0")
    if manual is not None and manual >= 0:
        return {
            "meta_oficial": manual,
            "meta_automatica": automatica,
            "origem_meta": "MANUAL",
        }
    return {
        "meta_oficial": automatica,
        "meta_automatica": automatica,
        "origem_meta": "AUTOMATICA",
    }


def calcular_meta_potencial_colaborador(tipo_usuario: Any, meta_manual: Any, soma_corretores: Any = None) -> Decimal:
    tipo = normalizar_tipo_usuario(tipo_usuario)
    if tipo == "CORRETOR":
        return arredondar_decimal(meta_manual)
    return arredondar_decimal(soma_corretores)


def preparar_nova_versao_meta(meta_anterior: dict[str, Any] | None, data_inicio: date, maior_versao: int | None) -> dict[str, Any]:
    versao = int(maior_versao or 0) + 1
    data_fim_anterior = None
    if meta_anterior:
        data_fim_anterior = data_inicio - timedelta(days=1)
    return {
        "versao": max(versao, 1),
        "data_fim_anterior": data_fim_anterior,
    }


def usuario_pode_gerenciar_subordinado(
    *,
    usuario_admin: bool,
    identificador_usuario: str,
    identificador_alvo: str,
    subordinados: set[str] | list[str] | tuple[str, ...],
) -> bool:
    if usuario_admin:
        return True
    return str(identificador_alvo) in {str(item) for item in subordinados} and str(identificador_alvo) != str(identificador_usuario)


def detectar_escopo_visualizacao_metas(
    *,
    perfis: list[str] | tuple[str, ...] | set[str] | None,
    usuario_admin: bool,
    usuario_manage: bool,
    usuario_gerenciais: bool = False,
    usuario_resultados: bool = False,
) -> str:
    perfis_normalizados = {remover_acentos(item).lower() for item in (perfis or []) if normalizar_texto(item)}
    eh_diretor = any("diretor" in item for item in perfis_normalizados)
    if usuario_admin or eh_diretor or (usuario_gerenciais and usuario_resultados and not usuario_manage):
        return "GLOBAL"
    if usuario_manage:
        return "GESTOR"
    return "PESSOAL"


def usuarios_permitidos_no_escopo(
    *,
    escopo: str,
    identificador_usuario: str,
    subordinados: set[str] | list[str] | tuple[str, ...] | None = None,
) -> set[str] | None:
    escopo_normalizado = normalizar_codigo(escopo or "PESSOAL")
    if escopo_normalizado == "GLOBAL":
        return None
    permitidos = {str(identificador_usuario)}
    if escopo_normalizado == "GESTOR":
        permitidos.update(str(item) for item in (subordinados or []) if item)
    return permitidos


def usuario_no_escopo_visualizacao(
    *,
    escopo: str,
    identificador_usuario: str,
    identificador_alvo: str | None,
    subordinados: set[str] | list[str] | tuple[str, ...] | None = None,
) -> bool:
    if not identificador_alvo:
        return normalizar_codigo(escopo or "PESSOAL") == "GLOBAL"
    permitidos = usuarios_permitidos_no_escopo(
        escopo=escopo,
        identificador_usuario=identificador_usuario,
        subordinados=subordinados,
    )
    if permitidos is None:
        return True
    return str(identificador_alvo) in permitidos


def filtrar_referencias_por_escopo(
    referencias: dict[str, Any],
    *,
    escopo: str,
    identificador_usuario: str,
    subordinados: set[str] | list[str] | tuple[str, ...] | None = None,
) -> dict[str, Any]:
    permitidos = usuarios_permitidos_no_escopo(
        escopo=escopo,
        identificador_usuario=identificador_usuario,
        subordinados=subordinados,
    )
    if permitidos is None:
        return dict(referencias or {})

    referencias_filtradas = dict(referencias or {})
    usuarios = [
        dict(item)
        for item in (referencias_filtradas.get("usuarios") or [])
        if str(item.get("identificador_usuario") or "") in permitidos
    ]
    equipes_permitidas = {
        str(codigo)
        for item in usuarios
        for codigo in (item.get("equipes") or [])
        if normalizar_texto(codigo)
    }
    referencias_filtradas["usuarios"] = usuarios
    referencias_filtradas["equipes"] = [
        dict(item)
        for item in (referencias_filtradas.get("equipes") or [])
        if str(item.get("codigo") or "") in equipes_permitidas
    ]
    return referencias_filtradas


def serializar_decimal(valor: Any) -> float | None:
    if valor is None:
        return None
    numero = decimalizar(valor, None)
    return float(numero) if numero is not None else None


def serializar_data(valor: Any) -> str | None:
    if not valor:
        return None
    if isinstance(valor, (date, datetime)):
        return valor.isoformat()
    return str(valor)


def enriquecer_calculo_meta(item: dict[str, Any]) -> dict[str, Any]:
    meta = item.get("meta_oficial", item.get("meta_valor"))
    realizado = item.get("valor_realizado")
    resultado_existe = item.get("resultado_existe", realizado is not None)
    atingimento = calcular_atingimento(realizado, meta)
    faltante = calcular_faltante(realizado, meta)
    item["atingimento_percentual"] = serializar_decimal(atingimento)
    item["faltante"] = serializar_decimal(faltante)
    item["status_resultado"] = classificar_status_meta(realizado, meta, bool(resultado_existe))
    return item


def normalizar_chaves_linha(linha: dict[str, Any]) -> dict[str, Any]:
    normalizada: dict[str, Any] = {}
    for chave, valor in (linha or {}).items():
        codigo = normalizar_codigo(chave)
        normalizada[codigo] = valor
    return normalizada


def normalizar_data_importacao(valor: Any) -> date | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


def validar_linha_importacao_colaborador(
    linha: dict[str, Any],
    *,
    indicadores_por_codigo: dict[str, Any],
    usuarios_por_nome: dict[str, Any],
) -> dict[str, Any]:
    dados = normalizar_chaves_linha(linha)
    erros: list[str] = []
    funcionario = normalizar_texto(dados.get("FUNCIONARIO"))
    indicador_codigo = normalizar_codigo(dados.get("INDICADOR"))
    meta = decimalizar(dados.get("META"), None)
    meta_potencial = decimalizar(dados.get("META_POTENCIAL"), None)
    data_inicio = normalizar_data_importacao(dados.get("DATA_INICIAL") or dados.get("DATA_INICIO"))
    data_fim = normalizar_data_importacao(dados.get("DATA_FIM"))

    usuario = usuarios_por_nome.get(remover_acentos(funcionario).lower())
    indicador = indicadores_por_codigo.get(indicador_codigo)

    if not funcionario:
        erros.append("Funcionário não informado.")
    elif not usuario:
        erros.append(f"Funcionário não encontrado: {funcionario}.")
    if not indicador_codigo:
        erros.append("Indicador não informado.")
    elif not indicador:
        erros.append(f"Indicador não encontrado: {indicador_codigo}.")
    if meta is None or meta < 0:
        erros.append("Meta inválida.")
    if data_inicio is None:
        erros.append("Data inicial inválida.")
    if data_fim is None:
        erros.append("Data fim inválida.")
    if data_inicio and data_fim and data_inicio > data_fim:
        erros.append("Data inicial maior que data fim.")

    return {
        "valida": not erros,
        "erros": erros,
        "dados": {
            "id_indicado_meta": dados.get("ID_INDICADO_META"),
            "usuario_id": usuario.get("identificador_usuario") if usuario else None,
            "funcionario": funcionario,
            "indicador_id": indicador.get("id") if indicador else None,
            "indicador": indicador_codigo,
            "meta_potencial": meta_potencial,
            "meta_valor": meta,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
        },
    }


def validar_linha_importacao_gerencial(
    linha: dict[str, Any],
    *,
    indicadores_por_codigo: dict[str, Any],
    usuarios_por_nome: dict[str, Any],
) -> dict[str, Any]:
    dados = normalizar_chaves_linha(linha)
    erros: list[str] = []
    indicador_codigo = normalizar_codigo(dados.get("INDICADOR_META") or dados.get("INDICADOR"))
    indicador = indicadores_por_codigo.get(indicador_codigo)
    pessoa_nome = normalizar_texto(dados.get("PESSOA"))
    pessoa = usuarios_por_nome.get(remover_acentos(pessoa_nome).lower()) if pessoa_nome else None
    tipo_meta = normalizar_codigo(dados.get("TIPO_META") or "GLOBAL")
    visao_meta = normalizar_texto(dados.get("VISAO_META"))
    data_inicio = normalizar_data_importacao(dados.get("DATA_INICIO") or dados.get("DATA_INICIAL"))
    data_fim = normalizar_data_importacao(dados.get("DATA_FIM"))
    meta = decimalizar(dados.get("META"), None)
    fato_1 = decimalizar(dados.get("FATO_1"), None)
    fato_2 = decimalizar(dados.get("FATO_2"), None)
    fato_consolidado = decimalizar(dados.get("FATO_CONSOLIDADO"), None)
    peso = decimalizar(dados.get("PESO"), None)

    if indicador_codigo and not indicador:
        erros.append(f"Indicador não encontrado: {indicador_codigo}.")
    if not indicador_codigo:
        erros.append("Indicador da meta não informado.")
    if tipo_meta not in TIPOS_META_GERENCIAL:
        erros.append("Tipo de meta inválido.")
    if pessoa_nome and not pessoa:
        erros.append(f"Pessoa não encontrada: {pessoa_nome}.")
    if meta is None and fato_consolidado is None:
        erros.append("Meta ou fato consolidado deve ser informado.")
    if meta is not None and meta < 0:
        erros.append("Meta inválida.")
    if fato_1 is not None and fato_1 < 0:
        erros.append("Fato 1 inválido.")
    if fato_2 is not None and fato_2 < 0:
        erros.append("Fato 2 inválido.")
    if fato_consolidado is not None and fato_consolidado < 0:
        erros.append("Fato consolidado inválido.")
    if peso is not None and peso < 0:
        erros.append("Peso inválido.")
    if data_inicio is None:
        erros.append("Data início inválida.")
    if data_fim is None:
        erros.append("Data fim inválida.")
    if data_inicio and data_fim and data_inicio > data_fim:
        erros.append("Data início maior que data fim.")

    return {
        "valida": not erros,
        "erros": erros,
        "dados": {
            "pessoa_id": pessoa.get("identificador_usuario") if pessoa else None,
            "pessoa": pessoa_nome,
            "visao_meta": visao_meta,
            "tipo_meta": tipo_meta,
            "regiao_id": normalizar_texto(dados.get("REGIAO")),
            "empreendimento_id": normalizar_texto(dados.get("EMPREENDIMENTO")),
            "indicador_id": indicador.get("id") if indicador else None,
            "indicador": indicador_codigo,
            "meta_regra": normalizar_regra_meta(dados.get("REGRA_META") or dados.get("META_REGRA") or ""),
            "meta_valor": meta,
            "fato_1": fato_1,
            "fato_2": fato_2,
            "fato_consolidado": fato_consolidado,
            "peso": peso,
            "observacao": normalizar_texto(dados.get("META_OBSERVACAO") or dados.get("OBSERVACAO")),
            "data_inicio": data_inicio,
            "data_fim": data_fim,
        },
    }


def extrair_linhas_planilha(conteudo: bytes, nome_arquivo: str) -> dict[str, list[dict[str, Any]]]:
    nome = str(nome_arquivo or "").lower()
    if nome.endswith(".csv") or nome.endswith(".tsv"):
        separador = "\t" if nome.endswith(".tsv") else ","
        texto = conteudo.decode("utf-8-sig", errors="replace")
        leitor = csv.DictReader(io.StringIO(texto), delimiter=separador)
        return {"metas_colaboradores": [dict(linha) for linha in leitor]}

    if not nome.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=422, detail="Arquivo de importação deve ser .xlsx, .xlsm, .csv ou .tsv.")

    try:
        from openpyxl import load_workbook
    except Exception as erro:  # pragma: no cover
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl indisponivel para importar planilha.") from erro

    workbook = load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    resultado: dict[str, list[dict[str, Any]]] = {}
    for planilha in workbook.worksheets:
        linhas = list(planilha.iter_rows(values_only=True))
        if not linhas:
            resultado[planilha.title] = []
            continue
        cabecalhos = [normalizar_texto(item) for item in linhas[0]]
        itens: list[dict[str, Any]] = []
        for valores in linhas[1:]:
            if not any(valor not in (None, "") for valor in valores):
                continue
            itens.append({cabecalhos[i]: valores[i] if i < len(valores) else None for i in range(len(cabecalhos)) if cabecalhos[i]})
        resultado[planilha.title] = itens
    return resultado
