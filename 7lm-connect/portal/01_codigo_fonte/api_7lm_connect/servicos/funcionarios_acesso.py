"""
Servicos do cadastro de funcionarios e quadro diario de acesso.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
import json
import re
import unicodedata
from typing import Any, Optional
from uuid import UUID

from fastapi import HTTPException


TIPOS_FUNCIONARIO = {"FUNCIONARIO", "CORRETOR", "SDR", "OUTRO"}
TIPOS_VINCULO = {"CLT", "PJ"}
STATUS_OPERACIONAIS = {"ATIVO", "AUSENTE", "FERIAS", "AFASTADO", "INATIVO", "PENDENTE"}
STATUS_LOGIN = {"LIBERADO", "BLOQUEADO", "PENDENTE", "SEM_EMAIL", "SEM_LOGIN"}
STATUS_VALIDACAO_CADASTRO = {"PENDENTE", "APROVADO", "REPROVADO"}


def _texto(valor: Any) -> str:
    return str(valor or "").strip()


def _texto_ou_none(valor: Any) -> Optional[str]:
    texto = _texto(valor)
    return texto or None


def _email_ou_none(valor: Any) -> Optional[str]:
    bruto = re.sub(r"\s+", "", _texto(valor).lower()).replace("mailto:", "")
    if not bruto:
        return None
    candidatos = re.split(r"[;,/|]+", bruto)
    for texto in candidatos:
        texto = texto.strip(" .")
        while ".com.com" in texto:
            texto = texto.replace(".com.com", ".com")
        while ".br.br" in texto:
            texto = texto.replace(".br.br", ".br")
        texto = re.sub(r"(\.com){2,}$", ".com", texto)
        if texto.count("@") != 1:
            continue
        dominio = texto.split("@", 1)[1]
        if "." not in dominio:
            continue
        return texto
    return None


def _normalizar_nome(valor: Any) -> str:
    texto = re.sub(r"\s+", " ", _texto(valor))
    texto = re.sub(r"[_]+", " ", texto)
    texto = re.sub(r"\s*-\s*", " - ", texto).strip()
    texto = re.sub(r"\s+-\s+(CLT|PJ|DESLIGAD[OA]|DESLIGAMENTO)\b.*$", "", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\b(CLT|PJ|DESLIGAD[OA]|DESLIGAMENTO)\b.*$", "", texto, flags=re.IGNORECASE).strip(" -")
    partes = re.split(r"(\s+|-)", texto.lower())
    return "".join(parte[:1].upper() + parte[1:] if parte and not parte.isspace() and parte != "-" else parte for parte in partes).strip()


def _nome_ou_none(valor: Any) -> Optional[str]:
    nome = _normalizar_nome(valor)
    return nome or None


def _chave_texto(valor: Any) -> str:
    texto = unicodedata.normalize("NFKD", _texto(valor))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", texto.lower()).strip()


def _lideranca_vaga(valor: Any) -> bool:
    chave = _chave_texto(valor)
    return chave in {"vago", "vaga aberta", "em aberto", "cargo vago"}


def _lideranca_ou_none(valor: Any) -> Optional[str]:
    if _lideranca_vaga(valor):
        return "Vago"
    return _nome_ou_none(valor)


def _documento_ou_none(valor: Any) -> Optional[str]:
    texto = "".join(ch for ch in _texto(valor) if ch.isdigit())
    return texto or None


def _matricula_ou_none(valor: Any) -> Optional[str]:
    texto = _texto(valor)
    if not texto:
        return None
    texto = re.sub(r"\s+", "", texto)
    texto = re.sub(r"[^0-9A-Za-z_.-]", "", texto)
    return texto.upper() or None


def _telefone_ou_none(valor: Any) -> Optional[str]:
    texto = _texto(valor)
    return texto or None


def _normalizar_chave(valor: Any) -> str:
    texto = unicodedata.normalize("NFKD", _texto(valor))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def _cargo_regra_equipe(valor: Any) -> str:
    cargo = _normalizar_chave(valor)
    if not cargo:
        return ""
    if "diretor" in cargo:
        return "DIRETOR_COMERCIAL"
    if "head" in cargo:
        return "HEAD_COMERCIAL"
    if "corretor" in cargo:
        return "CORRETOR"
    if "coordenador" in cargo:
        return "COORDENADOR"
    if "gerente_regional" in cargo:
        return "GERENTE_REGIONAL"
    if "gerente" in cargo:
        return "GERENTE"
    return ""


def _cargo_exige_equipe(valor: Any) -> bool:
    return _cargo_regra_equipe(valor) in {"CORRETOR", "COORDENADOR"}


def _cargo_usa_escopo_gestao(valor: Any) -> bool:
    return _cargo_regra_equipe(valor) in {
        "GERENTE",
        "GERENTE_REGIONAL",
        "HEAD_COMERCIAL",
        "DIRETOR_COMERCIAL",
    }


def _cargo_acesso_codigo(cargo: Any, tipo_funcionario: Any = None) -> str:
    regra = _cargo_regra_equipe(cargo)
    if regra:
        return regra
    tipo = _tipo_funcionario(tipo_funcionario)
    if tipo in {"CORRETOR", "SDR"}:
        return tipo
    return "COLABORADOR"


def _status_validacao(valor: Any, padrao: str = "PENDENTE") -> str:
    status = _texto(valor).upper().replace(" ", "_")
    if status in {"APROVAR", "APROVADA", "APROVADO"}:
        return "APROVADO"
    if status in {"REPROVAR", "REPROVADA", "REPROVADO"}:
        return "REPROVADO"
    if status in STATUS_VALIDACAO_CADASTRO:
        return status
    return padrao


def _data_ou_none(valor: Any) -> Optional[date]:
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = _texto(valor)
    if not texto:
        return None
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto[:10], formato).date()
        except ValueError:
            continue
    return None


def _data_hora_ou_none(valor: Any) -> Optional[datetime]:
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    texto = _texto(valor)
    if not texto:
        return None
    for formato in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto[:19], formato)
        except ValueError:
            continue
    return None


def _flag_ou_none(valor: Any) -> Optional[bool]:
    if isinstance(valor, bool):
        return valor
    texto = _texto(valor).lower()
    if not texto:
        return None
    if texto in {"s", "sim", "true", "1", "ativo", "liberado", "yes"}:
        return True
    if texto in {"n", "nao", "não", "false", "0", "inativo", "bloqueado", "no"}:
        return False
    return None


def _tipo_funcionario(valor: Any) -> str:
    tipo = _texto(valor).upper()
    return tipo if tipo in TIPOS_FUNCIONARIO else "OUTRO"


def _tipo_vinculo(valor: Any) -> Optional[str]:
    texto = _normalizar_chave(valor).upper()
    texto = texto.replace("Ô", "O").replace("Á", "A").replace("Â", "A").replace("Ã", "A")
    texto = texto.replace("_", "")
    if "AUTON" in texto:
        return "PJ"
    if "PJ" in texto:
        return "PJ"
    if "CLT" in texto:
        return "CLT"
    return texto if texto in TIPOS_VINCULO else None


def _uuid_ou_none(valor: Any) -> Optional[str]:
    if not valor:
        return None
    try:
        return str(UUID(str(valor)))
    except Exception:
        return None


def _status_operacional(valor: Any) -> str:
    status = _texto(valor).upper().replace("É", "E").replace("Ã", "A")
    status = status.replace(" ", "_")
    if status == "FÉRIAS":
        status = "FERIAS"
    return status if status in STATUS_OPERACIONAIS else "PENDENTE"


def _status_login(valor: Any) -> str:
    status = _texto(valor).upper().replace(" ", "_")
    if status in {"ATIVO", "COM_LOGIN", "SIM", "TRUE", "1"}:
        return "LIBERADO"
    if status in {"INATIVO", "NAO", "FALSE", "0", "SEM_ACESSO"}:
        return "BLOQUEADO"
    return status if status in STATUS_LOGIN else "PENDENTE"


def _bool_label(valor: Optional[bool], verdadeiro: str, falso: str, indefinido: str = "PENDENTE") -> str:
    if valor is True:
        return verdadeiro
    if valor is False:
        return falso
    return indefinido


def serializar_funcionario(linha: Any) -> dict[str, Any]:
    item = dict(linha)
    for chave in (
        "identificador_funcionario",
        "identificador_equipe_vigencia",
        "identificador_usuario",
        "identificador_usuario_vinculado",
        "validado_por",
        "perfil_acesso_excecao_por",
        "criado_por_usuario",
        "atualizado_por_usuario",
    ):
        if item.get(chave) is not None:
            item[chave] = str(item[chave])
    return item


def serializar_escopo_gestao(linha: Any) -> dict[str, Any]:
    item = dict(linha)
    for chave in (
        "identificador_escopo_gestao",
        "identificador_funcionario",
        "identificador_equipe_vigencia",
        "criado_por",
        "atualizado_por",
    ):
        if item.get(chave) is not None:
            item[chave] = str(item[chave])
    return item


def serializar_equipe_vigencia(linha: Any) -> dict[str, Any]:
    item = dict(linha)
    for chave in (
        "identificador_equipe_vigencia",
        "criado_por",
        "atualizado_por",
    ):
        if item.get(chave) is not None:
            item[chave] = str(item[chave])
    item["hc_planejado"] = int(item.get("hc_planejado") or 0)
    foco_planejado = item.get("foco_planejado")
    if isinstance(foco_planejado, str):
        try:
            foco_planejado = json.loads(foco_planejado)
        except json.JSONDecodeError:
            foco_planejado = []
    item["foco_planejado"] = foco_planejado if isinstance(foco_planejado, list) else []
    return item


def serializar_status_diario(linha: Any) -> dict[str, Any]:
    item = dict(linha)
    for chave in (
        "identificador_status",
        "identificador_funcionario",
        "identificador_usuario",
    ):
        if item.get(chave) is not None:
            item[chave] = str(item[chave])
    return item


def _normalizar_funcionario(dados: dict[str, Any]) -> dict[str, Any]:
    tipo = _tipo_funcionario(dados.get("tipo_funcionario") or dados.get("tipo") or "FUNCIONARIO")
    nome = _normalizar_nome(dados.get("nome") or dados.get("nome_completo"))
    email = _email_ou_none(dados.get("email") or dados.get("correio_eletronico"))
    documento = _documento_ou_none(dados.get("documento") or dados.get("cpf"))
    cnpj = _documento_ou_none(dados.get("cnpj") or dados.get("documento_empresa"))
    nome_empresa = _texto_ou_none(dados.get("nome_empresa") or dados.get("empresa") or dados.get("razao_social"))
    matricula = _matricula_ou_none(dados.get("matricula") or dados.get("coluna1") or dados.get("codigo_matricula"))
    tipo_vinculo = _tipo_vinculo(dados.get("tipo_vinculo") or dados.get("vinculo") or dados.get("pj_clt"))
    telefone = _telefone_ou_none(dados.get("telefone") or dados.get("celular") or dados.get("whatsapp"))
    origem_planilha = _texto_ou_none(dados.get("origem_planilha")) or "MANUAL"
    cadastro_manual = origem_planilha.upper() == "MANUAL"
    ativo = _flag_ou_none(dados.get("ativo")) if dados.get("ativo") is not None else True
    status_operacional = _status_operacional(dados.get("status_operacional") or ("INATIVO" if ativo is False else "ATIVO"))
    cadastro_inativo = ativo is False or status_operacional == "INATIVO"
    ativo_login = _flag_ou_none(dados.get("ativo_login"))
    if ativo_login is None:
        ativo_login = bool(ativo and email)
    if ativo is False:
        ativo_login = False

    if not nome:
        raise HTTPException(status_code=422, detail="Informe o nome da pessoa.")

    if cadastro_manual:
        if not documento:
            raise HTTPException(status_code=422, detail="Informe o CPF da pessoa.")
        if len(documento) != 11:
            raise HTTPException(status_code=422, detail="Informe um CPF valido com 11 digitos.")
        if tipo_vinculo == "PJ":
            if cnpj and len(cnpj) != 14:
                raise HTTPException(status_code=422, detail="Informe um CNPJ valido com 14 digitos.")
            if not cadastro_inativo:
                if not cnpj:
                    raise HTTPException(status_code=422, detail="Informe o CNPJ da empresa para pessoa PJ.")
                if not nome_empresa:
                    raise HTTPException(status_code=422, detail="Informe o nome da empresa para pessoa PJ.")
        else:
            cnpj = None
            nome_empresa = None
        if ativo_login and not email:
            raise HTTPException(status_code=422, detail="Informe o e-mail para liberar login.")
    elif not email and not documento and not matricula:
        raise HTTPException(status_code=422, detail="Informe ao menos e-mail, documento ou matricula da pessoa.")
    elif tipo_vinculo != "PJ":
        cnpj = None
        nome_empresa = None

    cargo = _texto_ou_none(dados.get("cargo") or dados.get("funcao") or dados.get("função"))
    imobiliaria = _texto_ou_none(dados.get("imobiliaria"))
    identificador_equipe_vigencia = _uuid_ou_none(dados.get("identificador_equipe_vigencia"))

    if status_operacional == "AFASTADO":
        imobiliaria = None
        identificador_equipe_vigencia = None
    elif cadastro_manual and _cargo_exige_equipe(cargo) and not imobiliaria and not identificador_equipe_vigencia:
        raise HTTPException(status_code=422, detail="Selecione uma equipe para corretor ou coordenador.")

    return {
        "tipo_funcionario": tipo,
        "tipo_vinculo": tipo_vinculo,
        "documento": documento,
        "cnpj": cnpj,
        "nome_empresa": nome_empresa,
        "matricula": matricula,
        "email": email,
        "nome": nome,
        "telefone": telefone,
        "cargo": cargo,
        "imobiliaria": imobiliaria,
        "identificador_equipe_vigencia": identificador_equipe_vigencia,
        "gestor_documento": None if status_operacional == "AFASTADO" else _documento_ou_none(dados.get("gestor_documento")),
        "gestor_email": None if status_operacional == "AFASTADO" else _email_ou_none(dados.get("gestor_email")),
        "gestor": None if status_operacional == "AFASTADO" else _nome_ou_none(dados.get("gestor")),
        "coordenador_documento": None if status_operacional == "AFASTADO" else _documento_ou_none(dados.get("coordenador_documento")),
        "coordenador_email": None if status_operacional == "AFASTADO" else _email_ou_none(dados.get("coordenador_email")),
        "coordenador": None if status_operacional == "AFASTADO" else _nome_ou_none(dados.get("coordenador")),
        "gerente_documento": None if status_operacional == "AFASTADO" else _documento_ou_none(dados.get("gerente_documento")),
        "gerente_email": None if status_operacional == "AFASTADO" else _email_ou_none(dados.get("gerente_email")),
        "gerente": None if status_operacional == "AFASTADO" else _nome_ou_none(dados.get("gerente")),
        "diretor_documento": None if status_operacional == "AFASTADO" else _documento_ou_none(dados.get("diretor_documento")),
        "diretor_email": None if status_operacional == "AFASTADO" else _email_ou_none(dados.get("diretor_email")),
        "diretor": None if status_operacional == "AFASTADO" else _nome_ou_none(dados.get("diretor")),
        "regional": None if status_operacional == "AFASTADO" else _texto_ou_none(dados.get("regional")),
        "regiao": None if status_operacional == "AFASTADO" else _texto_ou_none(dados.get("regiao")),
        "foco": None if status_operacional == "AFASTADO" else _normalizar_rotulo_foco(dados.get("foco")),
        "ativo_negocio": _flag_ou_none(dados.get("ativo_negocio")),
        "ativo": ativo,
        "ativo_login": ativo_login,
        "status_operacional": status_operacional,
        "data_cadastro_usuario": _data_hora_ou_none(dados.get("data_cadastro_usuario")),
        "data_admissao": _data_ou_none(dados.get("data_admissao") or dados.get("admissao") or dados.get("admissão")),
        "data_inicio_vigencia": _data_ou_none(dados.get("data_inicio_vigencia")),
        "data_fim_vigencia": _data_ou_none(dados.get("data_fim_vigencia")),
        "referencia_origem": _texto_ou_none(dados.get("referencia_origem") or dados.get("referencia")),
        "origem_planilha": origem_planilha,
        "cadastrado_por": _texto_ou_none(dados.get("cadastrado_por")),
        "observacao": _texto_ou_none(dados.get("observacao") or dados.get("observacoes") or dados.get("observações")),
        "identificador_usuario": _uuid_ou_none(dados.get("identificador_usuario")),
    }


def _normalizar_escopos_gestao(dados: Any, *, cargo: Any, status_operacional: Any) -> list[dict[str, Any]]:
    if status_operacional == "AFASTADO" or not _cargo_usa_escopo_gestao(cargo):
        return []
    if dados in (None, ""):
        return []
    if isinstance(dados, str):
        try:
            dados = json.loads(dados)
        except json.JSONDecodeError:
            raise HTTPException(status_code=422, detail="Escopos de gestao invalidos.")
    if not isinstance(dados, list):
        raise HTTPException(status_code=422, detail="Escopos de gestao devem ser uma lista.")

    escopos: list[dict[str, Any]] = []
    vistos: set[tuple[str, str, str]] = set()
    for indice, linha in enumerate(dados, start=1):
        if not isinstance(linha, dict):
            raise HTTPException(status_code=422, detail=f"Linha {indice} dos escopos de gestao invalida.")
        regiao = _normalizar_rotulo_foco(linha.get("regiao") or linha.get("regional") or linha.get("foco"))
        equipe = _texto_ou_none(linha.get("equipe") or linha.get("imobiliaria"))
        identificador_equipe_vigencia = _uuid_ou_none(linha.get("identificador_equipe_vigencia"))
        cargo_gestao = _texto_ou_none(linha.get("cargo_gestao") or cargo)
        if not regiao and not equipe and not identificador_equipe_vigencia:
            continue
        chave = (
            (identificador_equipe_vigencia or "").lower(),
            _normalizar_chave(equipe),
            _normalizar_chave(regiao),
        )
        if chave in vistos:
            continue
        vistos.add(chave)
        escopos.append(
            {
                "identificador_equipe_vigencia": identificador_equipe_vigencia,
                "regiao": regiao,
                "equipe": equipe,
                "cargo_gestao": cargo_gestao,
            }
        )
    return escopos


def _normalizar_status_equipe(valor: Any) -> str:
    status = _texto(valor).upper().replace(" ", "_")
    if status in {"INATIVO", "INATIVA", "ENCERRADO", "ENCERRADA"}:
        return "INATIVO"
    if status in {"PLANEJADO", "PLANEJADA", "PLANEJAMENTO", "PREVISTO", "PREVISTA"}:
        return "PLANEJADA"
    return "ATIVO"


def _normalizar_rotulo_foco(valor: Any) -> Optional[str]:
    texto = _texto_ou_none(valor)
    if not texto:
        return None
    chave = _normalizar_chave(texto)
    if chave in {"brasilia", "df", "distrito_federal"}:
        return "DF"
    return texto


def _inteiro_nao_negativo(valor: Any, *, campo: str, limite: int = 999) -> int:
    if valor in (None, ""):
        return 0
    try:
        numero = int(valor)
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail=f"Informe um numero valido para {campo}.")
    if numero < 0:
        raise HTTPException(status_code=422, detail=f"{campo} nao pode ser negativo.")
    if numero > limite:
        raise HTTPException(status_code=422, detail=f"{campo} nao pode ser maior que {limite}.")
    return numero


def _normalizar_foco_planejado(valor: Any) -> list[dict[str, Any]]:
    if not valor:
        return []
    if isinstance(valor, str):
        try:
            valor = json.loads(valor)
        except json.JSONDecodeError:
            raise HTTPException(status_code=422, detail="Planejamento por foco invalido.")
    if not isinstance(valor, list):
        raise HTTPException(status_code=422, detail="Planejamento por foco deve ser uma lista.")

    itens: list[dict[str, Any]] = []
    for indice, linha in enumerate(valor, start=1):
        if not isinstance(linha, dict):
            raise HTTPException(status_code=422, detail=f"Linha {indice} do foco planejado invalida.")
        foco = _normalizar_rotulo_foco(linha.get("foco") or linha.get("nivel") or linha.get("nome") or linha.get("regiao"))
        quantidade = _inteiro_nao_negativo(
            linha.get("quantidade") or linha.get("hc") or linha.get("vagas") or 0,
            campo=f"quantidade do foco {indice}",
        )
        if not foco and not quantidade:
            continue
        if not foco:
            raise HTTPException(status_code=422, detail=f"Informe o foco da linha {indice}.")
        itens.append({"foco": foco, "quantidade": quantidade})
    return itens


def _normalizar_equipe_vigencia(dados: dict[str, Any]) -> dict[str, Any]:
    equipe = _texto_ou_none(dados.get("equipe") or dados.get("imobiliaria"))
    if not equipe:
        raise HTTPException(status_code=422, detail="Informe o nome da equipe.")

    status_equipe = _normalizar_status_equipe(dados.get("status_equipe") or dados.get("status"))
    data_inicio = _data_ou_none(dados.get("data_inicio_vigencia") or dados.get("data_inicio"))
    if not data_inicio and status_equipe != "PLANEJADA":
        raise HTTPException(status_code=422, detail="Informe a data de inicio da equipe.")

    data_fim = _data_ou_none(dados.get("data_fim_vigencia") or dados.get("data_fim"))
    if data_inicio and data_fim and data_fim < data_inicio:
        raise HTTPException(status_code=422, detail="A data final da equipe nao pode ser anterior ao inicio.")

    foco_planejado = _normalizar_foco_planejado(dados.get("foco_planejado"))
    hc_planejado = _inteiro_nao_negativo(dados.get("hc_planejado"), campo="HC planejado")
    total_foco_planejado = sum(int(item["quantidade"]) for item in foco_planejado)
    if hc_planejado == 0 and total_foco_planejado:
        hc_planejado = total_foco_planejado
    if total_foco_planejado > hc_planejado:
        raise HTTPException(status_code=422, detail="A soma dos focos planejados nao pode ser maior que o HC planejado.")

    return {
        "equipe": equipe,
        "regiao": _texto_ou_none(dados.get("regiao") or dados.get("regional")),
        "gerente_vendas": _lideranca_ou_none(dados.get("gerente_vendas") or dados.get("gestor")),
        "gerente_comercial": _lideranca_ou_none(dados.get("gerente_comercial") or dados.get("coordenador")),
        "gerente_regional": _lideranca_ou_none(dados.get("gerente_regional") or dados.get("gerente")),
        "head_comercial": _nome_ou_none(dados.get("head_comercial")),
        "diretor_comercial": _nome_ou_none(dados.get("diretor_comercial") or dados.get("diretor")),
        "data_inicio_vigencia": data_inicio,
        "data_fim_vigencia": data_fim,
        "status_equipe": status_equipe,
        "ativo": _flag_ou_none(dados.get("ativo")) if dados.get("ativo") is not None else True,
        "origem_planilha": _texto_ou_none(dados.get("origem_planilha")) or "MANUAL",
        "observacao": _texto_ou_none(dados.get("observacao") or dados.get("observacoes")),
        "hc_planejado": hc_planejado,
        "foco_planejado": foco_planejado,
    }


async def _validar_sobreposicao_equipe(
    conexao,
    *,
    equipe: str,
    data_inicio: date,
    data_fim: Optional[date],
    identificador_equipe_vigencia: Optional[str] = None,
) -> None:
    existe = await conexao.fetchval(
        """
        select exists (
          select 1
            from sevenlm_connect.funcionario_equipe_vigencia
           where lower(equipe) = lower($1)
             and ativo = true
             and ($4::uuid is null or identificador_equipe_vigencia <> $4::uuid)
             and daterange(data_inicio_vigencia, coalesce(data_fim_vigencia, 'infinity'::date), '[]')
                 && daterange($2::date, coalesce($3::date, 'infinity'::date), '[]')
        )
        """,
        equipe,
        data_inicio,
        data_fim,
        identificador_equipe_vigencia,
    )
    if existe:
        raise HTTPException(status_code=409, detail="Ja existe uma vigencia ativa sobreposta para esta equipe.")


async def _garantir_equipe_planejada(conexao, item: dict[str, Any]) -> None:
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          add column if not exists hc_planejado integer not null default 0
        """
    )
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          add column if not exists foco_planejado jsonb not null default '[]'::jsonb
        """
    )
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          drop constraint if exists ck_funcionario_equipe_vigencia_hc_planejado
        """
    )
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          add constraint ck_funcionario_equipe_vigencia_hc_planejado
          check (hc_planejado >= 0 and hc_planejado <= 999)
        """
    )
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          drop constraint if exists ck_funcionario_equipe_vigencia_foco_planejado
        """
    )
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          add constraint ck_funcionario_equipe_vigencia_foco_planejado
          check (jsonb_typeof(foco_planejado) = 'array')
        """
    )
    if item.get("status_equipe") != "PLANEJADA" and item.get("data_inicio_vigencia") is not None:
        return
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          alter column data_inicio_vigencia drop not null
        """
    )
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          drop constraint if exists ck_funcionario_equipe_vigencia_status
        """
    )
    await conexao.execute(
        """
        alter table sevenlm_connect.funcionario_equipe_vigencia
          add constraint ck_funcionario_equipe_vigencia_status
          check (status_equipe in ('ATIVO', 'INATIVO', 'PLANEJADA'))
        """
    )


async def obter_equipe_vigente(
    conexao,
    *,
    equipe: Optional[str] = None,
    identificador_equipe_vigencia: Optional[str] = None,
    data_referencia: Optional[date] = None,
):
    if not equipe and not identificador_equipe_vigencia:
        return None

    referencia = data_referencia or date.today()

    if identificador_equipe_vigencia:
        return await conexao.fetchrow(
            """
            select *
              from sevenlm_connect.funcionario_equipe_vigencia
             where identificador_equipe_vigencia = $1::uuid
             limit 1
            """,
            identificador_equipe_vigencia,
        )

    return await conexao.fetchrow(
        """
        select *
          from sevenlm_connect.funcionario_equipe_vigencia
         where lower(equipe) = lower($1)
           and ativo = true
           and data_inicio_vigencia <= $2::date
           and (data_fim_vigencia is null or data_fim_vigencia >= $2::date)
         order by data_inicio_vigencia desc, data_hora_atualizado_em desc
         limit 1
        """,
        equipe,
        referencia,
    )


async def _aplicar_equipe_vigente_ao_funcionario(conexao, item: dict[str, Any]) -> dict[str, Any]:
    equipe = await obter_equipe_vigente(
        conexao,
        equipe=item.get("imobiliaria"),
        identificador_equipe_vigencia=item.get("identificador_equipe_vigencia"),
        data_referencia=item.get("data_inicio_vigencia") or date.today(),
    )
    if not equipe:
        return item

    item["identificador_equipe_vigencia"] = str(equipe["identificador_equipe_vigencia"])
    item["imobiliaria"] = equipe["equipe"]
    item["regiao"] = equipe["regiao"] or item.get("regiao")
    item["regional"] = equipe["regiao"] or item.get("regional")
    item["foco"] = _normalizar_rotulo_foco(item.get("foco") or equipe["regiao"] or item.get("regiao"))
    item["gestor"] = equipe["gerente_vendas"] or item.get("gestor")
    item["coordenador"] = equipe["gerente_comercial"] or item.get("coordenador")
    item["gerente"] = equipe["gerente_regional"] or item.get("gerente")
    item["diretor"] = equipe["diretor_comercial"] or item.get("diretor")
    return item


async def listar_equipes_vigencia(
    conexao,
    *,
    termo: str = "",
    equipe: str = "",
    status: str = "",
    apenas_vigentes: bool = False,
    data_referencia: Optional[date] = None,
    limite: int = 1000,
):
    termo_like = f"%{_texto(termo)}%"
    equipe_like = f"%{_texto(equipe)}%"
    status_normalizado = _normalizar_status_equipe(status) if _texto(status) else ""
    referencia = data_referencia or date.today()

    linhas = await conexao.fetch(
        """
        select *
          from sevenlm_connect.funcionario_equipe_vigencia
         where (
                $1 = '%%'
                or equipe ilike $1
                or coalesce(regiao, '') ilike $1
                or coalesce(gerente_vendas, '') ilike $1
                or coalesce(gerente_comercial, '') ilike $1
                or coalesce(gerente_regional, '') ilike $1
                or coalesce(head_comercial, '') ilike $1
                or coalesce(diretor_comercial, '') ilike $1
              )
           and ($2 = '%%' or equipe ilike $2)
           and ($3 = '' or status_equipe = $3)
           and (
                $4::boolean = false
                or (
                  ativo = true
                  and data_inicio_vigencia <= $5::date
                  and (data_fim_vigencia is null or data_fim_vigencia >= $5::date)
                )
              )
         order by equipe, data_inicio_vigencia desc, data_hora_atualizado_em desc
         limit $6
        """,
        termo_like,
        equipe_like,
        status_normalizado,
        apenas_vigentes,
        referencia,
        limite,
    )
    return [serializar_equipe_vigencia(linha) for linha in linhas]


async def salvar_equipe_vigencia(
    conexao,
    dados: dict[str, Any],
    identificador_equipe_vigencia: Optional[str] = None,
    atualizado_por: Optional[str] = None,
):
    item = _normalizar_equipe_vigencia(dados)
    await _garantir_equipe_planejada(conexao, item)
    if item["ativo"] and item["data_inicio_vigencia"]:
        await _validar_sobreposicao_equipe(
            conexao,
            equipe=item["equipe"],
            data_inicio=item["data_inicio_vigencia"],
            data_fim=item["data_fim_vigencia"],
            identificador_equipe_vigencia=identificador_equipe_vigencia,
        )

    valores = (
        item["equipe"],
        item["regiao"],
        item["gerente_vendas"],
        item["gerente_comercial"],
        item["gerente_regional"],
        item["head_comercial"],
        item["diretor_comercial"],
        item["data_inicio_vigencia"],
        item["data_fim_vigencia"],
        item["status_equipe"],
        item["ativo"],
        item["origem_planilha"],
        item["observacao"],
        item["hc_planejado"],
        json.dumps(item["foco_planejado"], ensure_ascii=False),
        atualizado_por,
    )

    if identificador_equipe_vigencia:
        linha = await conexao.fetchrow(
            """
            update sevenlm_connect.funcionario_equipe_vigencia
               set equipe = $2,
                   regiao = $3,
                   gerente_vendas = $4,
                   gerente_comercial = $5,
                   gerente_regional = $6,
                   head_comercial = $7,
                   diretor_comercial = $8,
                   data_inicio_vigencia = $9,
                   data_fim_vigencia = $10,
                   status_equipe = $11,
                   ativo = coalesce($12, true),
                   origem_planilha = coalesce($13, origem_planilha),
                   observacao = $14,
                   hc_planejado = $15,
                   foco_planejado = $16::jsonb,
                   atualizado_por = $17::uuid,
                   data_hora_atualizado_em = now()
             where identificador_equipe_vigencia = $1::uuid
             returning *
            """,
            identificador_equipe_vigencia,
            *valores,
        )
        if not linha:
            raise HTTPException(status_code=404, detail="Equipe nao encontrada.")
        return serializar_equipe_vigencia(linha), False

    linha = await conexao.fetchrow(
        """
        insert into sevenlm_connect.funcionario_equipe_vigencia (
            equipe,
            regiao,
            gerente_vendas,
            gerente_comercial,
            gerente_regional,
            head_comercial,
            diretor_comercial,
            data_inicio_vigencia,
            data_fim_vigencia,
            status_equipe,
            ativo,
            origem_planilha,
            observacao,
            hc_planejado,
            foco_planejado,
            criado_por,
            atualizado_por
        )
        values (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10,
            coalesce($11, true), $12, $13, $14, $15::jsonb, $16::uuid, $16::uuid
        )
        returning *
        """,
        *valores,
    )
    return serializar_equipe_vigencia(linha), True


async def _sincronizar_usuario_vinculado(conexao, funcionario: Any) -> None:
    item = dict(funcionario or {})
    identificador_usuario = item.get("identificador_usuario")
    if not identificador_usuario:
        return
    usuario_ativo = bool(item.get("ativo") is not False and item.get("ativo_login") is True)
    await conexao.execute(
        """
        update sevenlm_connect.usuario
           set indicador_ativo = $2
         where identificador_usuario = $1::uuid
        """,
        str(identificador_usuario),
        usuario_ativo,
    )
    if not usuario_ativo:
        await conexao.execute(
            """
            update sevenlm_connect.sessao
               set situacao_sessao = 'ENCERRADA',
                   data_hora_encerramento = coalesce(data_hora_encerramento, now())
             where identificador_usuario = $1::uuid
               and situacao_sessao = 'ATIVA'
            """,
            str(identificador_usuario),
        )


async def _sincronizar_lideranca_equipe_por_funcionario(conexao, item: dict[str, Any]) -> None:
    campo_por_cargo = {
        "GERENTE": "gerente_vendas",
        "COORDENADOR": "gerente_comercial",
        "GERENTE_REGIONAL": "gerente_regional",
    }
    campo = campo_por_cargo.get(_cargo_regra_equipe(item.get("cargo")))
    identificador_equipe_vigencia = item.get("identificador_equipe_vigencia")
    nome = _nome_ou_none(item.get("nome"))
    if not campo or not identificador_equipe_vigencia or not nome:
        return

    await conexao.execute(
        f"""
        update sevenlm_connect.funcionario_equipe_vigencia
           set {campo} = $2,
               data_hora_atualizado_em = now()
         where identificador_equipe_vigencia = $1::uuid
           and coalesce({campo}, '') is distinct from $2
        """,
        identificador_equipe_vigencia,
        nome,
    )


async def _garantir_funcionario_gestao_escopo(conexao) -> None:
    await conexao.execute(
        """
        create table if not exists sevenlm_connect.funcionario_gestao_escopo (
            identificador_escopo_gestao uuid primary key default gen_random_uuid(),
            identificador_funcionario uuid not null
                references sevenlm_connect.funcionario_acesso(identificador_funcionario)
                on delete cascade,
            identificador_equipe_vigencia uuid null
                references sevenlm_connect.funcionario_equipe_vigencia(identificador_equipe_vigencia)
                on delete set null,
            regiao text null,
            equipe text null,
            cargo_gestao text null,
            ativo boolean not null default true,
            criado_por uuid null references sevenlm_connect.usuario(identificador_usuario),
            atualizado_por uuid null references sevenlm_connect.usuario(identificador_usuario),
            data_hora_criado_em timestamptz not null default now(),
            data_hora_atualizado_em timestamptz not null default now()
        )
        """
    )
    await conexao.execute(
        """
        create index if not exists ix_funcionario_gestao_escopo_funcionario
            on sevenlm_connect.funcionario_gestao_escopo(identificador_funcionario)
        """
    )
    await conexao.execute(
        """
        create index if not exists ix_funcionario_gestao_escopo_equipe
            on sevenlm_connect.funcionario_gestao_escopo(identificador_equipe_vigencia)
        """
    )
    await conexao.execute(
        """
        create index if not exists ix_funcionario_gestao_escopo_regiao
            on sevenlm_connect.funcionario_gestao_escopo(regiao)
            where ativo = true
        """
    )


async def _resolver_escopo_equipe_vigente(conexao, escopo: dict[str, Any]) -> dict[str, Any]:
    equipe = await obter_equipe_vigente(
        conexao,
        equipe=escopo.get("equipe"),
        identificador_equipe_vigencia=escopo.get("identificador_equipe_vigencia"),
        data_referencia=date.today(),
    )
    if not equipe:
        return escopo
    return {
        **escopo,
        "identificador_equipe_vigencia": str(equipe["identificador_equipe_vigencia"]),
        "equipe": equipe["equipe"] or escopo.get("equipe"),
        "regiao": _normalizar_rotulo_foco(escopo.get("regiao") or equipe["regiao"]),
    }


async def _salvar_funcionario_escopos_gestao(
    conexao,
    *,
    identificador_funcionario: str,
    cargo: Any,
    status_operacional: Any,
    escopos_raw: Any,
    atualizado_por: Optional[str] = None,
) -> list[dict[str, Any]]:
    await _garantir_funcionario_gestao_escopo(conexao)
    escopos = _normalizar_escopos_gestao(
        escopos_raw,
        cargo=cargo,
        status_operacional=status_operacional,
    )
    escopos_resolvidos = [
        await _resolver_escopo_equipe_vigente(conexao, escopo)
        for escopo in escopos
    ]

    await conexao.execute(
        """
        update sevenlm_connect.funcionario_gestao_escopo
           set ativo = false,
               atualizado_por = $2::uuid,
               data_hora_atualizado_em = now()
         where identificador_funcionario = $1::uuid
           and ativo = true
        """,
        identificador_funcionario,
        atualizado_por,
    )

    if not escopos_resolvidos:
        return []

    linhas: list[dict[str, Any]] = []
    for escopo in escopos_resolvidos:
        linha = await conexao.fetchrow(
            """
            insert into sevenlm_connect.funcionario_gestao_escopo (
                identificador_funcionario,
                identificador_equipe_vigencia,
                regiao,
                equipe,
                cargo_gestao,
                criado_por,
                atualizado_por
            )
            values ($1::uuid, $2::uuid, $3, $4, $5, $6::uuid, $6::uuid)
            returning *
            """,
            identificador_funcionario,
            escopo.get("identificador_equipe_vigencia"),
            escopo.get("regiao"),
            escopo.get("equipe"),
            escopo.get("cargo_gestao"),
            atualizado_por,
        )
        linhas.append(serializar_escopo_gestao(linha))
    return linhas


async def _anexar_escopos_gestao_funcionarios(conexao, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not items:
        return items
    await _garantir_funcionario_gestao_escopo(conexao)
    ids = [item.get("identificador_funcionario") for item in items if item.get("identificador_funcionario")]
    if not ids:
        return items
    linhas = await conexao.fetch(
        """
        select
            ge.*,
            eq.equipe as equipe_vigencia_nome,
            eq.regiao as equipe_vigencia_regiao,
            eq.status_equipe as equipe_status
          from sevenlm_connect.funcionario_gestao_escopo ge
          left join sevenlm_connect.funcionario_equipe_vigencia eq
            on eq.identificador_equipe_vigencia = ge.identificador_equipe_vigencia
         where ge.identificador_funcionario = any($1::uuid[])
           and ge.ativo = true
         order by coalesce(ge.regiao, eq.regiao, ''), coalesce(ge.equipe, eq.equipe, '')
        """,
        ids,
    )
    por_funcionario: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for linha in linhas:
        escopo = serializar_escopo_gestao(linha)
        escopo["equipe"] = escopo.get("equipe") or escopo.get("equipe_vigencia_nome")
        escopo["regiao"] = _normalizar_rotulo_foco(escopo.get("regiao") or escopo.get("equipe_vigencia_regiao"))
        por_funcionario[str(escopo["identificador_funcionario"])].append(escopo)
    for item in items:
        item["escopos_gestao"] = por_funcionario.get(str(item.get("identificador_funcionario")), [])
    return items


def _aplicar_lideranca_propria_por_cargo(item: dict[str, Any]) -> dict[str, Any]:
    if item.get("status_operacional") == "AFASTADO":
        return item

    destino_por_cargo = {
        "GERENTE": ("gestor", "gestor_email", "gestor_documento"),
        "COORDENADOR": ("coordenador", "coordenador_email", "coordenador_documento"),
        "GERENTE_REGIONAL": ("gerente", "gerente_email", "gerente_documento"),
    }
    destino = destino_por_cargo.get(_cargo_regra_equipe(item.get("cargo")))
    if not destino:
        return item

    nome = _nome_ou_none(item.get("nome"))
    if not nome:
        return item

    item[destino[0]] = nome
    item[destino[1]] = item.get("email")
    item[destino[2]] = item.get("documento")
    return item


async def resolver_perfil_acesso_padrao(conexao, item: dict[str, Any]) -> dict[str, Any]:
    codigo_cargo = _cargo_acesso_codigo(item.get("cargo"), item.get("tipo_funcionario"))
    linha = await conexao.fetchrow(
        """
        select
            regra.codigo_cargo,
            regra.nome_perfil,
            perfil.identificador_perfil,
            perfil.descricao_perfil
          from sevenlm_connect.funcionario_cargo_acesso_padrao regra
          left join sevenlm_connect.perfil perfil
            on perfil.nome_perfil = regra.nome_perfil
         where regra.codigo_cargo = $1
           and coalesce(regra.ativo, true) = true
         limit 1
        """,
        codigo_cargo,
    )
    if linha:
        return dict(linha)

    fallback = {
        "CORRETOR": "Corretor",
        "SDR": "Corretor",
        "COORDENADOR": "Coordenador Comercial",
        "GERENTE": "Gerente Comercial",
        "GERENTE_REGIONAL": "Gestor Comercial",
        "HEAD_COMERCIAL": "Diretor Comercial",
        "DIRETOR_COMERCIAL": "Diretor Comercial",
        "COLABORADOR": "Acesso Básico",
    }.get(codigo_cargo, "Acesso Básico")
    perfil = await conexao.fetchrow(
        """
        select identificador_perfil, nome_perfil, descricao_perfil
          from sevenlm_connect.perfil
         where nome_perfil = $1
         limit 1
        """,
        fallback,
    )
    return {
        "codigo_cargo": codigo_cargo,
        "nome_perfil": fallback,
        "identificador_perfil": perfil["identificador_perfil"] if perfil else None,
        "descricao_perfil": perfil["descricao_perfil"] if perfil else None,
    }


async def _atualizar_perfil_acesso_sugerido(
    conexao,
    *,
    identificador_funcionario: str,
    item: dict[str, Any],
    atualizado_por: Optional[str] = None,
    novo_cadastro: bool = False,
):
    perfil = await resolver_perfil_acesso_padrao(conexao, item)
    linha = await conexao.fetchrow(
        """
        update sevenlm_connect.funcionario_acesso
           set perfil_acesso_padrao = $2,
               identificador_perfil_acesso = case
                   when coalesce(perfil_acesso_excecao, false) = false then $3
                   else identificador_perfil_acesso
               end,
               criado_por_usuario = case
                   when $4::boolean is true then coalesce($5::uuid, criado_por_usuario)
                   else criado_por_usuario
               end,
               atualizado_por_usuario = coalesce($5::uuid, atualizado_por_usuario),
               data_hora_atualizado_em = now()
         where identificador_funcionario = $1::uuid
         returning *
        """,
        identificador_funcionario,
        perfil.get("nome_perfil"),
        perfil.get("identificador_perfil"),
        novo_cadastro,
        atualizado_por,
    )
    return linha


async def validar_funcionario_cadastro(
    conexao,
    *,
    identificador_funcionario: str,
    aprovado: bool,
    validado_por: str,
    observacao: Optional[str] = None,
    identificador_usuario: Optional[str] = None,
    identificador_perfil_acesso: Optional[int] = None,
    perfil_acesso_nome: Optional[str] = None,
    perfil_excecao: bool = False,
):
    status_validacao = "APROVADO" if aprovado else "REPROVADO"
    funcionario_atual = await conexao.fetchrow(
        """
        select *
          from sevenlm_connect.funcionario_acesso
         where identificador_funcionario = $1::uuid
         limit 1
        """,
        identificador_funcionario,
    )
    if not funcionario_atual:
        raise HTTPException(status_code=404, detail="Funcionario nao encontrado.")

    perfil_nome_final = perfil_acesso_nome
    perfil_id_final = identificador_perfil_acesso
    if aprovado and not perfil_id_final:
        perfil = await resolver_perfil_acesso_padrao(conexao, dict(funcionario_atual))
        perfil_nome_final = perfil.get("nome_perfil")
        perfil_id_final = perfil.get("identificador_perfil")

    ativo_login = bool(aprovado and funcionario_atual.get("ativo") is not False and funcionario_atual.get("email"))
    linha = await conexao.fetchrow(
        """
        update sevenlm_connect.funcionario_acesso
           set status_validacao = $2,
               validacao_observacao = $3,
               validado_por = $4::uuid,
               data_hora_validacao = now(),
               perfil_acesso_padrao = coalesce($5, perfil_acesso_padrao),
               identificador_perfil_acesso = $6,
               perfil_acesso_excecao = $7,
               perfil_acesso_excecao_por = case when $7 is true then $4::uuid else perfil_acesso_excecao_por end,
               data_hora_perfil_excecao = case when $7 is true then now() else data_hora_perfil_excecao end,
               identificador_usuario = coalesce($8::uuid, identificador_usuario),
               ativo_login = $9,
               atualizado_por_usuario = $4::uuid,
               data_hora_atualizado_em = now()
         where identificador_funcionario = $1::uuid
         returning *
        """,
        identificador_funcionario,
        status_validacao,
        _texto_ou_none(observacao),
        validado_por,
        perfil_nome_final,
        perfil_id_final,
        perfil_excecao,
        identificador_usuario,
        ativo_login,
    )
    await _sincronizar_usuario_vinculado(conexao, linha)
    return serializar_funcionario(linha)


async def salvar_funcionario(
    conexao,
    dados: dict[str, Any],
    identificador_funcionario: Optional[str] = None,
    atualizado_por: Optional[str] = None,
):
    escopos_gestao_recebidos = "escopos_gestao" in dados
    escopos_gestao_raw = dados.get("escopos_gestao")
    item = _normalizar_funcionario(dados)
    item = await _aplicar_equipe_vigente_ao_funcionario(conexao, item)
    if _cargo_exige_equipe(item.get("cargo")) and item.get("imobiliaria") and not item.get("identificador_equipe_vigencia"):
        raise HTTPException(status_code=422, detail="Selecione uma equipe com vigencia ativa para esta pessoa.")
    item = _aplicar_lideranca_propria_por_cargo(item)

    valores = (
        item["tipo_funcionario"],
        item["tipo_vinculo"],
        item["documento"],
        item["matricula"],
        item["email"],
        item["nome"],
        item["telefone"],
        item["cargo"],
        item["imobiliaria"],
        item["identificador_equipe_vigencia"],
        item["gestor_documento"],
        item["gestor_email"],
        item["gestor"],
        item["coordenador_documento"],
        item["coordenador_email"],
        item["coordenador"],
        item["gerente_documento"],
        item["gerente_email"],
        item["gerente"],
        item["diretor_documento"],
        item["diretor_email"],
        item["diretor"],
        item["regional"],
        item["regiao"],
        item["foco"],
        item["ativo_negocio"],
        item["ativo"],
        item["ativo_login"],
        item["data_cadastro_usuario"],
        item["data_admissao"],
        item["data_inicio_vigencia"],
        item["data_fim_vigencia"],
        item["referencia_origem"],
        item["origem_planilha"],
        item["cadastrado_por"],
        item["observacao"],
        item["identificador_usuario"],
        item["cnpj"],
        item["nome_empresa"],
    )

    if identificador_funcionario:
        atualizado = await conexao.fetchrow(
            """
            update sevenlm_connect.funcionario_acesso
               set tipo_funcionario = $2,
                   tipo_vinculo = $3,
                   documento = $4,
                   matricula = $5,
                   email = $6,
                   nome = $7,
                   telefone = $8,
                   cargo = $9,
                   imobiliaria = $10,
                   identificador_equipe_vigencia = $11::uuid,
                   gestor_documento = $12,
                   gestor_email = $13,
                   gestor = $14,
                   coordenador_documento = $15,
                   coordenador_email = $16,
                   coordenador = $17,
                   gerente_documento = $18,
                   gerente_email = $19,
                   gerente = $20,
                   diretor_documento = $21,
                   diretor_email = $22,
                   diretor = $23,
                   regional = $24,
                   regiao = $25,
                   foco = $26,
                   ativo_negocio = $27,
                   ativo = coalesce($28, true),
                   ativo_login = $29,
                   data_cadastro_usuario = $30,
                   data_admissao = $31,
                   data_inicio_vigencia = $32,
                   data_fim_vigencia = $33,
                   referencia_origem = $34,
                   origem_planilha = coalesce($35, origem_planilha),
                   cadastrado_por = $36,
                   observacao = $37,
                   identificador_usuario = coalesce($38::uuid, identificador_usuario),
                   cnpj = $39,
                   nome_empresa = $40,
                   atualizado_por_usuario = coalesce($41::uuid, atualizado_por_usuario),
                   data_hora_atualizado_em = now()
             where identificador_funcionario = $1::uuid
             returning *
            """,
            identificador_funcionario,
            *valores,
            atualizado_por,
        )
        if not atualizado:
            raise HTTPException(status_code=404, detail="Funcionario nao encontrado.")
        atualizado = await _atualizar_perfil_acesso_sugerido(
            conexao,
            identificador_funcionario=str(atualizado["identificador_funcionario"]),
            item=item,
            atualizado_por=atualizado_por,
            novo_cadastro=False,
        ) or atualizado
        await _sincronizar_usuario_vinculado(conexao, atualizado)
        await _sincronizar_lideranca_equipe_por_funcionario(conexao, item)
        funcionario_serializado = serializar_funcionario(atualizado)
        if escopos_gestao_recebidos:
            funcionario_serializado["escopos_gestao"] = await _salvar_funcionario_escopos_gestao(
                conexao,
                identificador_funcionario=str(atualizado["identificador_funcionario"]),
                cargo=item.get("cargo"),
                status_operacional=item.get("status_operacional"),
                escopos_raw=escopos_gestao_raw,
                atualizado_por=atualizado_por,
            )
        else:
            await _anexar_escopos_gestao_funcionarios(conexao, [funcionario_serializado])
        return funcionario_serializado, False

    existente = await conexao.fetchrow(
        """
        select *
          from sevenlm_connect.funcionario_acesso
         where ($1::text is not null and documento = $1)
            or ($2::text is not null and matricula = $2)
            or ($3::text is not null and lower(coalesce(email::text, '')) = lower($3))
         order by
           case when $1::text is not null and documento = $1 then 1 else 2 end,
           case when $2::text is not null and matricula = $2 then 1 else 2 end,
           data_hora_atualizado_em desc
         limit 1
        """,
        item["documento"],
        item["matricula"],
        item["email"],
    )

    if existente:
        item_para_atualizar = {**item}
        if escopos_gestao_recebidos:
            item_para_atualizar["escopos_gestao"] = escopos_gestao_raw
        atualizado, _ = await salvar_funcionario(
            conexao,
            item_para_atualizar,
            identificador_funcionario=str(existente["identificador_funcionario"]),
            atualizado_por=atualizado_por,
        )
        return atualizado, False

    novo = await conexao.fetchrow(
        """
        insert into sevenlm_connect.funcionario_acesso (
            tipo_funcionario,
            tipo_vinculo,
            documento,
            matricula,
            email,
            nome,
            telefone,
            cargo,
            imobiliaria,
            identificador_equipe_vigencia,
            gestor_documento,
            gestor_email,
            gestor,
            coordenador_documento,
            coordenador_email,
            coordenador,
            gerente_documento,
            gerente_email,
            gerente,
            diretor_documento,
            diretor_email,
            diretor,
            regional,
            regiao,
            foco,
            ativo_negocio,
            ativo,
            ativo_login,
            data_cadastro_usuario,
            data_admissao,
            data_inicio_vigencia,
            data_fim_vigencia,
            referencia_origem,
            origem_planilha,
            cadastrado_por,
            observacao,
            identificador_usuario,
            cnpj,
            nome_empresa,
            data_hora_importacao
        )
        values (
            $1, $2, $3, $4, $5, $6, $7, $8, $9, $10::uuid,
            $11, $12, $13, $14, $15, $16, $17, $18, $19, $20,
            $21, $22, $23, $24, $25, $26, coalesce($27, true), ($28::boolean and false), $29,
            $30, $31, $32, $33, $34, $35, $36, $37::uuid, $38, $39, now()
        )
        returning *
        """,
        *valores,
    )
    novo = await _atualizar_perfil_acesso_sugerido(
        conexao,
        identificador_funcionario=str(novo["identificador_funcionario"]),
        item=item,
        atualizado_por=atualizado_por,
        novo_cadastro=True,
    ) or novo
    await _sincronizar_usuario_vinculado(conexao, novo)
    await _sincronizar_lideranca_equipe_por_funcionario(conexao, item)
    funcionario_serializado = serializar_funcionario(novo)
    if escopos_gestao_recebidos:
        funcionario_serializado["escopos_gestao"] = await _salvar_funcionario_escopos_gestao(
            conexao,
            identificador_funcionario=str(novo["identificador_funcionario"]),
            cargo=item.get("cargo"),
            status_operacional=item.get("status_operacional"),
            escopos_raw=escopos_gestao_raw,
            atualizado_por=atualizado_por,
        )
    else:
        await _anexar_escopos_gestao_funcionarios(conexao, [funcionario_serializado])
    return funcionario_serializado, True


def _registro_normalizado(cabecalho: Any, linha: Any) -> dict[str, Any]:
    nomes_colunas = [_normalizar_chave(coluna) for coluna in cabecalho]
    indices = {nome: posicao for posicao, nome in enumerate(nomes_colunas) if nome}
    return {
        nome: linha[posicao] if posicao < len(linha) else None
        for nome, posicao in indices.items()
    }


def _inferir_tipo_funcionario(registro: dict[str, Any], sheet_name: str) -> str:
    texto = " ".join(
        _texto(registro.get(chave))
        for chave in ("cargo", "funcao", "departamento_obra", "departamento", "setor")
    ).upper()
    if sheet_name.lower().startswith("sdr") or "SDR" in texto:
        return "SDR"
    if "CORRETOR" in texto:
        return "CORRETOR"
    return "FUNCIONARIO"


def _nome_por_email(contexto: dict[str, Any], email: Any) -> Optional[str]:
    email_normalizado = _email_ou_none(email)
    if not email_normalizado:
        return None
    nome = contexto.get("nomes_por_email", {}).get(email_normalizado)
    if nome:
        return nome
    local = email_normalizado.split("@", 1)[0].replace(".", " ").replace("_", " ")
    return _nome_ou_none(local)


def _extrair_contexto_planilha(workbook: Any) -> dict[str, Any]:
    contexto = {
        "setores": {},
        "responsaveis": {},
        "nomes_por_email": {},
    }

    for worksheet in workbook.worksheets:
        linhas = worksheet.iter_rows(values_only=True)
        try:
            cabecalho = next(linhas)
        except StopIteration:
            continue

        titulo = _normalizar_chave(worksheet.title)
        for linha in linhas:
            registro = _registro_normalizado(cabecalho, linha)
            if titulo == "setor":
                setor = _texto(registro.get("setor") or registro.get("departamento_obra"))
                if setor:
                    contexto["setores"][setor.upper()] = {
                        "gerente_email": _email_ou_none(registro.get("gestor_1")),
                        "diretor_email": _email_ou_none(registro.get("gestor_2")),
                        "local": _texto_ou_none(registro.get("local")),
                    }
            elif titulo == "responsavel":
                setor = _texto(registro.get("setor") or registro.get("departamento_obra"))
                if setor:
                    contexto["responsaveis"][setor.upper()] = _email_ou_none(registro.get("adm"))
            elif titulo == "funcionarios":
                email = _email_ou_none(registro.get("email"))
                nome = _nome_ou_none(registro.get("colaborador") or registro.get("nome"))
                if email and nome:
                    contexto["nomes_por_email"][email] = nome

    return contexto


def _mapear_linha_planilha(sheet_name: str, registro: dict[str, Any], contexto: Optional[dict[str, Any]] = None) -> Optional[dict[str, Any]]:
    contexto = contexto or {}
    eh_sdr = sheet_name.lower().startswith("sdr")
    nome = registro.get("sdr") if eh_sdr else (registro.get("colaborador") or registro.get("nome"))
    email = registro.get("email_sdr") if eh_sdr else registro.get("email")
    documento = registro.get("documento_sdr") if eh_sdr else (registro.get("cpf") or registro.get("documento"))
    matricula = registro.get("coluna1") or registro.get("matricula")
    departamento = registro.get("departamento_obra") or registro.get("departamento") or registro.get("setor")
    setor = contexto.get("setores", {}).get(_texto(departamento).upper(), {})
    coordenador_email = registro.get("coordenador_email") or contexto.get("responsaveis", {}).get(_texto(departamento).upper())
    gerente_email = registro.get("gerente_email") or setor.get("gerente_email")
    diretor_email = registro.get("diretor_email") or setor.get("diretor_email")
    registro_desligado = bool(re.search(r"(^|[\s_-])(DESLIGAD[OA]|DESLIGAMENTO|INATIV[OA])\b", _texto(nome), flags=re.IGNORECASE))

    if not (nome or email or documento or matricula):
        return None

    return {
        "tipo_funcionario": _inferir_tipo_funcionario(registro, sheet_name),
        "tipo_vinculo": registro.get("pj_clt") or registro.get("tipo_vinculo") or registro.get("vinculo"),
        "documento": documento,
        "matricula": matricula,
        "email": email,
        "nome": _normalizar_nome(nome),
        "telefone": registro.get("telefone") or registro.get("celular") or registro.get("whatsapp"),
        "cargo": registro.get("cargo") or registro.get("funcao") or registro.get("função"),
        "imobiliaria": registro.get("imobiliaria"),
        "gestor_documento": registro.get("gestor_documento"),
        "gestor_email": registro.get("gestor_email"),
        "gestor": registro.get("gestor"),
        "coordenador_documento": registro.get("coordenador_documento"),
        "coordenador_email": coordenador_email,
        "coordenador": registro.get("coordenador") or _nome_por_email(contexto, coordenador_email),
        "gerente_documento": registro.get("gerente_documento"),
        "gerente_email": gerente_email,
        "gerente": registro.get("gerente") or _nome_por_email(contexto, gerente_email),
        "diretor_documento": registro.get("diretor_documento"),
        "diretor_email": diretor_email,
        "diretor": registro.get("diretor") or _nome_por_email(contexto, diretor_email),
        "regional": registro.get("regional") or departamento,
        "regiao": registro.get("regiao") or registro.get("obrasede") or registro.get("local") or setor.get("local"),
        "ativo_negocio": registro.get("ativo_negocio") if registro.get("ativo_negocio") is not None else not registro_desligado,
        "ativo": registro.get("ativo") if registro.get("ativo") is not None else not registro_desligado,
        "ativo_login": registro.get("ativo_login"),
        "data_cadastro_usuario": registro.get("data_cadastro_usuario"),
        "data_admissao": registro.get("data_admissao") or registro.get("admissao") or registro.get("admissão"),
        "data_inicio_vigencia": registro.get("data_inicio_vigencia"),
        "data_fim_vigencia": registro.get("data_fim_vigencia"),
        "referencia_origem": registro.get("referencia") or registro.get("codigo"),
        "origem_planilha": sheet_name,
        "cadastrado_por": registro.get("cadastrado_por"),
        "observacao": registro.get("observacao") or registro.get("observacoes") or registro.get("observações"),
    }


def _chave_funcionario_planilha(item: dict[str, Any]) -> str:
    documento = _documento_ou_none(item.get("documento"))
    if documento:
        return f"{item.get('tipo_funcionario')}:doc:{documento}"
    matricula = _matricula_ou_none(item.get("matricula"))
    if matricula:
        return f"{item.get('tipo_funcionario')}:matricula:{matricula}"
    email = _email_ou_none(item.get("email"))
    if email:
        return f"{item.get('tipo_funcionario')}:email:{email}"
    return f"{item.get('tipo_funcionario')}:nome:{_texto(item.get('nome')).lower()}"


def extrair_funcionarios_de_planilha(conteudo: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as erro:
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl indisponivel para importar planilha.") from erro

    try:
        workbook = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as erro:
        raise HTTPException(status_code=400, detail="Arquivo Excel invalido ou corrompido.") from erro

    contexto = _extrair_contexto_planilha(workbook)
    candidatos: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for worksheet in workbook.worksheets:
        linhas = worksheet.iter_rows(values_only=True)
        try:
            cabecalho = next(linhas)
        except StopIteration:
            continue

        for linha in linhas:
            registro = _registro_normalizado(cabecalho, linha)
            item = _mapear_linha_planilha(worksheet.title, registro, contexto)
            if not item:
                continue
            chave = _chave_funcionario_planilha(item)
            candidatos[chave].append(item)

    hoje = date.today()
    resultado = []

    for itens in candidatos.values():
        def chave_ordenacao(item: dict[str, Any]):
            inicio = _data_ou_none(item.get("data_inicio_vigencia")) or date.min
            fim = _data_ou_none(item.get("data_fim_vigencia")) or date.min
            vigente = inicio <= hoje and (fim == date.min or fim >= hoje)
            return (1 if vigente else 0, fim, inicio)

        escolhido = sorted(itens, key=chave_ordenacao)[-1]
        resultado.append(_normalizar_funcionario(escolhido))

    return sorted(resultado, key=lambda item: (item["tipo_funcionario"], item["nome"].lower()))


async def importar_funcionarios_planilha(
    conexao,
    conteudo: bytes,
    atualizado_por: Optional[str] = None,
) -> dict[str, Any]:
    funcionarios = extrair_funcionarios_de_planilha(conteudo)
    criados = 0
    atualizados = 0

    for item in funcionarios:
        _, criado = await salvar_funcionario(conexao, item, atualizado_por=atualizado_por)
        if criado:
            criados += 1
        else:
            atualizados += 1

    return {
        "total_lido": len(funcionarios),
        "criados": criados,
        "atualizados": atualizados,
    }


async def listar_funcionarios(
    conexao,
    *,
    termo: str = "",
    tipo: str = "",
    status: str = "",
    validacao: str = "",
    vinculo: str = "",
    imobiliaria: str = "",
    login: str = "",
    limite: int = 100,
):
    termo_like = f"%{_texto(termo)}%"
    tipo_normalizado = _tipo_funcionario(tipo) if _texto(tipo) else ""
    status_normalizado = _texto(status).upper()
    validacao_normalizada = _status_validacao(validacao, "") if _texto(validacao) else ""
    vinculo_normalizado = _tipo_vinculo(vinculo) or ""
    imobiliaria_like = f"%{_texto(imobiliaria)}%"
    login_normalizado = _status_login(login) if _texto(login) else ""
    await _garantir_funcionario_gestao_escopo(conexao)

    linhas = await conexao.fetch(
        """
        select
            f.*,
            eq.equipe as equipe_vigencia_nome,
            eq.regiao as equipe_vigencia_regiao,
            eq.gerente_vendas as equipe_gerente_vendas,
            eq.gerente_comercial as equipe_gerente_comercial,
            eq.gerente_regional as equipe_gerente_regional,
            eq.head_comercial as equipe_head_comercial,
            eq.diretor_comercial as equipe_diretor_comercial,
            eq.data_inicio_vigencia as equipe_data_inicio_vigencia,
            eq.data_fim_vigencia as equipe_data_fim_vigencia,
            eq.status_equipe as equipe_status,
            case
              when coalesce(f.ativo, true) = false then 'INATIVO'
              else coalesce(s.status_operacional, 'ATIVO')
            end as status_operacional,
            u.identificador_usuario::text as identificador_usuario_vinculado,
            u.indicador_ativo as usuario_ativo,
            perfil_acesso.nome_perfil as perfil_acesso_nome,
            coalesce(usuario_validacao.nome_completo, case when f.validado_por is null then null else 'Sistema' end) as validado_por_nome,
            coalesce(usuario_criacao.nome_completo, 'Sistema') as criado_por_nome,
            coalesce(usuario_atualizacao.nome_completo, 'Sistema') as atualizado_por_nome
        from sevenlm_connect.funcionario_acesso f
        left join sevenlm_connect.funcionario_status_diario s
          on s.identificador_funcionario = f.identificador_funcionario
         and s.data_status = current_date
        left join sevenlm_connect.funcionario_equipe_vigencia eq
          on eq.identificador_equipe_vigencia = f.identificador_equipe_vigencia
        left join sevenlm_connect.usuario u
          on u.identificador_usuario = f.identificador_usuario
        left join sevenlm_connect.perfil perfil_acesso
          on perfil_acesso.identificador_perfil = f.identificador_perfil_acesso
        left join sevenlm_connect.usuario usuario_validacao
          on usuario_validacao.identificador_usuario = f.validado_por
        left join sevenlm_connect.usuario usuario_criacao
          on usuario_criacao.identificador_usuario = f.criado_por_usuario
        left join sevenlm_connect.usuario usuario_atualizacao
          on usuario_atualizacao.identificador_usuario = f.atualizado_por_usuario
        where (
                $1 = '%%'
                or f.nome ilike $1
                or coalesce(f.email::text, '') ilike $1
                or coalesce(f.documento, '') ilike $1
                or coalesce(f.cnpj, '') ilike $1
                or coalesce(f.nome_empresa, '') ilike $1
                or coalesce(f.matricula, '') ilike $1
                or coalesce(f.telefone, '') ilike $1
                or coalesce(f.cargo, '') ilike $1
                or coalesce(f.imobiliaria, '') ilike $1
                or coalesce(f.gestor, '') ilike $1
                or coalesce(f.coordenador, '') ilike $1
                or coalesce(f.gerente, '') ilike $1
                or coalesce(f.diretor, '') ilike $1
                or coalesce(eq.equipe, '') ilike $1
                or coalesce(eq.gerente_vendas, '') ilike $1
                or coalesce(eq.gerente_comercial, '') ilike $1
                or coalesce(eq.gerente_regional, '') ilike $1
                or coalesce(eq.head_comercial, '') ilike $1
                or coalesce(eq.diretor_comercial, '') ilike $1
                or exists (
                    select 1
                      from sevenlm_connect.funcionario_gestao_escopo ge_busca
                     where ge_busca.identificador_funcionario = f.identificador_funcionario
                       and ge_busca.ativo = true
                       and (
                            coalesce(ge_busca.regiao, '') ilike $1
                            or coalesce(ge_busca.equipe, '') ilike $1
                       )
                )
              )
          and ($2 = '' or f.tipo_funcionario = $2)
          and ($8 = '' or coalesce(f.status_validacao, 'PENDENTE') = $8)
          and (
                $3 = ''
                or ($3 = 'ATIVO' and case when coalesce(f.ativo, true) = false then 'INATIVO' else coalesce(s.status_operacional, 'ATIVO') end = 'ATIVO')
                or ($3 = 'INATIVO' and case when coalesce(f.ativo, true) = false then 'INATIVO' else coalesce(s.status_operacional, 'ATIVO') end = 'INATIVO')
                or ($3 = 'AFASTADO' and case when coalesce(f.ativo, true) = false then 'INATIVO' else coalesce(s.status_operacional, 'ATIVO') end = 'AFASTADO')
                or ($3 = 'COM_LOGIN' and f.identificador_usuario is not null)
                or ($3 = 'SEM_LOGIN' and f.identificador_usuario is null)
              )
          and ($4 = '' or coalesce(f.tipo_vinculo, '') = $4)
          and (
                $5 = '%%'
                or coalesce(f.imobiliaria, '') ilike $5
                or exists (
                    select 1
                      from sevenlm_connect.funcionario_gestao_escopo ge_filtro
                     where ge_filtro.identificador_funcionario = f.identificador_funcionario
                       and ge_filtro.ativo = true
                       and (
                            coalesce(ge_filtro.equipe, '') ilike $5
                            or coalesce(ge_filtro.regiao, '') ilike $5
                       )
                )
              )
          and (
                $6 = ''
                or ($6 = 'LIBERADO' and coalesce(f.ativo, true) = true and coalesce(f.ativo_login, false) = true and coalesce(f.email::text, '') <> '')
                or ($6 in ('BLOQUEADO', 'SEM_LOGIN') and coalesce(f.ativo_login, false) = false and coalesce(f.email::text, '') <> '')
                or ($6 = 'SEM_EMAIL' and coalesce(f.email::text, '') = '')
                or ($6 = 'PENDENTE' and f.ativo_login is null)
              )
        order by
          case coalesce(f.status_validacao, 'PENDENTE') when 'PENDENTE' then 0 when 'REPROVADO' then 2 else 1 end,
          f.ativo desc,
          f.nome
        limit $7
        """,
        termo_like,
        tipo_normalizado,
        status_normalizado,
        vinculo_normalizado,
        imobiliaria_like,
        login_normalizado,
        limite,
        validacao_normalizada,
    )
    items = [serializar_funcionario(linha) for linha in linhas]
    return await _anexar_escopos_gestao_funcionarios(conexao, items)


async def obter_funcionario(conexao, identificador_funcionario: str) -> dict[str, Any]:
    await _garantir_funcionario_gestao_escopo(conexao)
    linha = await conexao.fetchrow(
        """
        select
            f.*,
            eq.equipe as equipe_vigencia_nome,
            eq.regiao as equipe_vigencia_regiao,
            eq.gerente_vendas as equipe_gerente_vendas,
            eq.gerente_comercial as equipe_gerente_comercial,
            eq.gerente_regional as equipe_gerente_regional,
            eq.head_comercial as equipe_head_comercial,
            eq.diretor_comercial as equipe_diretor_comercial,
            eq.data_inicio_vigencia as equipe_data_inicio_vigencia,
            eq.data_fim_vigencia as equipe_data_fim_vigencia,
            eq.status_equipe as equipe_status,
            perfil_acesso.nome_perfil as perfil_acesso_nome,
            coalesce(usuario_validacao.nome_completo, case when f.validado_por is null then null else 'Sistema' end) as validado_por_nome,
            coalesce(usuario_criacao.nome_completo, 'Sistema') as criado_por_nome,
            coalesce(usuario_atualizacao.nome_completo, 'Sistema') as atualizado_por_nome
          from sevenlm_connect.funcionario_acesso f
          left join sevenlm_connect.funcionario_equipe_vigencia eq
            on eq.identificador_equipe_vigencia = f.identificador_equipe_vigencia
          left join sevenlm_connect.perfil perfil_acesso
            on perfil_acesso.identificador_perfil = f.identificador_perfil_acesso
          left join sevenlm_connect.usuario usuario_validacao
            on usuario_validacao.identificador_usuario = f.validado_por
          left join sevenlm_connect.usuario usuario_criacao
            on usuario_criacao.identificador_usuario = f.criado_por_usuario
          left join sevenlm_connect.usuario usuario_atualizacao
            on usuario_atualizacao.identificador_usuario = f.atualizado_por_usuario
         where f.identificador_funcionario = $1::uuid
         limit 1
        """,
        identificador_funcionario,
    )
    if not linha:
        raise HTTPException(status_code=404, detail="Funcionario nao encontrado.")
    item = serializar_funcionario(linha)
    await _anexar_escopos_gestao_funcionarios(conexao, [item])
    return item


async def gerar_quadro_diario(conexao, data_status: date) -> int:
    total = await conexao.fetchval(
        """
        select sevenlm_connect.fn_gerar_funcionario_status_diario($1::date)
        """,
        data_status,
    )
    return int(total or 0)


async def listar_quadro_diario(conexao, data_status: date, termo: str = "", tipo: str = ""):
    await gerar_quadro_diario(conexao, data_status)

    termo_like = f"%{_texto(termo)}%"
    tipo_normalizado = _tipo_funcionario(tipo) if _texto(tipo) else ""

    linhas = await conexao.fetch(
        """
        select
            s.*,
            f.identificador_usuario::text as identificador_usuario,
            f.nome,
            f.email::text as email,
            f.documento,
            f.matricula,
            f.tipo_funcionario,
            f.tipo_vinculo,
            f.imobiliaria,
            f.gestor,
            f.coordenador,
            f.gerente,
            f.diretor,
            f.regional,
            f.regiao,
            f.foco
        from sevenlm_connect.funcionario_status_diario s
        join sevenlm_connect.funcionario_acesso f
          on f.identificador_funcionario = s.identificador_funcionario
        where s.data_status = $1::date
          and (
                $2 = '%%'
                or f.nome ilike $2
                or coalesce(f.email::text, '') ilike $2
                or coalesce(f.documento, '') ilike $2
                or coalesce(f.matricula, '') ilike $2
                or coalesce(f.imobiliaria, '') ilike $2
                or coalesce(f.gestor, '') ilike $2
                or coalesce(f.coordenador, '') ilike $2
                or coalesce(f.gerente, '') ilike $2
                or coalesce(f.diretor, '') ilike $2
              )
          and ($3 = '' or f.tipo_funcionario = $3)
        order by f.tipo_funcionario, f.nome
        """,
        data_status,
        termo_like,
        tipo_normalizado,
    )
    return [serializar_status_diario(linha) for linha in linhas]


async def atualizar_status_diario(
    conexao,
    *,
    identificador_status: str,
    status_operacional: Optional[str] = None,
    status_negocio: Optional[str] = None,
    status_login: Optional[str] = None,
    observacao: Optional[str] = None,
    atualizado_por: Optional[str] = None,
):
    status_login_normalizado = _status_login(status_login) if status_login is not None else None
    linha = await conexao.fetchrow(
        """
        update sevenlm_connect.funcionario_status_diario
           set status_operacional = coalesce($2, status_operacional),
               status_negocio = coalesce($3, status_negocio),
               status_login = coalesce($4, status_login),
               observacao = $5,
               atualizado_por = $6::uuid,
               data_hora_atualizado_em = now()
         where identificador_status = $1::uuid
         returning *
        """,
        identificador_status,
        _status_operacional(status_operacional) if status_operacional is not None else None,
        _texto_ou_none(status_negocio),
        status_login_normalizado,
        observacao,
        atualizado_por,
    )
    if not linha:
        raise HTTPException(status_code=404, detail="Linha do quadro diario nao encontrada.")

    if status_login_normalizado == "LIBERADO" and not linha.get("snapshot_email"):
        raise HTTPException(status_code=422, detail="Informe o e-mail do funcionario antes de liberar login.")

    if status_login_normalizado in {"LIBERADO", "BLOQUEADO", "SEM_LOGIN", "SEM_EMAIL"}:
        ativo_login = status_login_normalizado == "LIBERADO"
        funcionario = await conexao.fetchrow(
            """
            update sevenlm_connect.funcionario_acesso f
               set ativo_login = $2,
                   data_hora_atualizado_em = now()
             where f.identificador_funcionario = $1::uuid
             returning *
            """,
            str(linha["identificador_funcionario"]),
            ativo_login,
        )
        await _sincronizar_usuario_vinculado(conexao, funcionario)

    return serializar_status_diario(linha)


def status_padrao_funcionario(item: dict[str, Any]) -> dict[str, str]:
    if item.get("status_validacao") == "PENDENTE":
        status_login = "PENDENTE"
    elif item.get("ativo") is False:
        status_login = "BLOQUEADO"
    elif not item.get("email"):
        status_login = "SEM_EMAIL"
    elif item.get("ativo_login") is True:
        status_login = "LIBERADO"
    elif item.get("ativo_login") is False:
        status_login = "BLOQUEADO"
    else:
        status_login = "PENDENTE"
    return {
        "status_negocio": _bool_label(item.get("ativo_negocio"), "ATIVO", "INATIVO"),
        "status_login": status_login,
        "status_operacional": _bool_label(item.get("ativo"), "ATIVO", "INATIVO"),
    }
