"""
Ajusta as regras oficiais do ciclo Maio/2026 a partir da planilha de gestores.

Fonte oficial de configuracao:
  /root/data-engineering/apps/commercial-dashboard/Regras Gestores e Coordenadores.xlsx

O script nao altera valores financeiros dos resultados. Ele apenas versiona Regra 01/02
dos comissionados mapeados e inativa regras orfas de Marco Narciso.
"""

from __future__ import annotations

import argparse
import asyncio
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
import hashlib
import json
from pathlib import Path
import re
import sys
import unicodedata
from typing import Any

from openpyxl import load_workbook


BASE_API = Path(__file__).resolve().parents[1]
PORTAL_ROOT = BASE_API.parents[1]
if str(BASE_API) not in sys.path:
    sys.path.insert(0, str(BASE_API))

from banco import encerrar_pool_de_conexoes, iniciar_pool_de_conexoes  # noqa: E402
from configuracoes import ESQUEMA_COMISSIONAMENTO  # noqa: E402
from repositorios.comissionamento import registrar_evento_comissionamento, salvar_publicacao_regra  # noqa: E402


CICLO_ID = "2026-05"
AGENTE = "ajustar_maio_2026_regras_gestores_oficial.py"
PLANILHA_PADRAO = Path("/root/data-engineering/apps/commercial-dashboard/Regras Gestores e Coordenadores.xlsx")
REGISTRO_SAIDA = PORTAL_ROOT / "03_registros" / "comissionamento" / "execucoes"
FONTE_OFICIAL = "Regras Gestores e Coordenadores.xlsx"

MAPEAMENTO_ABAS = {
    "Josué - Gerente AGL": "Josué Gomes De Souza",
    "Ana Cleia - Gerente AGL ": "Ana Cleia Nonato",
    "Francisco- Gerente AGL": "Francisco Lucielio De Queiroz",
    "Alana - Gerente FSA": "Alana Rabelo Da Costa",
    "Daiana - Gerente FSA": "Daiana Soares Da Rocha",
    "Rafael - Gerente FSA": "Rafael De Lucena Martins",
    "Thomaz - Gerente Loja FSA": "Thomaz Moreira Aquino",
    "Marco Taveira - Gerente Geral": "Marco Taveira",
}

ROTULOS_FAIXA = (
    ("0% a 39,99%", Decimal("0"), Decimal("39.99")),
    ("40% a 59,99%", Decimal("40"), Decimal("59.99")),
    ("60% a 79,99%", Decimal("60"), Decimal("79.99")),
    ("80% a 94,99%", Decimal("80"), Decimal("94.99")),
    ("95% a 104,99%", Decimal("95"), Decimal("104.99")),
    ("105% a 109,99%", Decimal("105"), Decimal("109.99")),
    ("110% a 114,99%", Decimal("110"), Decimal("114.99")),
    ("115% a 119,99%", Decimal("115"), Decimal("119.99")),
    ("120% a 129,99%", Decimal("120"), Decimal("129.99")),
    ("130% a 139,99%", Decimal("130"), Decimal("139.99")),
    ("+ que 140%", Decimal("140"), None),
)


def _normalizar(valor: Any) -> str:
    texto = str(valor or "").strip()
    texto = "".join(
        caractere
        for caractere in unicodedata.normalize("NFD", texto)
        if unicodedata.category(caractere) != "Mn"
    )
    texto = re.sub(r"[^a-zA-Z0-9@.]+", " ", texto)
    return " ".join(texto.lower().split())


def _slug(valor: Any) -> str:
    texto = _normalizar(valor)
    return re.sub(r"[^a-z0-9]+", "_", texto).strip("_") or "indicador"


def _decimal(valor: Any) -> Decimal:
    if valor in (None, ""):
        return Decimal("0")
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).replace("R$", "").strip()
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except Exception:
        return Decimal("0")


def _moeda(valor: Any) -> Decimal:
    return _decimal(valor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _numero(valor: Decimal) -> float:
    return float(valor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _json_default(valor: Any) -> Any:
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, datetime):
        return valor.isoformat()
    return str(valor)


def _json_obj(valor: Any) -> dict[str, Any]:
    if isinstance(valor, dict):
        return valor
    if isinstance(valor, str) and valor.strip():
        try:
            parsed = json.loads(valor)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}


def _json_list(valor: Any) -> list[dict[str, Any]]:
    if isinstance(valor, list):
        return [item for item in valor if isinstance(item, dict)]
    if isinstance(valor, str) and valor.strip():
        try:
            parsed = json.loads(valor)
            return [item for item in parsed if isinstance(item, dict)] if isinstance(parsed, list) else []
        except Exception:
            return []
    return []


def _md5_curto(valor: str) -> str:
    return hashlib.md5(valor.encode("utf-8")).hexdigest()[:12]


def _faixa_padrao(rotulo: Any) -> tuple[str, Decimal, Decimal | None] | None:
    norm = _normalizar(rotulo)
    if norm in {"120", "120"} or ">=120" in str(rotulo or "") or norm == "120":
        return ("120% a 129,99%", Decimal("120"), Decimal("129.99"))
    if norm in {"que 140", "140"}:
        return ("+ que 140%", Decimal("140"), None)
    for item in ROTULOS_FAIXA:
        if _normalizar(item[0]) == norm:
            return item
    return None


def _faixa_por_percentual(percentual: Decimal) -> tuple[str, Decimal, Decimal | None]:
    for item in ROTULOS_FAIXA:
        _, minimo, maximo = item
        if percentual >= minimo and (maximo is None or percentual <= maximo):
            return item
    return ROTULOS_FAIXA[0]


@dataclass
class RegraOficial:
    aba: str
    nome_comissionado: str
    objetivo: Decimal
    faixas: list[dict[str, Any]]
    ips: list[dict[str, Any]]
    divergencias: list[str]


def _linha_texto(ws, linha: int, colunas: int = 8) -> str:
    return " ".join(str(ws.cell(linha, coluna).value) for coluna in range(1, colunas + 1) if ws.cell(linha, coluna).value not in (None, ""))


def _extrair_objetivo(ws) -> Decimal:
    for linha in range(1, ws.max_row + 1):
        if _normalizar(ws.cell(linha, 1).value) == "objetivo":
            valores = [
                _decimal(ws.cell(linha, coluna).value)
                for coluna in range(2, ws.max_column + 1)
                if _decimal(ws.cell(linha, coluna).value) > 0
            ]
            return sum(valores, Decimal("0"))
    return Decimal("0")


def _extrair_faixas(ws) -> list[dict[str, Any]]:
    faixas: list[dict[str, Any]] = []
    for linha in range(1, ws.max_row + 1):
        rotulo_raw = ws.cell(linha, 2).value
        faixa = _faixa_padrao(rotulo_raw)
        if not faixa:
            continue
        rotulo, minimo, maximo = faixa
        valor = _moeda(ws.cell(linha, 3).value)
        faixas.append({
            "id": f"faixa_{len(faixas) + 1}",
            "rotulo": rotulo,
            "percentual_minimo": _numero(minimo),
            "percentual_maximo": _numero(maximo) if maximo is not None else 999,
            "valor_bonus": _numero(valor),
            "valor_faixa": _numero(valor),
            "ativa": False,
            "fonte_linha": linha,
        })
    return faixas


def _parse_ip(descricao: str) -> tuple[str, str, Decimal]:
    texto = _normalizar(descricao)
    if "ipc" in texto:
        alvo = _decimal((re.search(r"ipc\s*([0-9]+(?:[,.][0-9]+)?)", descricao, re.I) or [None, "1"])[1])
        return "ipc", ">=", alvo
    if "sobrepreco" in texto:
        alvo = _decimal((re.search(r"r\$?\s*([0-9][0-9.,]*)", descricao, re.I) or [None, "0"])[1])
        return "sobrepreco_medio", ">=", alvo
    if "imobs com venda" in texto or "imob com venda" in texto:
        alvo = _decimal((re.search(r">=?\s*([0-9]+)", descricao) or [None, "1"])[1])
        return "imobiliarias_com_venda", ">=", alvo
    if "vendas canal" in texto or "venda canal" in texto:
        alvo = _decimal((re.search(r">=?\s*([0-9]+)", descricao) or [None, "1"])[1])
        return "vendas_canal", ">=", alvo
    if "repasse" in texto:
        alvo = _decimal((re.search(r"([0-9]+)\s*repasses?", descricao, re.I) or [None, "1"])[1])
        return "repasses", ">=", alvo
    return _slug(descricao), ">=", Decimal("1")


def _extrair_ips(ws, aba: str) -> list[dict[str, Any]]:
    ips: list[dict[str, Any]] = []
    for linha in range(1, ws.max_row + 1):
        marcador = str(ws.cell(linha, 1).value or "").strip()
        if not _normalizar(marcador).startswith("ip"):
            continue
        descricao = str(ws.cell(linha, 2).value or marcador).strip()
        valor = _moeda(ws.cell(linha + 1, 3).value)
        indicador, operador, alvo = _parse_ip(descricao)
        nome = f"{marcador} - {descricao}" if descricao and descricao != marcador else marcador
        ips.append({
            "ip_id": f"maio-2026-oficial-{_slug(aba)}-ip-{len(ips) + 1}",
            "nome": nome,
            "indicador": indicador,
            "indicador_texto": descricao or marcador,
            "tipo_comissao": "numero",
            "operador": operador,
            "alvo": _numero(alvo),
            "realizado": 0,
            "atingiu": False,
            "valor_bonus": _numero(valor),
            "data_inicial": "2026-05-01",
            "data_fim": "2026-05-31",
            "periodo_corte": "2026-05",
            "fonte_realizado": "atingimento_nao_inferido",
            "fonte_template": FONTE_OFICIAL,
            "fonte_linha": linha,
        })
    return ips


def _carregar_regras_oficiais(caminho: Path) -> dict[str, RegraOficial]:
    wb = load_workbook(caminho, read_only=True, data_only=True)
    regras: dict[str, RegraOficial] = {}
    for aba, nome in MAPEAMENTO_ABAS.items():
        if aba not in wb.sheetnames:
            regras[nome] = RegraOficial(aba, nome, Decimal("0"), [], [], [f"Aba nao encontrada: {aba}"])
            continue
        ws = wb[aba]
        divergencias: list[str] = []
        objetivo = _extrair_objetivo(ws)
        faixas = _extrair_faixas(ws)
        ips = _extrair_ips(ws, aba)
        if objetivo <= 0:
            divergencias.append("Objetivo oficial nao encontrado.")
        if not faixas:
            divergencias.append("Escada da Regra 01 nao encontrada.")
        if not ips:
            divergencias.append("IPs da Regra 02 nao encontrados.")
        if aba == "Thomaz - Gerente Loja FSA":
            titulo = _linha_texto(ws, 1)
            if "taveira" in _normalizar(titulo):
                divergencias.append("Titulo interno da aba aponta Marco Taveira; mapeamento usa Thomaz pelo nome da aba.")
        regras[nome] = RegraOficial(aba, nome, objetivo, faixas, ips, divergencias)
    return regras


def _compatibilidade_ip(ip_novo: dict[str, Any], ips_atuais: list[dict[str, Any]]) -> dict[str, Any] | None:
    chave_nova = _normalizar(f"{ip_novo.get('nome')} {ip_novo.get('indicador')} {ip_novo.get('indicador_texto')}")
    for atual in ips_atuais or []:
        chave_atual = _normalizar(f"{atual.get('nome')} {atual.get('indicador')} {atual.get('indicador_texto')}")
        if chave_nova and chave_atual and (chave_nova in chave_atual or chave_atual in chave_nova):
            return atual
    return None


def _montar_regra_01(resultado: dict[str, Any], regra_atual: dict[str, Any], oficial: RegraOficial) -> dict[str, Any]:
    objetivo = oficial.objetivo
    realizado = _decimal(regra_atual.get("realizado"))
    percentual = (realizado / objetivo * Decimal("100")).quantize(Decimal("0.01")) if objetivo > 0 else Decimal("0")
    faixa_aplicada, _, _ = _faixa_por_percentual(percentual)
    faixas = []
    valor_calculado = Decimal("0")
    for item in oficial.faixas:
        faixa = dict(item)
        faixa["realizado_atual"] = _numero(realizado)
        faixa["objetivo_base"] = _numero(objetivo)
        minimo = _decimal(faixa.get("percentual_minimo"))
        maximo = _decimal(faixa.get("percentual_maximo")) if faixa.get("percentual_maximo") != 999 else None
        necessario_minimo = objetivo * minimo / Decimal("100") if objetivo > 0 else Decimal("0")
        necessario_maximo = objetivo * maximo / Decimal("100") if objetivo > 0 and maximo is not None else None
        faixa["necessario_minimo"] = _numero(necessario_minimo)
        faixa["necessario_maximo"] = None if necessario_maximo is None else _numero(necessario_maximo)
        faixa["faltam_para_minimo"] = _numero(max(necessario_minimo - realizado, Decimal("0")))
        faixa["ativa"] = _normalizar(faixa.get("rotulo")) == _normalizar(faixa_aplicada)
        if faixa["ativa"]:
            valor_calculado = _moeda(faixa.get("valor_bonus"))
        faixas.append(faixa)
    return {
        "meta_id": f"{resultado['resultado_id']}-r01-maio-2026-oficial",
        "nome": "Regra 01 - Maio/2026 Oficial",
        "indicador": regra_atual.get("indicador") or "repasses",
        "substituir_faixas": True,
        "objetivo": _numero(objetivo),
        "realizado": _numero(realizado),
        "peso": 1,
        "percentual_atingimento": _numero(percentual),
        "faixa_aplicada": faixa_aplicada,
        "valor_faixa": _numero(valor_calculado),
        "valor_calculado": _numero(valor_calculado),
        "faixas": faixas,
        "fonte_realizado": regra_atual.get("fonte_realizado") or "preservado_regra_atual_maio",
        "fonte_template": FONTE_OFICIAL,
        "aba_fonte": oficial.aba,
        "realizado_preservado_de": regra_atual.get("meta_id") or regra_atual.get("nome") or "regra_atual",
    }


def _montar_regra_02(resultado: dict[str, Any], regra_atual: dict[str, Any], ips_atuais: list[dict[str, Any]], oficial: RegraOficial) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    ips: list[dict[str, Any]] = []
    for item in oficial.ips:
        ip = dict(item)
        compativel = _compatibilidade_ip(ip, ips_atuais)
        if compativel:
            ip["realizado"] = float(_decimal(compativel.get("realizado")))
            ip["atingiu"] = bool(compativel.get("atingiu"))
            ip["fonte_realizado"] = compativel.get("fonte_realizado") or "preservado_regra_atual_maio"
            ip["compatibilidade"] = "preservado_por_nome_indicador"
        else:
            ip["compatibilidade"] = "atingimento_nao_inferido"
        ips.append(ip)
    total = sum((_decimal(ip.get("valor_bonus")) for ip in ips if ip.get("atingiu")), Decimal("0"))
    regra_02 = {
        "meta_id": f"{resultado['resultado_id']}-r02-maio-2026-oficial",
        "nome": "Regra 02 - IPs Maio/2026 Oficial",
        "indicador": "ips_gestores_coordenadores",
        "objetivo": len(ips),
        "realizado": sum(1 for ip in ips if ip.get("atingiu")),
        "valor_calculado": _numero(total),
        "ips": ips,
        "fonte_template": FONTE_OFICIAL,
        "aba_fonte": oficial.aba,
        "observacao": regra_atual.get("observacao") or "Atingimento preservado quando compativel; demais IPs sem inferencia.",
    }
    return regra_02, ips


async def _carregar_estado(conexao) -> tuple[list[dict[str, Any]], dict[str, dict[str, dict[str, Any]]]]:
    resultados = [dict(linha) for linha in await conexao.fetch(
        f"""
        select *
        from {ESQUEMA_COMISSIONAMENTO}.resultados
        where ciclo_id = $1
        order by nome
        """,
        CICLO_ID,
    )]
    regras: dict[str, dict[str, dict[str, Any]]] = {}
    linhas = await conexao.fetch(
        f"""
        select *
        from {ESQUEMA_COMISSIONAMENTO}.regras_publicadas
        where ciclo_id = $1 and ativo is true
        order by comissionado_nome, regra_tipo
        """,
        CICLO_ID,
    )
    for linha in linhas:
        item = dict(linha)
        regras.setdefault(str(item.get("comissionado_id") or ""), {})[str(item.get("regra_tipo"))] = item
        regras.setdefault(_normalizar(item.get("comissionado_nome")), {})[str(item.get("regra_tipo"))] = item
    return resultados, regras


async def ajustar(aplicar: bool, planilha: Path) -> dict[str, Any]:
    regras_oficiais = _carregar_regras_oficiais(planilha)
    pool = await iniciar_pool_de_conexoes()
    try:
        async with pool.acquire() as conexao:
            resultados, regras_atuais = await _carregar_estado(conexao)
            resultado_por_nome = {_normalizar(item.get("nome")): item for item in resultados}
            total_antes = {
                "resultados": len(resultados),
                "valor_bruto": _numero(sum(_moeda(item.get("valor_bruto")) for item in resultados)),
                "valor_liquido": _numero(sum(_moeda(item.get("valor_liquido")) for item in resultados)),
            }
            ajustados: list[dict[str, Any]] = []
            preservados: list[dict[str, Any]] = []
            novas_regras: list[dict[str, Any]] = []

            for nome, oficial in regras_oficiais.items():
                resultado = resultado_por_nome.get(_normalizar(nome))
                if not resultado:
                    preservados.append({
                        "nome": nome,
                        "motivo": "regra_oficial_sem_resultado_no_ciclo",
                        "aba": oficial.aba,
                        "divergencias": oficial.divergencias,
                    })
                    continue
                atuais = regras_atuais.get(str(resultado["resultado_id"])) or regras_atuais.get(_normalizar(nome)) or {}
                regra_01_linha = atuais.get("regra_01") or {}
                regra_02_linha = atuais.get("regra_02") or {}
                regra_01_atual = _json_obj(regra_01_linha.get("regra_01"))
                regra_02_atual = _json_obj(regra_02_linha.get("regra_02"))
                ips_atuais = _json_list(regra_02_linha.get("regra_02_ips"))
                regra_01 = _montar_regra_01(resultado, regra_01_atual, oficial)
                regra_02, ips = _montar_regra_02(resultado, regra_02_atual, ips_atuais, oficial)
                novas_regras.append({
                    "resultado": resultado,
                    "oficial": oficial,
                    "regra_01": regra_01,
                    "regra_02": regra_02,
                    "ips": ips,
                })
                ajustados.append({
                    "resultado_id": resultado["resultado_id"],
                    "nome": nome,
                    "aba": oficial.aba,
                    "objetivo_oficial": _numero(oficial.objetivo),
                    "realizado_preservado": regra_01["realizado"],
                    "percentual_recalculado": regra_01["percentual_atingimento"],
                    "faixa_aplicada": regra_01["faixa_aplicada"],
                    "valor_regra_01_oficial": regra_01["valor_calculado"],
                    "ips_oficiais": len(ips),
                    "ips_sem_inferencia": sum(1 for ip in ips if ip.get("compatibilidade") == "atingimento_nao_inferido"),
                    "divergencias": oficial.divergencias,
                })

            nomes_oficiais = {_normalizar(nome) for nome in regras_oficiais}
            for resultado in resultados:
                if _normalizar(resultado.get("nome")) not in nomes_oficiais:
                    preservados.append({
                        "resultado_id": resultado["resultado_id"],
                        "nome": resultado.get("nome"),
                        "motivo": "sem_regra_oficial_xlsx",
                    })

            orfas = [dict(linha) for linha in await conexao.fetch(
                f"""
                select rp.regra_publicada_id::text as regra_publicada_id,
                       rp.comissionado_id,
                       rp.comissionado_nome,
                       rp.regra_tipo
                from {ESQUEMA_COMISSIONAMENTO}.regras_publicadas rp
                left join {ESQUEMA_COMISSIONAMENTO}.resultados r
                  on r.ciclo_id = rp.ciclo_id
                 and r.resultado_id = rp.comissionado_id
                where rp.ciclo_id = $1
                  and rp.ativo is true
                  and r.resultado_id is null
                order by rp.comissionado_nome, rp.regra_tipo
                """,
                CICLO_ID,
            )]

            relatorio = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "modo": "apply" if aplicar else "dry_run",
                "ciclo_id": CICLO_ID,
                "planilha": str(planilha),
                "fonte_oficial": FONTE_OFICIAL,
                "total_antes": total_antes,
                "total_depois": total_antes,
                "ajustados": ajustados,
                "preservados": preservados,
                "orfas_para_inativar": orfas,
                "resumo": {
                    "resultados": len(resultados),
                    "regras_oficiais_mapeadas": len(novas_regras),
                    "preservados": len(preservados),
                    "orfas": len(orfas),
                    "valor_bruto_preservado": total_antes["valor_bruto"],
                    "valor_liquido_preservado": total_antes["valor_liquido"],
                },
            }

            if aplicar:
                async with conexao.transaction():
                    for item in novas_regras:
                        resultado = item["resultado"]
                        oficial = item["oficial"]
                        for regra_tipo in ("regra_01", "regra_02"):
                            await salvar_publicacao_regra(
                                conexao,
                                ESQUEMA_COMISSIONAMENTO,
                                ciclo_id=CICLO_ID,
                                comissionado_id=resultado["resultado_id"],
                                comissionado_nome=resultado.get("nome"),
                                regra_tipo=regra_tipo,
                                regra_01=item["regra_01"] if regra_tipo == "regra_01" else {},
                                regra_02=item["regra_02"] if regra_tipo == "regra_02" else {},
                                regra_02_ips=item["ips"] if regra_tipo == "regra_02" else [],
                                regra_02_ips_removidos=[],
                                payload={
                                    "fonte": AGENTE,
                                    "fonte_oficial": FONTE_OFICIAL,
                                    "aba": oficial.aba,
                                    "financeiro_preservado": True,
                                    "regra_01": item["regra_01"] if regra_tipo == "regra_01" else None,
                                    "regra_02": item["regra_02"] if regra_tipo == "regra_02" else None,
                                },
                                motivo="Ajuste oficial Maio/2026 pela planilha Regras Gestores e Coordenadores.",
                                comentario="Financeiro preservado; realizado atual mantido; regras oficiais versionadas.",
                                usuario_id=None,
                            )
                    if orfas:
                        await conexao.execute(
                            f"""
                            update {ESQUEMA_COMISSIONAMENTO}.regras_publicadas rp
                            set ativo = false
                            where rp.ciclo_id = $1
                              and rp.ativo is true
                              and not exists (
                                select 1
                                from {ESQUEMA_COMISSIONAMENTO}.resultados r
                                where r.ciclo_id = rp.ciclo_id
                                  and r.resultado_id = rp.comissionado_id
                              )
                            """,
                            CICLO_ID,
                        )
                    await registrar_evento_comissionamento(
                        conexao,
                        ESQUEMA_COMISSIONAMENTO,
                        tipo_evento="ciclo_maio_2026_regras_gestores_oficial_ajustado",
                        usuario_id=None,
                        sessao_id=None,
                        endereco_ip="127.0.0.1",
                        agente_do_usuario=AGENTE,
                        antes={"ciclo_id": CICLO_ID, **total_antes},
                        depois={"ciclo_id": CICLO_ID, **total_antes, "ajustados": len(ajustados), "orfas_inativadas": len(orfas)},
                        comentario="Ajuste controlado de regras oficiais de Maio/2026 sem alterar financeiro.",
                    )

            REGISTRO_SAIDA.mkdir(parents=True, exist_ok=True)
            destino = REGISTRO_SAIDA / f"comissionamento_maio_2026_regras_gestores_oficial_{'apply' if aplicar else 'dry_run'}.json"
            destino.write_text(json.dumps(relatorio, ensure_ascii=False, indent=2, default=_json_default), encoding="utf-8")
            relatorio["arquivo"] = str(destino)
            return relatorio
    finally:
        await encerrar_pool_de_conexoes()


async def main() -> None:
    parser = argparse.ArgumentParser(description="Ajusta Maio/2026 com Regras Gestores e Coordenadores.xlsx.")
    parser.add_argument("--apply", action="store_true", help="Grava as alteracoes. Sem esta opcao, roda dry-run.")
    parser.add_argument("--planilha", type=Path, default=PLANILHA_PADRAO)
    args = parser.parse_args()
    relatorio = await ajustar(args.apply, args.planilha)
    print(json.dumps({
        "modo": relatorio["modo"],
        "ciclo_id": relatorio["ciclo_id"],
        "resumo": relatorio["resumo"],
        "arquivo": relatorio["arquivo"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    asyncio.run(main())
