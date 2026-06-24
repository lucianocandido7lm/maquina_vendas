"""
Importa a planilha manual de repasses/historico mensal para a produtividade G&C.

Uso:
  python importar_dashboard_gc_produtividade_historico_manual.py --arquivo C:\\caminho\\repasses.6.6.xlsx
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import math
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import asyncpg
from openpyxl import load_workbook


API_DIR = Path(__file__).resolve().parents[1]
PORTAL_ROOT = API_DIR.parents[1]
if str(API_DIR) not in sys.path:
    sys.path.insert(0, str(API_DIR))

from configuracoes import (  # noqa: E402
    ESQUEMA_COMERCIAL,
    NOME_BANCO,
    PORTA_BANCO,
    SENHA_BANCO,
    SERVIDOR_BANCO,
    USUARIO_BANCO,
)


SHEET_BASE = "20260606112522_6a242dd21911c"
SHEET_HISTORICO = "dinamica"
ANO_HISTORICO_PADRAO = 2026
STATUS_INVALIDO_RE = re.compile(r"DISTRATO|CANCEL", re.IGNORECASE)
COLUNAS_BASE_REPASSES = {
    "situacao_do_repasse",
    "data_de_assinatura",
    "reserva",
    "regiao",
    "empreendimento",
    "imobiliaria",
    "corretor",
    "cliente",
    "documento_do_cliente",
}


def _quote_ident(valor: str) -> str:
    return '"' + str(valor).replace('"', '""') + '"'


def _qualificar(schema: str, tabela: str) -> str:
    return f"{_quote_ident(schema)}.{_quote_ident(tabela)}"


def _normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and math.isnan(valor):
        return ""
    return str(valor).strip()


def _slug(valor: Any) -> str:
    texto = "".join(
        caractere
        for caractere in unicodedata.normalize("NFD", _normalizar_texto(valor))
        if unicodedata.category(caractere) != "Mn"
    ).lower()
    return re.sub(r"[^a-z0-9]+", "_", texto).strip("_")


def _slug_corretor(valor: Any) -> str:
    slug = _slug(valor)
    for padrao in (
        r"_(?:clt_)?desligad[oa]s?$",
        r"_clt$",
        r"_pj$",
        r"_autonom[oa]$",
    ):
        slug = re.sub(padrao, "", slug).strip("_")
    return slug


def _chave_registro(*partes: Any) -> str:
    bruto = "|".join(_normalizar_texto(parte) for parte in partes)
    return hashlib.sha1(bruto.encode("utf-8", errors="replace")).hexdigest()


def _mes_ref(data_assinatura: date) -> date:
    return date(data_assinatura.year, data_assinatura.month, 1)


def _colunas_normalizadas(cabecalho: list[Any] | tuple[Any, ...]) -> dict[str, int]:
    return {
        _slug(coluna): indice
        for indice, coluna in enumerate(cabecalho)
        if _normalizar_texto(coluna)
    }


def _pegar_coluna(colunas: dict[str, int], nome: str) -> int:
    coluna = colunas.get(nome)
    if coluna is None:
        raise RuntimeError(f"Coluna obrigatoria ausente na planilha: {nome}")
    return coluna


def _detectar_aba_por_colunas(caminho: Path, colunas_obrigatorias: set[str], preferida: str | None = None) -> str:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    if preferida and preferida in workbook.sheetnames:
        ws_preferida = workbook[preferida]
        colunas_preferida = _colunas_normalizadas(next(ws_preferida.iter_rows(min_row=1, max_row=1, values_only=True)))
        if colunas_obrigatorias.issubset(set(colunas_preferida)):
            return preferida

    for nome_aba in workbook.sheetnames:
        ws = workbook[nome_aba]
        colunas = _colunas_normalizadas(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        if colunas_obrigatorias.issubset(set(colunas)):
            return nome_aba
    raise RuntimeError(
        "Nenhuma aba da planilha possui as colunas obrigatorias da base de repasses."
    )


def _data_planilha(valor: Any) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = _normalizar_texto(valor)
    if not texto:
        return None
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(texto[:10], formato).date()
        except ValueError:
            continue
    return None


def _extrair_repasses(caminho: Path, fonte: str) -> list[tuple[Any, ...]]:
    sheet_base = _detectar_aba_por_colunas(caminho, COLUNAS_BASE_REPASSES, SHEET_BASE)
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    ws = workbook[sheet_base]
    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas)
    colunas = _colunas_normalizadas(cabecalho)
    col_status = _pegar_coluna(colunas, "situacao_do_repasse")
    col_data = _pegar_coluna(colunas, "data_de_assinatura")
    col_reserva = _pegar_coluna(colunas, "reserva")
    col_regiao = _pegar_coluna(colunas, "regiao")
    col_empreendimento = _pegar_coluna(colunas, "empreendimento")
    col_imobiliaria = _pegar_coluna(colunas, "imobiliaria")
    col_corretor = _pegar_coluna(colunas, "corretor")
    col_cliente = _pegar_coluna(colunas, "cliente")
    col_documento = _pegar_coluna(colunas, "documento_do_cliente")

    registros: list[tuple[Any, ...]] = []
    for linha in linhas:
        data_assinatura = _data_planilha(linha[col_data] if col_data < len(linha) else None)
        if not data_assinatura:
            continue
        status = _normalizar_texto(linha[col_status] if col_status < len(linha) else None)
        if STATUS_INVALIDO_RE.search(status):
            continue
        reserva = _normalizar_texto(linha[col_reserva] if col_reserva < len(linha) else None)
        imobiliaria = _normalizar_texto(linha[col_imobiliaria] if col_imobiliaria < len(linha) else None) or "Sem equipe"
        corretor = _normalizar_texto(linha[col_corretor] if col_corretor < len(linha) else None) or "Sem corretor"
        chave = _chave_registro(fonte, reserva, data_assinatura.isoformat(), imobiliaria, corretor)
        registros.append(
            (
                chave,
                reserva,
                data_assinatura,
                _mes_ref(data_assinatura),
                status,
                _normalizar_texto(linha[col_regiao] if col_regiao < len(linha) else None),
                _normalizar_texto(linha[col_empreendimento] if col_empreendimento < len(linha) else None),
                imobiliaria,
                corretor,
                _normalizar_texto(linha[col_cliente] if col_cliente < len(linha) else None),
                _normalizar_texto(linha[col_documento] if col_documento < len(linha) else None),
                fonte,
            )
        )
    return registros


async def _mapear_equipes(conn: asyncpg.Connection) -> dict[str, dict[str, str]]:
    mapa: dict[str, dict[str, str]] = {}
    try:
        linhas_hierarquia = await conn.fetch(
            f"""
            select distinct on (lower(imobiliaria_corretor))
                   imobiliaria_corretor as equipe,
                   regiao_corretor as regiao,
                   coalesce(nullif(trim(gerente_nome), ''), nullif(trim(gestor_corretor), '')) as gerente,
                   coalesce(nullif(trim(coordenador_nome), ''), nullif(trim(coordenador_corretor), '')) as coordenador,
                   mes_referencia
              from {_qualificar(ESQUEMA_COMERCIAL, "dashboard_gc_produtividade_hierarquia")}
             where imobiliaria_corretor is not null
             order by lower(imobiliaria_corretor), mes_referencia desc
            """
        )
    except Exception:
        linhas_hierarquia = []

    for linha in linhas_hierarquia:
        chave = _slug(linha["equipe"])
        if not chave or chave in mapa:
            continue
        mapa[chave] = {
            "regiao": _normalizar_texto(linha["regiao"]),
            "gerente": _normalizar_texto(linha["gerente"]) or "Sem gerente",
            "coordenador": _normalizar_texto(linha["coordenador"]) or "Sem coordenador",
        }

    linhas = await conn.fetch(
        """
        select equipe, regiao, gerente_vendas, gerente_comercial, status_equipe, ativo, data_inicio_vigencia
          from sevenlm_connect.funcionario_equipe_vigencia
         order by lower(equipe), data_inicio_vigencia desc
        """
    )
    for linha in linhas:
        chave = _slug(linha["equipe"])
        if not chave or chave in mapa:
            continue
        mapa[chave] = {
            "regiao": _normalizar_texto(linha["regiao"]),
            "gerente": _normalizar_texto(linha["gerente_vendas"]) or "Sem gerente",
            "coordenador": _normalizar_texto(linha["gerente_comercial"]) or "Sem coordenador",
        }
    return mapa


def _tipo_por_equipe(equipe: str) -> str:
    chave = _slug(equipe)
    if "imobiliaria" in chave:
        return "Imobiliaria"
    if "canal_virtual" in chave:
        return "Canal virtual"
    if "autonom" in chave:
        return "Autonomo"
    return "Equipe propria"


def _meses_entre(inicio: date, fim: date) -> list[date]:
    meses: list[date] = []
    atual = inicio
    while atual <= fim:
        meses.append(atual)
        proximo_indice = atual.year * 12 + atual.month
        atual = date(proximo_indice // 12, (proximo_indice % 12) + 1, 1)
    return meses


def _extrair_historico(caminho: Path, ano: int, fonte: str, equipes: dict[str, dict[str, str]]) -> list[tuple[Any, ...]]:
    workbook = load_workbook(caminho, read_only=True, data_only=True)
    if SHEET_HISTORICO not in workbook.sheetnames:
        return []
    ws = workbook[SHEET_HISTORICO]
    meses = [date(ano, mes, 1) for mes in range(1, 7)]
    registros: list[tuple[Any, ...]] = []

    for linha in ws.iter_rows(values_only=True):
        equipe = _normalizar_texto(linha[0] if len(linha) > 0 else None)
        corretor = _normalizar_texto(linha[1] if len(linha) > 1 else None)
        if not equipe or not corretor:
            continue
        if _slug(equipe) in {"situacao_do_repasse", "contagem_de_reserva", "imobiliaria"}:
            continue

        contagens: dict[date, int] = {}
        for offset, mes_ref in enumerate(meses, start=2):
            valor = linha[offset] if offset < len(linha) else None
            if not _normalizar_texto(valor):
                contagens[mes_ref] = 0
                continue
            try:
                contagens[mes_ref] = int(float(valor))
            except Exception:
                contagens[mes_ref] = 0

        meses_com_repasse = [mes_ref for mes_ref, valor in contagens.items() if valor > 0]
        if not meses_com_repasse:
            continue

        chave_equipe = _slug(equipe)
        chave_corretor = _slug_corretor(corretor)
        mapa_equipe = equipes.get(chave_equipe, {})
        for mes_ref in _meses_entre(min(meses_com_repasse), max(meses_com_repasse)):
            registros.append(
                (
                    mes_ref,
                    equipe,
                    chave_equipe,
                    corretor,
                    chave_corretor,
                    mapa_equipe.get("gerente") or "Sem gerente",
                    mapa_equipe.get("coordenador") or "Sem coordenador",
                    mapa_equipe.get("regiao") or "Sem regiao",
                    _tipo_por_equipe(equipe),
                    int(contagens.get(mes_ref) or 0),
                    True,
                    fonte,
                )
            )
    return registros


async def _garantir_tabelas(conn: asyncpg.Connection) -> None:
    caminho_migracao = PORTAL_ROOT / "Servidor" / "migracao_20260606_dashboard_gc_produtividade_historico_manual.sql"
    await conn.execute(caminho_migracao.read_text(encoding="utf-8"))


async def _copiar_registros(
    conn: asyncpg.Connection,
    *,
    tabela: str,
    colunas: list[str],
    registros: list[tuple[Any, ...]],
    campo_fonte: str,
    fonte: str,
) -> int:
    tabela_qualificada = _qualificar(ESQUEMA_COMERCIAL, tabela)
    tabela_temp = f"tmp_{tabela}"
    await conn.execute(f"create temp table {_quote_ident(tabela_temp)} (like {tabela_qualificada} including defaults) on commit drop")
    if registros:
        await conn.copy_records_to_table(tabela_temp, records=registros, columns=colunas)
    await conn.execute(f"delete from {tabela_qualificada} where {_quote_ident(campo_fonte)} = $1", fonte)
    if registros:
        lista_colunas = ", ".join(_quote_ident(coluna) for coluna in colunas)
        await conn.execute(
            f"""
            insert into {tabela_qualificada} ({lista_colunas})
            select {lista_colunas}
              from {_quote_ident(tabela_temp)}
            """
        )
    return int(await conn.fetchval(f"select count(*)::bigint from {tabela_qualificada} where {_quote_ident(campo_fonte)} = $1", fonte))


async def _copiar_repasses_importados(
    conn: asyncpg.Connection,
    *,
    registros: list[tuple[Any, ...]],
    fonte: str,
) -> int:
    tabela = "dashboard_gc_produtividade_repasses_importados"
    tabela_qualificada = _qualificar(ESQUEMA_COMERCIAL, tabela)
    tabela_temp = f"tmp_{tabela}"
    colunas = [
        "chave_registro",
        "reserva",
        "data_assinatura",
        "mes_referencia",
        "situacao_repasse",
        "regiao",
        "empreendimento",
        "imobiliaria",
        "corretor",
        "cliente",
        "documento_cliente",
        "fonte_arquivo",
    ]
    await conn.execute(f"create temp table {_quote_ident(tabela_temp)} (like {tabela_qualificada} including defaults) on commit drop")
    if registros:
        await conn.copy_records_to_table(tabela_temp, records=registros, columns=colunas)
        await conn.execute(
            f"""
            delete from {tabela_qualificada} destino
             where exists (
                   select 1
                     from {_quote_ident(tabela_temp)} origem
                    where origem.mes_referencia = destino.mes_referencia
             )
            """
        )
        lista_colunas = ", ".join(_quote_ident(coluna) for coluna in colunas)
        await conn.execute(
            f"""
            insert into {tabela_qualificada} ({lista_colunas})
            select {lista_colunas}
              from {_quote_ident(tabela_temp)}
            """
        )
    else:
        await conn.execute(f"delete from {tabela_qualificada} where fonte_arquivo = $1", fonte)
    return int(await conn.fetchval(f"select count(*)::bigint from {tabela_qualificada} where fonte_arquivo = $1", fonte))


async def importar(caminho: Path, ano: int) -> dict[str, int]:
    fonte = caminho.name
    repasses = _extrair_repasses(caminho, fonte)
    conn = await asyncpg.connect(
        user=USUARIO_BANCO,
        password=SENHA_BANCO,
        host=SERVIDOR_BANCO,
        port=PORTA_BANCO,
        database=NOME_BANCO,
        timeout=30,
        command_timeout=300,
        server_settings={"application_name": "7lm_gc_prod_manual_import"},
    )
    try:
        async with conn.transaction():
            await _garantir_tabelas(conn)
            equipes = await _mapear_equipes(conn)
            historico = _extrair_historico(caminho, ano, fonte, equipes)
            total_repasses = await _copiar_repasses_importados(
                registros=repasses,
                conn=conn,
                fonte=fonte,
            )
            total_historico = await _copiar_registros(
                conn,
                tabela="dashboard_gc_produtividade_historico_corretor_equipe",
                colunas=[
                    "mes_referencia",
                    "equipe",
                    "equipe_key",
                    "corretor",
                    "corretor_key",
                    "gerente",
                    "coordenador",
                    "regiao",
                    "tipo",
                    "repasses_mes",
                    "ativo_no_mes",
                    "origem_planilha",
                ],
                registros=historico,
                campo_fonte="origem_planilha",
                fonte=fonte,
            )
        return {"repasses_importados": total_repasses, "historico_importado": total_historico}
    finally:
        await conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Importa historico manual da produtividade G&C.")
    parser.add_argument("--arquivo", required=True, help="Caminho do arquivo XLSX recebido dos gestores.")
    parser.add_argument("--ano", type=int, default=ANO_HISTORICO_PADRAO, help="Ano dos meses jan-jun da aba dinamica.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    caminho = Path(args.arquivo).expanduser()
    if not caminho.exists():
        raise SystemExit(f"Arquivo nao encontrado: {caminho}")
    resultado = asyncio.run(importar(caminho, args.ano))
    print(resultado)


if __name__ == "__main__":
    main()
